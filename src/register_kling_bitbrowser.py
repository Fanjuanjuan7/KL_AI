import datetime
import json
import os
import platform
import re
import socket
import threading
import time
import traceback
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from typing import Any, Callable, Dict, List, Optional, Tuple

import psutil
import requests
import socks
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# 设置全局 socket 超时，防止网络操作无限阻塞
default_socket_timeout = 30
socket.setdefaulttimeout(default_socket_timeout)

try:
    from src.captcha_receiver import MailExtractor
    from src.email_pool import EmailPool
except ImportError:
    from captcha_receiver import MailExtractor
    from email_pool import EmailPool


# =============================================================================
# CONSTANTS
# =============================================================================

# Time constants (in seconds)
DEFAULT_TIMEOUT_SEC = 200  # 降低超时到90秒，避免长时间阻塞
DEFAULT_POLL_INTERVAL_SEC = 0.5
PAGE_READY_TIMEOUT_SEC = 12
CONNECTIVITY_MAX_WAIT_SEC = 30  # 增加网络检测等待时间
SLIDER_MAX_WAIT_SEC = 60
CODE_EXTRACTION_TIMEOUT_SEC = 400
BROWSER_START_WAIT_SEC = 3  # 增加启动等待时间到8秒
POST_SLIDER_WAIT_SEC = 5
POST_SUBMIT_WAIT_SEC = 5  # 改为5秒，确保页面跳转完成

# Time constants (in milliseconds)
DEFAULT_TIMEOUT_MS = 100000
DEFAULT_POLL_MS = 500
ELEMENT_TIMEOUT_MS = 12000
SHORT_TIMEOUT_MS = 8000

# Retry constants
MAX_SLIDER_RETRIES = 8
MAX_REGISTRATION_ATTEMPTS = 3
MAX_CODE_RETRIES = 2
MAX_BROWSER_DELETE_RETRIES = 5
SAFE_CLICK_RETRIES = 2

# Resource limits
DEFAULT_MAX_CPU_PERCENT = 90   # 整机 CPU 使用率超过此值(%)才视为繁忙（原 30 过低，正常负载即误报）
DEFAULT_MAX_MEM_MB = 300        # 整机“可用内存”低于此值(MB)才视为繁忙（注意：是可用内存，非本进程内存）

# Slider pass cache timeout (seconds)
SLIDER_CACHE_TIMEOUT_SEC = 300

# Success URL keywords
SUCCESS_URL_KEYWORDS = ["dashboard", "all-tools", "home", "portal", "success"]

# XPath patterns
RESEND_CODE_XPATH_PATTERNS = [
    "//a[contains(text(), 'Resend Code') or contains(text(), '重新发送') or contains(text(), '再发一条')]",
    "//*[contains(text(), 'Resend')]",
    "//*[contains(@class, 'highlight')]",
]

# Error messages
ERROR_EMAIL_UNAVAILABLE = "Email unavailable"
ERROR_TARGET_REACHED = "target_reached"
ERROR_STOPPED = "stopped"
ERROR_PROXY_CONNECTIVITY_FAILED = "proxy_connectivity_failed"
ERROR_ENGLISH_OPTION_CLICK_FAILED = "english_option_click_failed"
ERROR_MORE_TOOLS_CLICK_FAILED = "more_tools_click_failed"
ERROR_SIGNIN_CLICK_FAILED = "signin_click_failed"
ERROR_SIGNIN_EMAIL_CLICK_FAILED = "signin_email_click_failed"
ERROR_SIGNUP_CLICK_FAILED = "signup_click_failed"
ERROR_EMAIL_INPUT_FAILED = "email_input_failed"
ERROR_PASSWORD_INPUT_FAILED = "password_input_failed"
ERROR_CONFIRM_INPUT_FAILED = "confirm_input_failed"
ERROR_NEXT_CLICK_FAILED = "next_click_failed"
ERROR_EMAIL_USED_PROMPT = "email_used_prompt"
ERROR_SLIDER_FAILED = "slider_failed"
ERROR_SLIDER_REAPPEARED = "slider_reappeared"
ERROR_MISSING_CREDENTIALS = "missing_credentials"
ERROR_CODE_NOT_FOUND = "code_not_found"
ERROR_CODE_INPUT_NOT_VISIBLE = "code_input_not_visible"
ERROR_CODE_INPUT_ERROR = "code_input_error"
ERROR_URL_JUMP_FAILED = "url_jump_failed"


# =============================================================================
# CLASSES
# =============================================================================


class WindowsDumpManager:
    """Context manager for creating crash dump files on Windows."""

    def __init__(self, logger: Optional[Callable[[str], None]], email: str) -> None:
        self.logger = logger
        self.email = email

    def __enter__(self) -> "WindowsDumpManager":
        return self

    def __exit__(
        self,
        exc_type: Optional[type],
        exc_val: Optional[BaseException],
        exc_tb: Optional[Any],
    ) -> None:
        if exc_type:
            try:
                # Create dump file
                dump_dir = os.path.join(os.getcwd(), "logs", "dumps")
                os.makedirs(dump_dir, exist_ok=True)
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"dump_{timestamp}_{self.email}.txt"
                filepath = os.path.join(dump_dir, filename)

                with open(filepath, "w", encoding="utf-8") as f:
                    f.write(f"Timestamp: {timestamp}\n")
                    f.write(f"Email: {self.email}\n")
                    f.write(f"Exception: {exc_type.__name__}: {exc_val}\n")
                    f.write("-" * 50 + "\n")
                    traceback.print_tb(exc_tb, file=f)
                    f.write("-" * 50 + "\n")
                    f.write(f"System: {platform.system()} {platform.release()}\n")

                if self.logger:
                    self.logger(f"已生成错误转储文件: {filepath}")
            except Exception as e:
                if self.logger:
                    self.logger(f"生成转储文件失败: {e}")


class StepRunner:
    """Executes steps with resource monitoring and timeout handling."""

    def __init__(
        self,
        logger: Optional[Callable[[str], None]],
        stop_event: Optional[threading.Event],
        max_cpu: int = DEFAULT_MAX_CPU_PERCENT,
        max_mem: int = DEFAULT_MAX_MEM_MB,
    ) -> None:
        self.logger = logger
        self.stop_event = stop_event
        self.max_cpu = max_cpu
        self.max_mem = max_mem

    def _check_resources(self) -> None:
        """检查整机资源，仅在机器确实快撑不住时才短暂停顿。

        历史坑：旧实现用 psutil.Process(os.getpid()).memory_info().rss（本 Python 进程
        自身内存）与 200MB 上限比较。带 GUI + 多个浏览器时本进程内存天然 > 200MB，
        于是几乎每次都误报“资源繁忙”并白白等待，看起来像卡死/报错。
        现改为测量【整机可用内存】，阈值也放宽到合理范围，正常负载下不会再误报。
        """
        try:
            for _ in range(3):
                cpu = psutil.cpu_percent(interval=0.1)

                # 整机可用内存（MB），而非本进程 RSS
                mem_avail = psutil.virtual_memory().available / 1024 / 1024

                # 可用内存充足 且 CPU 未打满 → 直接放行，不等待
                if mem_avail > self.max_mem and cpu < self.max_cpu:
                    return

                if self.logger:
                    self.logger(
                        f"资源繁忙 (CPU={cpu:.1f}%, 可用内存={mem_avail:.1f}MB)，等待释放..."
                    )
                time.sleep(1)
        except (psutil.Error, OSError):
            pass

    def run(
        self, name: str, func: Callable[[], Any], timeout: int = DEFAULT_TIMEOUT_SEC
    ) -> Any:
        """Execute a function with timeout and resource checking."""
        if self.stop_event and self.stop_event.is_set():
            raise RuntimeError(ERROR_STOPPED)

        self._check_resources()

        if self.logger:
            self.logger(f"--- 阶段开始: {name} ---")

        t0 = time.time()
        try:
            # Avoid using 'with' context manager which enforces wait=True on exit
            executor = ThreadPoolExecutor(max_workers=1)
            try:
                future = executor.submit(func)
                res = future.result(timeout=timeout)
                executor.shutdown(wait=True)
                return res
            except TimeoutError:
                if self.logger:
                    self.logger(f"阶段 {name} 超时 ({timeout}s)")
                # Force shutdown without waiting for the stuck thread
                executor.shutdown(wait=False)
                raise RuntimeError(f"Step {name} timed out")
            except Exception:
                executor.shutdown(wait=False)
                raise
        except TimeoutError:
            # Re-raise nicely formatted error
            raise RuntimeError(f"Step {name} timed out")
        except Exception as e:
            if self.logger:
                self.logger(f"阶段 {name} 异常: {e}")
            raise e
        finally:
            dur = time.time() - t0
            if self.logger:
                self.logger(f"--- 阶段结束: {name} (耗时 {dur:.2f}s) ---")
            time.sleep(0.1)


@dataclass
class RegistrationEvents:
    """Event callbacks for registration process."""

    on_success: Optional[Callable[[str], None]] = None  # email
    on_failure: Optional[Callable[[str, str], None]] = None  # email, reason
    on_log: Optional[Callable[[str], None]] = None
    on_finish: Optional[Callable[[str, bool, str], None]] = (
        None  # email, success, message
    )


# =============================================================================
# PROCESS UTILITIES
# =============================================================================


def get_pid_by_port(port: int) -> Optional[int]:
    """Find PID of the process listening on a specific TCP port."""
    try:
        # psutil.net_connections requires root/admin on some OS/versions for all PIDs
        # But for user-owned processes it often works.
        for conn in psutil.net_connections(kind="inet"):
            if conn.laddr.port == port and conn.status == psutil.CONN_LISTEN:
                return conn.pid
    except (psutil.AccessDenied, psutil.NoSuchProcess):
        pass
    return None


def kill_process_tree(pid: int, logger: Optional[Callable[[str], None]] = None) -> None:
    """Kill a process and all its children forcibly."""
    try:
        if not psutil.pid_exists(pid):
            return
        parent = psutil.Process(pid)
        children = parent.children(recursive=True)
        for child in children:
            try:
                child.kill()
            except psutil.NoSuchProcess:
                pass
        parent.kill()
        try:
            parent.wait(3)
        except psutil.TimeoutExpired:
            pass
        if logger:
            logger(f"已强制结束进程 PID={pid}")
    except psutil.NoSuchProcess:
        pass
    except Exception as e:
        if logger:
            logger(f"结束进程 PID={pid} 失败: {e}")


def kill_all_bitbrowsers(logger: Optional[Callable[[str], None]] = None) -> None:
    """Force kill all BitBrowser processes."""
    target_names = ["BitBrowser.exe", "BitBrowser"]
    killed_count = 0
    for proc in psutil.process_iter(["pid", "name"]):
        try:
            if proc.info["name"] in target_names:
                kill_process_tree(proc.info["pid"], logger)
                killed_count += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass
    if logger and killed_count > 0:
        logger(f"已清理 {killed_count} 个残留 BitBrowser 进程")


class BitBrowserClient:
    """Client for interacting with BitBrowser API."""

    def __init__(self, base_url: str, secret: Optional[str] = None) -> None:
        self.base_url = base_url.rstrip("/")
        self.secret = secret
        self.session = requests.Session()

        # ============================================================
        # BitBrowser API 全局限流
        #
        # 注册任务可以 8 并发，但 BitBrowser 本地 API 不允许 8 个线程
        # 同时 create/open/close/delete。
        #
        # 最多允许 2 个 API 请求同时执行。
        # 并且两个 API 请求的启动时间至少间隔 0.6 秒。
        # ============================================================
        self._api_semaphore = threading.BoundedSemaphore(2)
        self._api_rate_lock = threading.Lock()
        self._last_api_request_time = 0.0
        self._api_min_interval = 0.6

        # HTTP连接池不用开太大，真正的并发由上面的 semaphore 控制
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=10,
            pool_maxsize=10,
            max_retries=0,
        )
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)

    def _headers(self) -> Dict[str, str]:
        h = {"Content-Type": "application/json", "Accept": "application/json"}
        if self.secret:
            h["Authorization"] = self.secret
        return h

    def _request(
        self,
        method: str,
        endpoint: str,
        data: Optional[Dict[str, Any]] = None,
        timeout: int = 60,
        max_retries: int = 5,
    ) -> requests.Response:
        """
        BitBrowser API统一请求入口。

        功能：
        1. 最多2个BitBrowser API请求同时执行
        2. API请求启动间隔至少0.6秒
        3. HTTP 429自动退避重试
        4. 连接失败/超时自动重试
        5. 不影响Selenium/邮箱接码等其他并发
        """

        url = f"{self.base_url}{endpoint}"
        last_exception = None

        max_retries = max(1, int(max_retries))

        for attempt in range(1, max_retries + 1):

            # ========================================================
            # 429重试的每一轮，都重新进入API限流队列
            # ========================================================
            with self._api_semaphore:

                # ----------------------------------------------------
                # 控制请求启动频率
                # 即使8个线程同时来到这里，也会依次间隔0.6秒发送
                # ----------------------------------------------------
                with self._api_rate_lock:
                    now = time.monotonic()

                    elapsed_since_last = now - self._last_api_request_time
                    wait_for_rate = self._api_min_interval - elapsed_since_last

                    if wait_for_rate > 0:
                        time.sleep(wait_for_rate)

                    self._last_api_request_time = time.monotonic()

                try:
                    t0 = time.time()

                    if method.upper() == "POST":
                        r = self.session.post(
                            url,
                            headers=self._headers(),
                            data=json.dumps(data) if data is not None else None,
                            timeout=timeout,
                        )
                    else:
                        r = self.session.get(
                            url,
                            headers=self._headers(),
                            timeout=timeout,
                        )

                    elapsed = time.time() - t0

                    if elapsed > 1.0:
                        print(
                            f"Slow BitBrowser API: "
                            f"{method} {endpoint} took {elapsed:.2f}s"
                        )

                    # =================================================
                    # HTTP 429：不要直接杀掉注册任务
                    # =================================================
                    if r.status_code == 429:
                        last_exception = requests.exceptions.HTTPError(
                            f"429 Too Many Requests: {method} {url}",
                            response=r,
                        )

                        if attempt >= max_retries:
                            print(
                                f"BitBrowser API 429，"
                                f"已达到最大重试次数 "
                                f"{attempt}/{max_retries}: {endpoint}"
                            )
                            raise last_exception

                        # 优先读取服务器 Retry-After
                        retry_after = None
                        try:
                            retry_after_value = r.headers.get("Retry-After")
                            if retry_after_value:
                                retry_after = float(retry_after_value)
                        except Exception:
                            retry_after = None

                        # 没有 Retry-After 时：
                        # 第1次 2秒
                        # 第2次 4秒
                        # 第3次 6秒
                        # 第4次 8秒
                        wait_time = (
                            retry_after
                            if retry_after is not None
                            else min(attempt * 2, 10)
                        )

                        print(
                            f"⚠️ BitBrowser API触发429: {endpoint} | "
                            f"尝试 {attempt}/{max_retries}，"
                            f"{wait_time:.1f}秒后重试..."
                        )

                    else:
                        # 其他 HTTP 错误正常抛出
                        if r.status_code == 405:
                            print(
                                f"HTTP 405 Method Not Allowed: "
                                f"{method} {url}"
                            )

                        r.raise_for_status()
                        return r

                except requests.exceptions.HTTPError as e:
                    # 429已经在上面处理
                    if (
                        getattr(e, "response", None) is not None
                        and e.response.status_code == 429
                    ):
                        last_exception = e

                        if attempt >= max_retries:
                            raise

                        # wait_time 会在 semaphore 外面计算
                        wait_time = min(attempt * 2, 10)

                    else:
                        # 400 / 401 / 403 / 404 / 405 等错误
                        # 不应该反复轰API
                        if (
                            getattr(e, "response", None) is not None
                            and e.response.status_code == 405
                        ):
                            print(
                                f"HTTP 405 Error Details: "
                                f"{method} {url} - {e.response.text}"
                            )

                        raise

                except (
                    requests.exceptions.ConnectionError,
                    requests.exceptions.Timeout,
                ) as e:
                    last_exception = e

                    if attempt >= max_retries:
                        raise

                    wait_time = min(2 ** (attempt - 1), 8)

                    print(
                        f"BitBrowser API连接失败 "
                        f"{attempt}/{max_retries}: "
                        f"{method} {endpoint} | {e}"
                    )

                except requests.exceptions.RequestException:
                    raise

            # ========================================================
            # 非常重要：
            # 睡眠放在 semaphore 外面。
            #
            # 否则一个429线程睡8秒，会一直霸占API名额。
            # ========================================================
            if attempt < max_retries:
                print(
                    f"等待 {wait_time:.1f} 秒后重试 "
                    f"BitBrowser API: {endpoint}"
                )
                time.sleep(wait_time)

        if last_exception:
            raise last_exception

        raise RuntimeError(f"BitBrowser API请求失败: {url}")

    def update_browser(
        self,
        name: str,
        proxy: Optional[Dict[str, Any]] = None,
        enable_udp: bool = False,
        cmd_args: Optional[List[str]] = None,
        save_traffic: bool = False,  # <--- 必须加上这一行
    ) -> str:
        payload: Dict[str, Any] = {
            "name": name,
            "remark": "",
            "proxyMethod": 2,
            "proxyType": "noproxy",
            "isUDP": 1 if enable_udp else 0,
            "syncTabs": True,
            "syncCookies": True,
            "syncLocalStorage": True,
            "syncIndexedDb": True,
            "browserFingerPrint": {"coreVersion": "124"},
        }

        # --- 新增：对接比特浏览器原生省流量 API ---
        if save_traffic:
            payload.update(
                {
                    "abortImage": True,  # 开启图片拦截
                    "abortImageMaxSize": 100,  # ⚠️ 关键：允许加载 100KB 以下的图片（确保滑块正常）
                    "abortAudio": True,  # 禁用声音
                    "abortTranslate": True,  # 禁用谷歌翻译弹窗
                    "abortCertificate": True,  # 禁用保存密码弹窗
                }
            )
        # ---------------------------------------

        if cmd_args:
            payload["cmdArgs"] = cmd_args
        if proxy:
            payload.update(proxy)
        r = self._request(
            "POST",
            "/browser/update",
            payload,
            timeout=30,
            max_retries=5,
        )
        data = r.json()
        d = data.get("data")
        bid = None
        if isinstance(d, dict):
            bid = d.get("id")
        elif isinstance(d, str):
            bid = d
        if not bid:
            bid = data.get("id")
        if not bid:
            raise RuntimeError(f"update_browser failed: {data}")
        return bid

    def create_browser(
        self,
        name: str,
        proxy: Optional[Dict[str, Any]] = None,
        enable_udp: bool = False,
        cmd_args: Optional[List[str]] = None,
        save_traffic: bool = False,  # <--- 必须加上这一行
    ) -> str:
        payload: Dict[str, Any] = {
            "name": name,
            "remark": "",
            "proxyMethod": 2,
            "proxyType": "noproxy",
            "isUDP": 1 if enable_udp else 0,
            "syncTabs": True,
            "syncCookies": True,
            "syncLocalStorage": True,
            "syncIndexedDb": True,
            "browserFingerPrint": {"coreVersion": "124"},
        }

        # --- 新增：对接比特浏览器原生省流量 API ---
        if save_traffic:
            payload.update(
                {
                    "abortImage": True,
                    "abortImageMaxSize": 100,  # 允许小图（滑块），拦截大图
                    "abortAudio": True,
                    "abortTranslate": True,
                    "abortCertificate": True,
                }
            )

        if cmd_args:
            payload["cmdArgs"] = cmd_args
        if proxy:
            payload.update(proxy)
        try:
            r = self._request(
                "POST",
                "/browser/update",
                payload,
                timeout=30,
                max_retries=5,
            )
        except requests.exceptions.HTTPError as e:
            # 404 = 本地API没有「创建窗口」(/browser/update) 这个路由，
            # 几乎都是比特浏览器客户端/账号侧的问题，翻译成人话，避免误以为是程序bug
            resp = getattr(e, "response", None)
            if resp is not None and resp.status_code == 404:
                raise RuntimeError(
                    "比特浏览器「创建窗口」接口(/browser/update)返回404：本地API不可用。"
                    "请排查：①比特浏览器客户端是否已启动并登录账号；"
                    "②设置→高级→「本地API接口」是否开启；"
                    "③客户端版本是否完整（可重启或重装客户端后重试）。"
                ) from e
            raise
        data = r.json()
        d = data.get("data")
        bid = None
        if isinstance(d, dict):
            bid = d.get("id")
        elif isinstance(d, str):
            bid = d
        if not bid:
            bid = data.get("id")
        if not bid:
            raise RuntimeError(f"create_browser failed: {data}")
        return bid

    def open_browser(
        self,
        browser_id: str,
        headless_mode: bool = False,
    ) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "id": browser_id,
        }

        # ============================================================
        # BitBrowser 官方真正无头模式
        # ============================================================
        if headless_mode:
            payload.update(
                {
                    "args": ["--headless"],
                    "queue": True,
                    "ignoreDefaultUrls": True,
                }
            )

        # 保留之前增加的 API 限流 / 429 自动重试
        r = self._request(
            "POST",
            "/browser/open",
            payload,
            timeout=60,
            max_retries=5,
        )

        data = r.json()
        d = data.get("data")

        if isinstance(d, dict):
            return d

        raise RuntimeError(
            f"open_browser 返回异常: browser_id={browser_id}, response={data}"
        )

    def close_browser(self, browser_id: str) -> None:
        """
        关闭浏览器窗口。
        失败必须向上抛异常，不能静默 pass。
        """
        if not browser_id:
            return

        r = self._request(
            "POST",
            "/browser/close",
            {"id": browser_id},
            timeout=15,
            max_retries=1,
        )

        # HTTP 200 但 API 业务层失败时，也视为失败
        try:
            data = r.json()
            if isinstance(data, dict) and data.get("success") is False:
                raise RuntimeError(f"close_browser failed: {data}")
        except ValueError:
            # 有些版本接口成功时可能没有 JSON，HTTP 2xx 就算成功
            pass

    def delete_browser(self, browser_id: str) -> None:
        """
        删除浏览器 Profile。
        删除失败必须抛异常，让外层真正执行重试。
        """
        if not browser_id:
            return

        r = self._request(
            "POST",
            "/browser/delete",
            {"id": browser_id},
            timeout=30,
            max_retries=1,
        )

        # HTTP 200 但 API 返回业务失败，也不能假装删除成功
        try:
            data = r.json()
            if isinstance(data, dict) and data.get("success") is False:
                raise RuntimeError(f"delete_browser failed: {data}")
        except ValueError:
            pass

    def detail_browser(self, browser_id: str) -> Dict[str, Any]:
        try:
            r = self._request("POST", "/browser/detail", {"id": browser_id}, timeout=30)
            data = r.json()
            d = data.get("data")
            if isinstance(d, dict):
                return d
            return {}
        except (requests.exceptions.RequestException, RuntimeError):
            return {}


# =============================================================================
# FILE UTILITIES
# =============================================================================


def read_rows(input_path: str) -> List[Dict[str, Any]]:
    """Read rows from various file formats (CSV, JSON, XLSX)."""
    # Check for "----" format (New EmailPool format)
    try:
        with open(input_path, "r", encoding="utf-8") as f:
            first_line = f.readline()
            if "----" in first_line:
                try:
                    from src.email_pool import EmailPool

                    pool = EmailPool(input_path)
                    return pool.get_all_rows()
                except ImportError:
                    pass
    except (IOError, UnicodeDecodeError):
        pass

    if input_path.lower().endswith(".csv"):
        import csv

        rows: List[Dict[str, Any]] = []
        with open(input_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        return rows
    if input_path.lower().endswith(".json"):
        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            if isinstance(data, dict) and "rows" in data:
                return data["rows"]
        return []
    if input_path.lower().endswith(".xlsx"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(input_path)
            sheet = wb.active
            headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
            rows = []
            for r in sheet.iter_rows(min_row=2):
                row = {
                    headers[i]: (r[i].value if i < len(r) else None)
                    for i in range(len(headers))
                }
                rows.append(row)
            return rows
        except Exception:
            return []
    return []


def write_rows_csv(input_path: str, rows: List[Dict[str, Any]]) -> None:
    """Write rows to CSV file."""
    import csv

    if not rows:
        return
    # Fix: Collect all unique keys from all rows to ensure fieldnames cover everything
    headers = set()
    for r in rows:
        headers.update(r.keys())

    # Sort headers for consistency, prioritize common fields
    header_list = sorted(list(headers))
    priority_fields = ["email", "password", "status", "host", "port", "msg"]
    for f in reversed(priority_fields):
        if f in header_list:
            header_list.remove(f)
            header_list.insert(0, f)

    with open(input_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header_list)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


# =============================================================================
# WEBDRIVER UTILITIES
# =============================================================================


def element_exists(
    driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int
) -> bool:
    """Check if an element exists in the DOM."""
    try:
        eff_timeout = min(timeout_ms, 60000)
        eff_poll = max(0.2, poll_ms / 1000.0)
        WebDriverWait(driver, eff_timeout / 1000.0, poll_frequency=eff_poll).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        return True
    except Exception:
        return False


def element_visible(
    driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int
) -> bool:
    """Check if an element is visible on the page."""
    try:
        eff_timeout = min(timeout_ms, 60000)
        eff_poll = max(0.2, poll_ms / 1000.0)
        WebDriverWait(driver, eff_timeout / 1000.0, poll_frequency=eff_poll).until(
            EC.visibility_of_element_located((By.XPATH, xpath))
        )
        return True
    except Exception:
        return False


def element_exists_visible(
    driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int
) -> bool:
    """Check if an element exists and is visible."""
    if not element_exists(driver, xpath, timeout_ms, poll_ms):
        return False
    return element_visible(driver, xpath, timeout_ms, poll_ms)


def wait_page_ready(
    driver: webdriver.Remote, timeout_sec: int = PAGE_READY_TIMEOUT_SEC
) -> bool:
    """Wait for page to reach interactive or complete ready state."""
    try:
        WebDriverWait(driver, timeout_sec).until(
            lambda d: (
                d.execute_script("return document.readyState")
                in ("interactive", "complete")
            )
        )
        return True
    except Exception:
        return False


def log_page_timing(
    driver: webdriver.Remote, logger: Optional[Callable[[str], None]]
) -> None:
    """Log page timing metrics."""
    if not logger:
        return
    try:
        data = driver.execute_script("""
            try {
              var t = performance.timing || {};
              var navs = performance.getEntriesByType && performance.getEntriesByType('navigation');
              var nav = (navs && navs.length) ? navs[0] : null;
              var ns = nav ? nav.startTime : (t.navigationStart || 0);
              var dcl = nav ? nav.domContentLoadedEventEnd : (t.domContentLoadedEventEnd || 0);
              var le = nav ? nav.loadEventEnd : (t.loadEventEnd || 0);
              return {
                readyState: document.readyState,
                navigationStart: ns,
                domContentLoadedEventEnd: dcl,
                loadEventEnd: le
              };
            } catch (e) {
              return { readyState: document.readyState };
            }
        """)
        ns = float(data.get("navigationStart") or 0.0)
        dcl = float(data.get("domContentLoadedEventEnd") or 0.0)
        le = float(data.get("loadEventEnd") or 0.0)
        rs = data.get("readyState")

        def fmt(ms: float) -> str:
            return f"{int(ms)}ms" if ms > 0 else "n/a"

        logger(
            f"PageTiming: readyState={rs}, domContentLoaded={fmt(dcl - ns)}, loadComplete={fmt(le - ns)}"
        )
    except Exception:
        pass


def log_resource_status(
    driver: webdriver.Remote, logger: Optional[Callable[[str], None]], limit: int = 200
) -> None:
    """Log resource loading status."""
    if not logger:
        return
    try:
        data = driver.execute_script("""
            try {
              var es = performance.getEntries ? performance.getEntries() : [];
              var out = {};
              for (var i=0;i<es.length;i++){
                var e = es[i];
                var t = e.entryType || e.initiatorType || 'unknown';
                out[t] = (out[t]||0) + 1;
              }
              return {counts: out, total: es.length};
            } catch (e) { return {error: ''+e}; }
        """)
        counts = data.get("counts") or {}
        total = data.get("total") or 0
        try:
            logger(
                f"Resources: total={total}, counts={json.dumps(counts, ensure_ascii=False)}"
            )
        except Exception:
            logger(f"Resources: total={total}")
    except Exception:
        pass


def log_memory_usage(logger: Optional[Callable[[str], None]] = None) -> None:
    """Log current process memory usage."""
    if not logger:
        return
    try:
        process = psutil.Process(os.getpid())
        mem_info = process.memory_info()
        rss_mb = mem_info.rss / 1024 / 1024
        logger(f"当前进程内存占用: {rss_mb:.2f} MB")
    except (psutil.Error, OSError):
        pass


def log_resource_phase_timings(
    driver: webdriver.Remote,
    logger: Optional[Callable[[str], None]],
    contains: Optional[str] = None,
    limit: int = 30,
) -> None:
    """Log detailed resource timing information."""
    if not logger:
        return
    try:
        data = driver.execute_script(
            """
            try {
              var es = performance.getEntriesByType ? performance.getEntriesByType('resource') : [];
              var out = [];
              var kw = arguments[0] || '';
              var lim = arguments[1] || 30;
              for (var i=0;i<es.length;i++){
                var e = es[i];
                if (kw && e.name && e.name.indexOf(kw) === -1) continue;
                out.push({
                  name: e.name,
                  initiatorType: e.initiatorType,
                  duration: e.duration,
                  dns: (e.domainLookupEnd && e.domainLookupStart) ? (e.domainLookupEnd - e.domainLookupStart) : 0,
                  tcp: (e.connectEnd && e.connectStart) ? (e.connectEnd - e.connectStart) : 0,
                  ssl: (e.connectEnd && e.secureConnectionStart) ? (e.connectEnd - e.secureConnectionStart) : 0,
                  ttfb: (e.responseStart && e.requestStart) ? (e.responseStart - e.requestStart) : 0,
                  download: (e.responseEnd && e.responseStart) ? (e.responseEnd - e.responseStart) : 0
                });
              }
              if (out.length > lim) out = out.slice(out.length - lim);
              return out;
            } catch (e) { return []; }
        """,
            contains or "",
            limit,
        )
        if not isinstance(data, list) or not data:
            return
        key = contains or "all"
        logger(f"ResourceTiming[{key}] count={len(data)}")
        for e in data:
            try:
                name = str(e.get("name") or "")
                it = str(e.get("initiatorType") or "")
                dur = float(e.get("duration") or 0.0)
                dns = float(e.get("dns") or 0.0)
                tcp = float(e.get("tcp") or 0.0)
                ssl = float(e.get("ssl") or 0.0)
                ttfb = float(e.get("ttfb") or 0.0)
                dl = float(e.get("download") or 0.0)
                logger(
                    f"Timing {it} {int(dur)}ms dns={int(dns)} tcp={int(tcp)} ssl={int(ssl)} ttfb={int(ttfb)} dl={int(dl)} url={name}"
                )
            except (ValueError, TypeError):
                pass
    except Exception:
        pass


def ensure_artifact_dir() -> str:
    """Ensure the test artifacts directory exists."""
    base = os.path.dirname(__file__)
    p = os.path.join(base, "test_artifacts")
    try:
        os.makedirs(p, exist_ok=True)
    except (OSError, PermissionError):
        pass
    return p


def take_screenshot(
    driver: webdriver.Remote, name: str, logger: Optional[Callable[[str], None]] = None
) -> None:
    """Take a screenshot and save it to the artifacts directory."""
    d = ensure_artifact_dir()
    path = os.path.join(d, name)
    try:
        driver.save_screenshot(path)
        if logger:
            logger(f"Screenshot: {path}")
    except Exception as e:
        if logger:
            logger(f"Screenshot error: {e}")


def write_artifact_json(
    name: str, data: Any, logger: Optional[Callable[[str], None]] = None
) -> None:
    """Write data as JSON to the artifacts directory."""
    d = ensure_artifact_dir()
    path = os.path.join(d, name)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        if logger:
            logger(f"Artifact: {path}")
    except Exception as e:
        if logger:
            logger(f"Artifact error: {e}")


def check_connectivity(
    driver: webdriver.Remote,
    url: str,
    logger: Optional[Callable[[str], None]],
    max_wait: int = CONNECTIVITY_MAX_WAIT_SEC,
    udp_enabled: bool = False,
) -> bool:
    """Check network connectivity by navigating to a URL."""
    ok = True
    try:
        # 1. 强制页面加载超时时间，防止 driver.get 自身卡死
        try:
            driver.set_page_load_timeout(15)
        except Exception:
            pass

        try:
            driver.get(url)
        except Exception:
            driver.execute_script("window.location.href=arguments[0];", url)

        # ==========================================
        # 🚀 核心优化：代理失效极速熔断 (Fast-Fail)
        # ==========================================
        time.sleep(1)  # 给浏览器 1 秒钟渲染错误页面
        try:
            # 抓取页面源码（全部转大写方便匹配）
            page_source = driver.page_source.upper()

            # 定义 Chrome 常见的代理/网络致命错误代码
            fatal_errors = [
                "ERR_SOCKS_CONNECTION_FAILED",
                "ERR_PROXY_CONNECTION_FAILED",
                "ERR_CONNECTION_TIMED_OUT",
                "ERR_CONNECTION_CLOSED",
                "ERR_CONNECTION_RESET",
                "ERR_TUNNEL_CONNECTION_FAILED",
            ]

            for err in fatal_errors:
                if err in page_source:
                    if logger:
                        logger(
                            f"🚫 极速熔断: 检测到致命网络错误 [{err}]，代理失效，立即放弃！"
                        )
                    return False  # 瞬间返回 False，终止当前任务
        except Exception:
            pass
        # ==========================================

        wait_page_ready(driver, max_wait)
        log_page_timing(driver, logger)
        log_resource_status(driver, logger)
        try:
            init_handles = driver.window_handles
            if logger:
                logger(f"初始窗口数量: {len(init_handles)}")
            if len(init_handles) > 1 and (not udp_enabled):
                try:
                    driver.switch_to.window(init_handles[0])
                    if logger:
                        logger("检测到初始阶段存在额外标签，已切回主标签")
                except Exception:
                    pass
        except Exception:
            pass
        take_screenshot(driver, "bitbrowser_initial.png", logger)
    except Exception as e:
        ok = False
        if logger:
            logger(f"Connectivity nav error: {e}")
    return ok


def xpath_count(driver: webdriver.Remote, xpath: str) -> int:
    """Count elements matching an XPath."""
    try:
        return int(
            driver.execute_script(
                """
            try {
              var r = document.evaluate(arguments[0], document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
              return r ? r.snapshotLength : 0;
            } catch (e) { return -1; }
        """,
                xpath,
            )
        )
    except Exception:
        return -1


def safe_click(
    driver: webdriver.Remote,
    xpath: str,
    timeout_ms: int,
    poll_ms: int,
    logger: Optional[Callable[[str], None]] = None,
    retries: int = SAFE_CLICK_RETRIES,
) -> bool:
    """Safely click an element with multiple fallback strategies."""
    for i in range(max(1, retries)):
        try:
            WebDriverWait(
                driver,
                timeout_ms / 1000.0,
                poll_frequency=max(0.2, poll_ms / 1000.0),
            ).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
            return True
        except Exception:
            try:
                el = WebDriverWait(
                    driver,
                    timeout_ms / 1000.0,
                    poll_frequency=max(0.2, poll_ms / 1000.0),
                ).until(EC.presence_of_element_located((By.XPATH, xpath)))
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", el
                )
                try:
                    el.click()
                    return True
                except Exception:
                    driver.execute_script("arguments[0].click();", el)
                    return True
            except Exception as e:
                if logger:
                    logger(f"点击失败重试 {i + 1}/{retries}: {e}")
        time.sleep(0.5)
    return False


def first_present_xpath(
    driver: webdriver.Remote, xpaths: List[Optional[str]], timeout_ms: int, poll_ms: int
) -> Optional[str]:
    """并行高频探测：找出列表中第一个出现的可见 XPath，不再串行死等。"""
    valid_xpaths = [xp for xp in xpaths if xp]
    if not valid_xpaths:
        return None

    # 彻底去掉 12 秒的硬限制，完全尊重 UI 面板传进来的 timeout_ms（比如 30000 或 60000）
    eff_timeout = timeout_ms / 1000.0
    eff_poll = max(0.2, poll_ms / 1000.0)
    end_time = time.time() + eff_timeout

    while time.time() < end_time:
        for xp in valid_xpaths:
            try:
                # 瞬间探测，如果不存在立刻抛异常进入下一个，绝对不阻塞
                elements = driver.find_elements(By.XPATH, xp)
                for el in elements:
                    if el.is_displayed():
                        return xp
            except Exception:
                continue
        time.sleep(eff_poll)

    return None


def safe_click_any(
    driver: webdriver.Remote,
    xpaths: List[Optional[str]],
    timeout_ms: int,
    poll_ms: int,
    logger: Optional[Callable[[str], None]] = None,
    retries: int = SAFE_CLICK_RETRIES,
) -> bool:
    """Safely click the first visible element from a list of XPaths."""
    xp = first_present_xpath(driver, xpaths, timeout_ms, poll_ms)
    if not xp:
        return False
    return safe_click(driver, xp, timeout_ms, poll_ms, logger, retries)


def js_click_xpath(driver: webdriver.Remote, xpath: str) -> bool:
    """Click an element using JavaScript execution."""
    try:
        r = driver.execute_script(
            """
          var xp = arguments[0];
          try {
            var res = document.evaluate(xp, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            var el = res && res.singleNodeValue;
            if (el) {
              try { el.scrollIntoView({block:'center'}); } catch(e) {}
              try { el.click(); return true; } catch(e) {}
              try { el.dispatchEvent(new MouseEvent('click', {bubbles:true})); return true; } catch(e) {}
            }
          } catch (e) {}
          return false;
        """,
            xpath,
        )
        return bool(r)
    except Exception:
        return False


def safe_send_keys(
    driver: webdriver.Remote,
    xpath: str,
    text: str,
    timeout_ms: int,
    poll_ms: int,
    logger: Optional[Callable[[str], None]] = None,
    retries: int = 1,
) -> bool:
    """Safely send keys to an input element."""
    for i in range(max(1, retries)):
        try:
            el = WebDriverWait(
                driver,
                timeout_ms / 1000.0,
                poll_frequency=max(0.2, poll_ms / 1000.0),
            ).until(EC.presence_of_element_located((By.XPATH, xpath)))
            el.send_keys(text)
            return True
        except Exception as e:
            if logger:
                logger(f"输入失败重试 {i + 1}/{retries}: {e}")
        time.sleep(0.5)
    return False


def safe_send_keys_any(
    driver: webdriver.Remote,
    xpaths: List[Optional[str]],
    text: str,
    timeout_ms: int,
    poll_ms: int,
    logger: Optional[Callable[[str], None]] = None,
    retries: int = 1,
) -> bool:
    """智能并行输入：先用 first_present_xpath 极速定位，再安全输入。"""
    xp = first_present_xpath(driver, xpaths, timeout_ms, poll_ms)
    if not xp:
        return False
    return safe_send_keys(driver, xp, text, timeout_ms, poll_ms, logger, retries)


def validate_xpaths(
    driver: webdriver.Remote,
    xpaths: Dict[str, str],
    logger: Optional[Callable[[str], None]],
    keys: List[str],
    timeout_ms: int,
    poll_ms: int,
) -> Dict[str, bool]:
    """Validate multiple XPath locators."""
    result: Dict[str, bool] = {}
    for k in keys:
        xp = xpaths.get(k) or ""
        ok = bool(xp) and element_exists_visible(
            driver, xp, min(timeout_ms, 8000), poll_ms
        )
        result[k] = ok
        if logger:
            logger(f"XPath验证 {k}: {'OK' if ok else 'FAIL'}")
    return result


# =============================================================================
# SLIDER UTILITIES
# =============================================================================


def _human_drag_track(
    distance: int, duration_ms: int = 900, jitter_px: int = 1, overshoot_px: int = 5
) -> List[int]:
    """Generate a human-like drag track for slider solving."""
    steps = 15
    arr: List[int] = []
    moved = 0
    for i in range(steps):
        remain = distance - moved
        if remain <= 0:
            break
        prog = i / steps
        base = int((1 - (prog - 0.5) ** 2) * 8) + 1
        jitter = ((i % 3) - 1) * jitter_px
        dx = max(1, base + jitter)
        moved += dx
        arr.append(dx)
    while sum(arr) < distance:
        arr.append(1)
    arr.append(overshoot_px)
    arr.append(-overshoot_px + 1)
    return arr


_SLIDER_PASS_CACHE: Dict[str, Dict[str, Any]] = {}


def solve_slider(
    driver: webdriver.Remote,
    xpaths: Dict[str, str],
    timeout_ms: int,
    poll_ms: int,
    logger: Optional[Callable[[str], None]] = None,
    stop_event: Optional[threading.Event] = None,
) -> bool:
    """
    Attempt to solve a slider captcha.

    Args:
        driver: Selenium WebDriver instance
        xpaths: Dictionary containing slider element XPaths
        timeout_ms: Maximum time to wait for slider in milliseconds
        poll_ms: Polling interval in milliseconds
        logger: Optional logging function

    Returns:
        True if slider was solved successfully, False otherwise
    """
    t0 = time.time()
    ok_xpath = (
        xpaths.get("code_url_element")
        or xpaths.get("next_btn")
        or xpaths.get("password_input")
    )

    extra_ok_locators = [
        "//input[@autocomplete='one-time-code']",
        "//input[contains(@placeholder, 'verification') or contains(@placeholder, 'code') or contains(@placeholder, '验证码')]",
        "//input[@type='text' and string-length(@maxlength)='6']",
    ]

    def _slider_container_visible(short_timeout_ms: int = 600) -> bool:
        try:
            try:
                driver.switch_to.default_content()
            except Exception:
                pass
            if element_visible(
                driver, xpaths["slider_container"], short_timeout_ms, poll_ms
            ):
                return True
        except Exception:
            pass
        try:
            try:
                driver.switch_to.default_content()
            except Exception:
                pass
            if element_exists(
                driver, xpaths["slider_iframe"], short_timeout_ms, poll_ms
            ):
                try:
                    iframe = driver.find_element(By.XPATH, xpaths["slider_iframe"])
                    driver.switch_to.frame(iframe)
                except Exception:
                    return False
                try:
                    return element_visible(
                        driver, xpaths["slider_container"], short_timeout_ms, poll_ms
                    )
                finally:
                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass
        except Exception:
            pass
        return False

    def _ok_visible(short_timeout_ms: int) -> bool:
        # Check primary xpath
        if ok_xpath:
            try:
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                if element_visible(driver, ok_xpath, short_timeout_ms, poll_ms):
                    return True
            except Exception:
                pass

        # Check extra locators (only if we are looking for code input)
        # We assume if 'code_url_element' is in xpaths, we are at that stage
        if xpaths.get("code_url_element"):
            for loc in extra_ok_locators:
                try:
                    if element_visible(driver, loc, 100, poll_ms):
                        return True
                except Exception:
                    pass

        return False

    try:
        if _ok_visible(min(1500, timeout_ms)):
            if logger:
                logger("Slider: 已处于后续步骤，跳过滑块等待")
            return True
    except Exception:
        pass

    cache_key = ""
    domain_filter = None
    try:
        sid = str(getattr(driver, "session_id", "") or "")
        cur = ""
        try:
            cur = driver.current_url or ""
        except Exception:
            cur = ""
        host = ""
        try:
            host = (urllib.parse.urlparse(cur).netloc or "").strip()
        except Exception:
            host = ""
        domain_filter = host or None
        cache_key = f"{sid}|{host}"
    except Exception:
        cache_key = ""

    try:
        cached = _SLIDER_PASS_CACHE.get(cache_key) if cache_key else None
        if cached and (
            time.time() - float(cached.get("t") or 0.0) < SLIDER_CACHE_TIMEOUT_SEC
        ):
            if not _slider_container_visible(600):
                if logger:
                    logger("Slider: 命中通过缓存，且当前未检测到滑块")
                if cached.get("ok") is True:
                    return True
    except Exception:
        pass

    appear_wait_ms = min(max(timeout_ms, 15000), 60000)
    start = time.time()
    while (time.time() - start) * 1000 < appear_wait_ms:
        if stop_event and stop_event.is_set():
            if logger:
                logger("Slider: 收到停止信号，终止等待")
            raise RuntimeError(ERROR_STOPPED)
        try:
            if _ok_visible(800):
                if logger:
                    logger("Slider: 等待期间检测到后续步骤，视为通过")
                if cache_key:
                    _SLIDER_PASS_CACHE[cache_key] = {"ok": True, "t": time.time()}
                return True
        except Exception:
            pass
        if element_visible(
            driver, xpaths["slider_iframe"], 800, poll_ms
        ) or element_visible(driver, xpaths["slider_container"], 800, poll_ms):
            break
        time.sleep(max(0.2, poll_ms / 1000.0))
    else:
        if logger:
            logger(f"Slider: {int(appear_wait_ms)}ms 内未出现滑块")
            log_page_timing(driver, logger)
            log_resource_phase_timings(driver, logger, domain_filter, 30)
        return False
    try:
        iframe_probe_ms = min(timeout_ms, 6000)
        if element_exists(driver, xpaths["slider_iframe"], iframe_probe_ms, poll_ms):
            iframe = driver.find_element(By.XPATH, xpaths["slider_iframe"])
            driver.switch_to.frame(iframe)
        for i in range(10):
            # 每次尝试重新定位容器与句柄，避免刷新后引用失效
            if stop_event and stop_event.is_set():
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                if logger:
                    logger("Slider: 收到停止信号，终止滑块处理")
                raise RuntimeError(ERROR_STOPPED)
            try:
                # 确保位于正确的文档或iframe中
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                if element_exists(driver, xpaths["slider_iframe"], 1500, poll_ms):
                    try:
                        iframe = driver.find_element(By.XPATH, xpaths["slider_iframe"])
                        driver.switch_to.frame(iframe)
                    except Exception:
                        pass
                container = None
                for xp in [
                    xpaths["slider_container"],
                    "//*[contains(@class,'slider-shadow')]",
                    "//*[contains(@class,'kwai-captcha-slider-wrapper')]",
                ]:
                    try:
                        container = driver.find_element(By.XPATH, xp)
                        if container:
                            break
                    except Exception:
                        pass
                if not container:
                    time.sleep(1.5)  # 找不到容器时增加等待
                    continue
                handle_wait_ms = min(timeout_ms, 4000)
                handle = None
                handle_candidates = [
                    xpaths["slider_handle"],
                    "//*[contains(@class,'slider-btn')]",
                    "//*[contains(@class,'btn-icon')]",
                ]
                # 显式等待滑块句柄出现
                try:
                    WebDriverWait(driver, 5).until(
                        lambda d: any(
                            element_exists(d, hx, 500, poll_ms)
                            for hx in handle_candidates
                        )
                    )
                except Exception:
                    pass

                for hx in handle_candidates:
                    try:
                        handle = WebDriverWait(
                            driver,
                            handle_wait_ms / 1000.0,
                            poll_frequency=poll_ms / 1000.0,
                        ).until(EC.presence_of_element_located((By.XPATH, hx)))
                        if handle:
                            break
                    except Exception:
                        handle = None
                try:
                    width = driver.execute_script(
                        "return Math.floor(arguments[0].getBoundingClientRect().width)||arguments[0].offsetWidth||200;",
                        container,
                    )
                except Exception:
                    width = container.size.get("width") or 200
                handle_w = 24
                try:
                    if handle is not None:
                        hw = handle.size.get("width")
                        if hw:
                            handle_w = hw
                except Exception:
                    pass
                dist = 238
            except Exception:
                dist = 238
            try:
                actions = ActionChains(driver)
                if handle is not None:
                    actions.move_to_element(handle).pause(0.01).move_by_offset(
                        2, 0
                    ).click_and_hold(handle).pause(0.01)
                else:
                    h = container.size.get("height") or 20
                    actions.move_to_element_with_offset(container, 5, int(h / 2)).pause(
                        0.01
                    ).click_and_hold().pause(0.01)
                # Faster slider: reduce steps significantly
                steps = max(2, int(dist / 100))
                step_len = max(40, int(dist / steps))
                moved = 0
                # Use a single action chain for smoother and faster movement
                for _ in range(steps):
                    left = dist - moved
                    dx = min(step_len, left)
                    actions.move_by_offset(dx, 0)
                    moved += dx
                actions.release().perform()

                if _ok_visible(1500):
                    if cache_key:
                        _SLIDER_PASS_CACHE[cache_key] = {"ok": True, "t": time.time()}
                    if logger:
                        logger(
                            f"Slider: 通过(检测到后续元素)，耗时 {time.time() - t0:.2f}s"
                        )
                    return True

                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass

                if not _slider_container_visible(600):
                    if stop_event and stop_event.is_set():
                        if logger:
                            logger("Slider: 收到停止信号，终止等待")
                        raise RuntimeError(ERROR_STOPPED)
                    if _ok_visible(1200):
                        if cache_key:
                            _SLIDER_PASS_CACHE[cache_key] = {
                                "ok": True,
                                "t": time.time(),
                            }
                        if logger:
                            logger(
                                f"Slider: 通过(检测到后续元素)，耗时 {time.time() - t0:.2f}s"
                            )
                        return True

                    # 优化: 滑块消失且一段时间未重现，视为通过，避免死等后续元素导致重试
                    time.sleep(1.0)
                    if not _slider_container_visible(500):
                        # 二次确认: 等待 3s 确保不是因刷新导致的短暂消失
                        time.sleep(3.0)
                        if not _slider_container_visible(500):
                            if logger:
                                logger(
                                    f"Slider: 通过(滑块消失并确认)，耗时 {time.time() - t0:.2f}s"
                                )
                            if cache_key:
                                _SLIDER_PASS_CACHE[cache_key] = {
                                    "ok": True,
                                    "t": time.time(),
                                }
                            return True

                    time.sleep(0.4)
                    continue

                time.sleep(0.1)
            except Exception:
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                time.sleep(0.2)
    except Exception:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    if _ok_visible(min(5000, timeout_ms)):
        if cache_key:
            _SLIDER_PASS_CACHE[cache_key] = {"ok": True, "t": time.time()}
        if logger:
            logger(f"Slider: 通过(最终兜底检测)，耗时 {time.time() - t0:.2f}s")
        return True
    if logger:
        logger(f"Slider: 失败，耗时 {time.time() - t0:.2f}s")
        log_page_timing(driver, logger)
        log_resource_phase_timings(driver, logger, domain_filter, 30)
        log_performance_network(driver, logger, limit=120, domain_filter=domain_filter)
    return False


# =============================================================================
# BROWSER UTILITIES
# =============================================================================


def open_attached_driver(
    open_data: Dict[str, Any],
    max_retries: int = 3,
    logger: Optional[Callable[[str], None]] = None,
) -> webdriver.Chrome:
    """Open a Chrome driver attached to an existing browser instance with retry mechanism."""
    driver_path = open_data.get("driver")
    debugger_address = open_data.get("http")
    if not driver_path or not debugger_address:
        raise RuntimeError(f"No driver/http returned: {open_data}")

    # 验证 debugger_address 格式
    if not debugger_address.startswith("http://") and not debugger_address.startswith(
        "https://"
    ):
        debugger_address = f"http://{debugger_address}"

    last_exception = None
    for attempt in range(max_retries):
        try:
            if logger and attempt > 0:
                logger(f"WebDriver连接重试 {attempt + 1}/{max_retries}...")

            options = webdriver.ChromeOptions()
            options.debugger_address = debugger_address.replace("http://", "").replace(
                "https://", ""
            )
            try:
                options.page_load_strategy = "none"
            except Exception:
                pass
            try:
                options.set_capability(
                    "goog:loggingPrefs", {"browser": "ALL", "performance": "ALL"}
                )
            except Exception:
                pass

            # 先检查调试端口是否可连接
            import urllib.request

            debug_check_url = f"{debugger_address}/json/version"
            try:
                with urllib.request.urlopen(debug_check_url, timeout=5) as response:
                    if response.status != 200:
                        raise RuntimeError(f"调试端口未就绪: {debugger_address}")
            except Exception as e:
                if logger:
                    logger(f"调试端口检查失败: {e}")
                # 不阻断，继续尝试连接

            service = Service(executable_path=driver_path)
            driver = webdriver.Chrome(service=service, options=options)

            # 验证连接成功
            try:
                _ = driver.current_url  # 测试连接
            except Exception as e:
                driver.quit()
                raise RuntimeError(f"WebDriver连接验证失败: {e}")

            # 设置超时 - 防止长时间阻塞
            try:
                driver.set_page_load_timeout(60)  # 页面加载超时30秒
            except Exception:
                pass
            try:
                driver.set_script_timeout(60)  # JS执行超时30秒
            except Exception:
                pass
            try:
                driver.implicitly_wait(0)  # 禁用隐式等待，使用显式等待
            except Exception:
                pass

            if logger:
                logger(f"WebDriver连接成功: {debugger_address}")
            return driver

        except Exception as e:
            last_exception = e
            wait_time = 2**attempt
            if logger:
                logger(f"WebDriver连接失败(尝试 {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                if logger:
                    logger(f"等待 {wait_time} 秒后重试...")
                time.sleep(wait_time)

    raise last_exception if last_exception else RuntimeError("无法连接浏览器")


def create_cdp_tab(
    debugger_address: str, url: str, logger: Optional[Callable[[str], None]] = None
) -> bool:
    """Create a new tab via Chrome DevTools Protocol."""
    try:
        if debugger_address.startswith("http://"):
            base = debugger_address
        else:
            base = f"http://{debugger_address}"
        encoded = urllib.parse.quote(url, safe="")
        target_url = f"{base}/json/new?{encoded}"

        if logger:
            logger(f"调试接口(PUT): {target_url}")

        # Try PUT first (Standard CDP)
        try:
            r = requests.put(target_url, timeout=2)
            if r.status_code == 200:
                return True
            if r.status_code != 405:
                if logger:
                    logger(f"PUT失败: {r.status_code}")
        except requests.exceptions.RequestException as e:
            if logger:
                logger(f"PUT请求异常: {e}")

        # Fallback to GET if PUT failed (Compatibility)
        if logger:
            logger(f"调试接口(GET重试): {target_url}")
        r = requests.get(target_url, timeout=2)
        if logger:
            logger(f"GET响应: {r.status_code}")
        return r.status_code == 200
    except requests.exceptions.RequestException as e:
        if logger:
            logger(f"调试接口最终异常: {e}")
        return False


def open_tab_via_debugger(
    debugger_address: str, url: str, logger: Optional[Callable[[str], None]] = None
) -> bool:
    """Open a new tab via debugger protocol."""
    return create_cdp_tab(debugger_address, url, logger)


def proxy_payload(
    host: str,
    port: str,
    username: Optional[str],
    password: Optional[str],
    protocol: str = "socks5",
) -> Dict[str, Any]:
    """Construct proxy payload for BitBrowser."""
    if not host or not port:
        return {"proxyType": "noproxy"}

    # Map protocol to bitbrowser proxyType if needed
    # Common types: socks5, http, https
    ptype = protocol.lower() if protocol else "socks5"
    if ptype not in ("socks5", "http", "https", "ssh"):
        ptype = "socks5"

    p: Dict[str, Any] = {
        "proxyType": ptype,
        "host": host,
        "port": str(port),
        "proxyHost": host,
        "proxyPort": str(port),
    }
    if username:
        p["proxyUserName"] = username
    if password:
        p["proxyPassword"] = password

    # Construct proxy string URL
    if username and password:
        p["proxy"] = f"{ptype}://{username}:{password}@{host}:{port}"
    else:
        p["proxy"] = f"{ptype}://{host}:{port}"
    return p


def log_window_urls(
    driver: webdriver.Remote, logger: Optional[Callable[[str], None]]
) -> None:
    """Log all window URLs."""
    if not logger:
        return
    try:
        handles = driver.window_handles
        logger(f"当前窗口句柄数量: {len(handles)}")
        for i, h in enumerate(handles):
            try:
                driver.switch_to.window(h)
                url = driver.current_url
            except Exception:
                url = "(不可获取URL)"
            logger(f"句柄[{i}]: {url}")
    except Exception as e:
        logger(f"枚举窗口失败: {e}")


def log_browser_console(
    driver: webdriver.Remote, logger: Optional[Callable[[str], None]], limit: int = 50
) -> None:
    """Log browser console messages."""
    if not logger:
        return
    try:
        logs = driver.get_log("browser")
        if not logs:
            return
        size = len(logs)
        start = max(0, size - limit)
        for entry in logs[start:]:
            try:
                ts = entry.get("timestamp")
                lvl = entry.get("level")
                msg = entry.get("message")
                logger(f"Console[{lvl}]: {msg}")
            except Exception:
                pass
    except Exception:
        pass


def log_performance_network(
    driver: webdriver.Remote,
    logger: Optional[Callable[[str], None]],
    limit: int = 200,
    domain_filter: Optional[str] = None,
) -> None:
    """Log performance network events."""
    if not logger:
        return

    def redact(h: Any) -> Any:
        if not isinstance(h, dict):
            return h
        out = {}
        for k, v in h.items():
            try:
                kl = str(k).lower()
            except Exception:
                kl = ""
            if kl in (
                "authorization",
                "cookie",
                "set-cookie",
                "x-api-key",
                "proxy-authorization",
            ):
                out[k] = "[REDACTED]"
            else:
                out[k] = v
        return out

    try:
        logs = driver.get_log("performance")
        if not logs:
            return
        size = len(logs)
        start = max(0, size - limit)
        for entry in logs[start:]:
            try:
                msg_str = entry.get("message") or ""
                data = json.loads(msg_str)
                m = data.get("message", {})
                method = m.get("method")
                params = m.get("params", {})
                if method in ("Network.requestWillBeSent", "Network.responseReceived"):
                    req = params.get("request", {})
                    res = params.get("response", {})
                    url = req.get("url") or res.get("url") or ""
                    if domain_filter and (domain_filter not in url):
                        continue
                    if method == "Network.requestWillBeSent":
                        logger(f"HTTP[REQ] {req.get('method')} {url}")
                        hdrs = req.get("headers") or {}
                        try:
                            logger(
                                f"HTTP[REQ-HEADERS] {json.dumps(redact(hdrs), ensure_ascii=False)}"
                            )
                        except Exception:
                            pass
                    if method == "Network.responseReceived":
                        status = res.get("status")
                        logger(f"HTTP[RES] {status} {url}")
                        hdrs = res.get("headers") or {}
                        try:
                            logger(
                                f"HTTP[RES-HEADERS] {json.dumps(redact(hdrs), ensure_ascii=False)}"
                            )
                        except Exception:
                            pass
            except Exception:
                pass
    except Exception:
        pass


def log_response_bodies(
    driver: webdriver.Remote,
    logger: Optional[Callable[[str], None]],
    limit: int = 10,
    domain_filter: Optional[str] = None,
) -> None:
    """Log response bodies from performance logs."""
    if not logger:
        return
    try:
        logs = driver.get_log("performance")
        if not logs:
            return
        ids = []
        for entry in reversed(logs):
            try:
                data = json.loads(entry.get("message") or "{}")
                m = data.get("message", {})
                method = m.get("method")
                params = m.get("params", {})
                if method == "Network.responseReceived":
                    res = params.get("response", {})
                    url = res.get("url") or ""
                    if domain_filter and (domain_filter not in url):
                        continue
                    rid = params.get("requestId")
                    if rid:
                        ids.append((rid, url))
                        if len(ids) >= limit:
                            break
            except Exception:
                pass
        for rid, url in ids:
            try:
                body = driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": rid}
                )
                txt = body.get("body", "")
                logger(f"HTTP[BODY] {url} len={len(txt)}")
            except Exception:
                pass
    except Exception:
        pass


def extract_verification_code_unified(
    driver: webdriver.Remote,
    email_addr: str,
    password: str,
    resend_xpath: Optional[str],
    logger: Optional[Callable[[str], None]],
    timeout: int = CODE_EXTRACTION_TIMEOUT_SEC,
    proxy_config: Optional[Dict[str, Any]] = None,
    stop_event: Optional[threading.Event] = None,
    email_pool: Optional[Any] = None,
) -> Optional[str]:
    if logger:
        logger(f"Unified Captcha: Start (Email: {email_addr}) Mode: IMAP")

    try:
        main_handle = driver.current_window_handle
    except Exception:
        main_handle = None

    use_imap_proxy = os.getenv("DISABLE_IMAP_PROXY", "false").lower() != "true"
    imap_proxy = proxy_config if use_imap_proxy else None

    def _mark_problem(reason: str) -> None:
        """把当前邮箱立即标记为问题邮箱。"""
        if not email_pool:
            return

        reason = str(reason or "未知错误")
        reason = reason.replace("\n", " ").replace("\r", " ")
        reason = reason[:180]

        try:
            if hasattr(email_pool, "update_email_status"):
                email_pool.update_email_status(email_addr, "problem", reason=reason)
            elif hasattr(email_pool, "update_status"):
                email_pool.update_status(email_addr, "problem", reason=reason)
        except Exception as e:
            if logger:
                logger(f"Unified Captcha: 标记问题邮箱失败: {e}")

    def _is_fatal_oauth_error(err_text: str) -> bool:
        """判断是否为 OAuth/token 致命错误。"""
        s = (err_text or "").lower()
        fatal_keywords = [
            "failed to get access token",
            "oauth exchange failed",
            "authenticationerror",
            "invalid_grant",
            "invalid_client",
            "bad request",
            "400 client error",
            "401 client error",
            "403 client error",
        ]
        return any(k in s for k in fatal_keywords)

    def _precheck_oauth_token() -> None:
        """
        Outlook OAuth 邮箱先快速检测 token。
        如果微软直接返回 400/401/403，说明 token 已废，立即 problem，不再等 400 秒。
        """
        if "|||" not in str(password or ""):
            return

        try:
            client_id, refresh_token = str(password).split("|||", 1)
            client_id = client_id.strip()
            refresh_token = refresh_token.strip()
        except Exception:
            _mark_problem("OAuth格式错误: 缺少 client_id 或 refresh_token")
            raise RuntimeError("oauth_token_failed")

        if not client_id or not refresh_token:
            _mark_problem("OAuth格式错误: client_id 或 refresh_token 为空")
            raise RuntimeError("oauth_token_failed")

        try:
            import requests

            url = "https://login.microsoftonline.com/common/oauth2/v2.0/token"
            data = {
                "client_id": client_id,
                "grant_type": "refresh_token",
                "refresh_token": refresh_token,
            }

            session = requests.Session()
            session.trust_env = False

            if logger:
                logger(f"Unified Captcha: OAuth预检 Access Token: {email_addr}")

            res = session.post(url, data=data, timeout=15)

            if res.status_code in (400, 401, 403):
                body = ""
                try:
                    body = res.text[:300]
                except Exception:
                    pass

                reason = f"OAuth token失效: HTTP {res.status_code} {body}"
                if logger:
                    logger(f"Unified Captcha: ❌ {reason}")

                _mark_problem(reason)
                raise RuntimeError("oauth_token_failed")

            # 其他非 2xx 先不立刻判死刑，交给后面的 IMAP 流程处理
            if not res.ok and logger:
                logger(
                    f"Unified Captcha: OAuth预检非成功状态 HTTP {res.status_code}，继续走正常接码流程"
                )

        except RuntimeError:
            raise
        except Exception as e:
            # 网络抖动/微软接口临时异常，不直接标 problem，继续走原接码流程
            if logger:
                logger(f"Unified Captcha: OAuth预检异常，继续正常流程: {e}")

    # 关键修复：OAuth 坏号先预检，坏号立即退出，不再占窗口等 400 秒
    _precheck_oauth_token()

    extractor = None

    try:
        extractor = MailExtractor(
            email_addr,
            password,
            proxy_config=imap_proxy,
            logger=logger,
        )
    except Exception as e:
        err = f"{type(e).__name__}: {e}"

        if _is_fatal_oauth_error(err):
            if logger:
                logger(f"Unified Captcha: ❌ OAuth初始化失败，立即标记问题邮箱: {err}")
            _mark_problem(f"OAuth初始化失败: {err}")
            raise RuntimeError("oauth_token_failed")

        if logger:
            logger(f"Unified Captcha: ❌ 初始化邮箱连接失败: {e}")

        _mark_problem(f"IMAP登录失败: {str(e)}")
        return None

    try:
        if logger:
            logger(f"Unified Captcha: 开始监听邮件 (最大等待 {timeout} 秒)...")

        t_poll_start = time.time()
        code = None

        while time.time() - t_poll_start < timeout:
            if stop_event and stop_event.is_set():
                return None

            try:
                if not extractor:
                    extractor = MailExtractor(
                        email_addr,
                        password,
                        proxy_config=imap_proxy,
                        logger=logger,
                    )

                code = extractor.get_latest_verification_code()

                if (
                    code
                    and code not in ["未匹配到验证码", "未找到邮件"]
                    and not str(code).startswith("错误:")
                ):
                    if logger:
                        logger(f"Unified Captcha: ✅ 成功获取验证码: {code}")

                    try:
                        if main_handle:
                            driver.switch_to.window(main_handle)
                    except Exception:
                        pass

                    return code

            except Exception as e:
                err = f"{type(e).__name__}: {e}"

                if _is_fatal_oauth_error(err):
                    if logger:
                        logger(f"Unified Captcha: ❌ OAuth失效，立即标记问题邮箱并跳过: {err}")

                    _mark_problem(f"OAuth失效: {err}")
                    raise RuntimeError("oauth_token_failed")

                if logger:
                    logger(f"Unified Captcha: 邮箱读取异常，继续重试: {err}")

                extractor = None

            time.sleep(3)

        if logger:
            logger(f"Unified Captcha: ❌ {timeout} 秒内未收到验证码，放弃当前邮箱。")

        return None

    finally:
        if extractor:
            try:
                extractor.close()
            except Exception:
                pass


# =============================================================================
# REGISTRATION STEP FUNCTIONS
# =============================================================================


def step_verify(
    ctx: Dict[str, Any],
    row: Dict[str, Any],
    xpaths: Dict[str, str],
    client: BitBrowserClient,
    platform_url: str,
    timeout_ms: int,
    poll_ms: int,
    logger: Optional[Callable[[str], None]],
    stop_event: Optional[threading.Event],
    target_check: Optional[Callable[[], bool]],
    email_pool: Optional[Any],
    headless_mode: bool,
    udp_enabled: bool,
) -> None:
    """
    Step 1: Verification and initialization.

    Validates email, creates/opens browser, and navigates to platform.
    """
    email = str(row.get("email") or row.get("账号") or "").strip()
    password = str(row.get("password") or row.get("密码") or "").strip()
    host = str(row.get("host") or row.get("代理IP") or "").strip()
    port = str(row.get("port") or row.get("端口") or "").strip()
    proxy_username = str(row.get("proxyUserName") or row.get("用户名") or "").strip()
    proxy_password = str(row.get("proxyPassword") or row.get("密码2") or "").strip()
    protocol = str(row.get("protocol") or row.get("proxyType") or "socks5").strip()
    window_name = str(
        row.get("windowName") or row.get("窗口名称") or email or "win"
    ).strip()

    # Email Validation Logic
    if email_pool:
        is_valid, reason = email_pool.check_email_availability(email)
        if not is_valid:
            # Allow 'processing' as it is set by the current task runner immediately before execution
            if reason == "正在处理中":
                pass
            else:
                log_time = time.strftime("%Y-%m-%d %H:%M:%S")
                audit_msg = f"Audit: Invalid Email Rejected | Time: {log_time} | IP: {host} | Email: {email} | Reason: {reason}"
                if logger:
                    logger(audit_msg)
                raise RuntimeError(f"{ERROR_EMAIL_UNAVAILABLE}: {reason}")

    # Construct proxy_config
    proxy_config = None
    if host and port:
        try:
            ptype = socks.SOCKS5
            if protocol.lower() == "http":
                ptype = socks.HTTP
            elif protocol.lower() == "socks4":
                ptype = socks.SOCKS4

            proxy_config = {
                "proxy_type": ptype,
                "addr": host,
                "port": int(port),
                "username": proxy_username if proxy_username else None,
                "password": proxy_password if proxy_password else None,
                "rdns": True,
            }
        except (ValueError, TypeError) as e:
            if logger:
                logger(f"代理配置异常: {e}")

    # Early target check
    if target_check and target_check():
        raise RuntimeError(ERROR_TARGET_REACHED)

    if stop_event and stop_event.is_set():
        raise RuntimeError(ERROR_STOPPED)

    if logger:
        logger(f"create_profile {window_name}")

    # Optimization: Disable QUIC
    # 原有的 browser_cmd_args 逻辑保持不动（作为双重保险）
    browser_cmd_args = ["--disable-quic", "--window-size=800,550"]
    if ctx.get("save_traffic"):
        browser_cmd_args.append("--blink-settings=imagesEnabled=false")

    browser_id = None
    try:
        # ⚠️ 注意这里：增加了 save_traffic=ctx.get("save_traffic")
        browser_id = client.update_browser(
            window_name,
            proxy_payload(host, port, proxy_username, proxy_password, protocol),
            enable_udp=udp_enabled,
            cmd_args=browser_cmd_args,
            save_traffic=ctx.get("save_traffic"),
        )
    except (requests.exceptions.RequestException, RuntimeError):
        try:
            # ⚠️ 这里也要改
            browser_id = client.create_browser(
                window_name,
                proxy_payload(host, port, proxy_username, proxy_password, protocol),
                enable_udp=udp_enabled,
                cmd_args=browser_cmd_args,
                save_traffic=ctx.get("save_traffic"),
            )
        except Exception as create_err:
            if logger:
                logger(f"创建窗口失败: {create_err}")
            raise RuntimeError(f"create_browser failed: {create_err}")

    ctx["browser_id"] = browser_id
    if logger:
        logger(f"profile_id {browser_id}")

    if headless_mode:
        if logger:
            logger("🚀 已启用 BitBrowser 官方无头模式 (--headless)")

    open_data = client.open_browser(
        browser_id,
        headless_mode=headless_mode,
    )

    ctx["open_data"] = open_data

    # Try to get PID for cleanup
    browser_pid = open_data.get("pid")
    if not browser_pid and open_data.get("http"):
        try:
            addr = open_data.get("http")
            if ":" in addr:
                port = int(addr.split(":")[-1])
                browser_pid = get_pid_by_port(port)
        except Exception:
            pass
    ctx["browser_pid"] = browser_pid

    if logger:
        logger(f"open_browser {browser_id} -> {open_data} (PID={browser_pid})")

    if not (open_data.get("driver") and open_data.get("http")):
        raise RuntimeError(f"open_browser 未返回 driver/http: {open_data}")

    # 等待浏览器调试端口就绪
    debugger_addr = open_data.get("http", "")
    if logger:
        logger(f"等待浏览器调试端口就绪: {debugger_addr}")

    import urllib.request

    port_ready = False
    for port_check in range(10):  # 最多等待10次
        try:
            check_url = (
                f"http://{debugger_addr}/json/version"
                if not debugger_addr.startswith("http")
                else f"{debugger_addr}/json/version"
            )
            with urllib.request.urlopen(check_url, timeout=2) as resp:
                if resp.status == 200:
                    port_ready = True
                    if logger:
                        logger(f"浏览器调试端口已就绪")
                    break
        except Exception:
            pass
        time.sleep(1)

    if not port_ready and logger:
        logger("警告: 浏览器调试端口可能未就绪，继续尝试连接...")

    # Wait for browser to fully start
    time.sleep(BROWSER_START_WAIT_SEC)

    driver = None
    try:
        driver = open_attached_driver(open_data, max_retries=3, logger=logger)
    except Exception as attach_err:
        if logger:
            logger(f"连接浏览器最终失败: {attach_err}")
        raise RuntimeError(f"无法连接到比特浏览器: {attach_err}")

    ctx["driver"] = driver
    ctx["t_attached"] = time.time()
    ctx["gate_state"] = "attached"

    # 强制设置比特浏览器窗口大小
    try:
        driver.set_window_size(800, 650)
    except Exception:
        pass

    ctx["driver"] = driver
    ctx["t_attached"] = time.time()
    ctx["gate_state"] = "attached"

    # ============================================================
    # 浏览器渲染尺寸
    #
    # 真无头模式使用桌面分辨率，避免 Kling 进入手机/窄屏响应式布局。
    # 普通模式继续保持原来的较小窗口，避免占用桌面。
    # ============================================================
    try:
        if headless_mode:
            target_width = 1440
            target_height = 900
        else:
            target_width = 800
            target_height = 650

        driver.set_window_size(target_width, target_height)

        # 读取浏览器真正生效的尺寸
        window_size = driver.get_window_size()

        viewport_info = driver.execute_script(
            """
            return {
                innerWidth: window.innerWidth,
                innerHeight: window.innerHeight,
                outerWidth: window.outerWidth,
                outerHeight: window.outerHeight,
                devicePixelRatio: window.devicePixelRatio,
                screenWidth: screen.width,
                screenHeight: screen.height
            };
            """
        )

        if logger:
            logger(
                f"浏览器尺寸: "
                f"headless={headless_mode}, "
                f"window={window_size}, "
                f"viewport={viewport_info}"
            )

    except Exception as e:
        if logger:
            logger(f"⚠️ 设置/读取浏览器尺寸失败: {e}")

    if logger:
        logger("步骤: 打开平台网址")

    if not check_connectivity(
        driver, platform_url, logger, max_wait=45, udp_enabled=udp_enabled
    ):
        take_screenshot(driver, f"{window_name}_connectivity_failed.png", logger)
        raise RuntimeError(ERROR_PROXY_CONNECTIVITY_FAILED)

    if logger:
        logger("步骤: 打开平台网址")
    if not check_connectivity(
        driver, platform_url, logger, max_wait=45, udp_enabled=udp_enabled
    ):
        take_screenshot(driver, f"{window_name}_connectivity_failed.png", logger)
        raise RuntimeError(ERROR_PROXY_CONNECTIVITY_FAILED)

    # DOM Check (智能弹性等待：不设死板时间，根据 UI 的超时设置灵活探测核心元素)
    try:
        if logger:
            logger("步骤: 智能等待页面核心元素加载...")
            # [健壮性优化] 极速探测 Cloudflare 拦截 (Fast-Fail)，绝不浪费时间死等
        page_title = driver.title.lower()
        if (
            "just a moment" in page_title
            or "attention required" in page_title
            or "cloudflare" in page_title
        ):
            if logger:
                logger(
                    "🚫 致命拦截: 检测到 Cloudflare 验证盾，当前 IP 信誉极差，立即放弃本轮！"
                )
            raise RuntimeError("cloudflare_blocked")
        # 将我们关心的核心入口汇总
        core_xpaths = [
            xpaths.get("language_menu"),
            xpaths.get("language_menu_alt"),
            xpaths.get("signin_btn"),
            xpaths.get("Creative Studio"),
            "//*[contains(text(), 'Sign In') or contains(text(), '登录')]",
        ]

        # 使用你 UI 上配置的最大超时时间 (timeout_ms) 进行智能探测
        found_element = first_present_xpath(driver, core_xpaths, timeout_ms, poll_ms)

        if not found_element:
            if logger:
                logger("DOM状态检测超时：页面可能卡白屏或代理极慢，尝试刷新补救...")
            try:
                driver.refresh()
                wait_page_ready(driver, 30)
            except Exception:
                pass

            # 刷新后进行最后一次探测
            found_element = first_present_xpath(
                driver, core_xpaths, timeout_ms, poll_ms
            )

            if not found_element:
                if logger:
                    logger("警告：二次刷新后核心元素仍未就绪，强制进入下一步(可能报错)")
        else:
            if logger:
                logger("页面核心元素已就绪，进入交互流程。")

    except Exception as e:
        if logger:
            logger(f"DOM 校验阶段发生异常: {e}")


def step_write(
    ctx: Dict[str, Any],
    row: Dict[str, Any],
    xpaths: Dict[str, str],
    platform_url: str,
    timeout_ms: int,
    poll_ms: int,
    logger: Optional[Callable[[str], None]],
    stop_event: Optional[threading.Event],
) -> None:
    """
    Step 2: Write and interact with registration form.

    Fills in email, password, and handles slider captcha.
    """
    driver = ctx["driver"]
    open_data = ctx["open_data"]
    email = str(row.get("email") or row.get("账号") or "").strip()
    password = str(row.get("password") or row.get("密码") or "").strip()

    # Check Ready
    ready = False
    for xp_key in ("language_menu", "signin_btn", "Creative Studio"):
        try:
            xp_val = xpaths.get(xp_key)
            if xp_val and element_exists(driver, xp_val, 8000, poll_ms):
                ready = True
                break
        except Exception:
            pass

    if not ready:
        try:
            driver.refresh()
        except Exception:
            pass
        wait_page_ready(driver, 20)

    # Debugger Tab
    enable_debugger_open = bool(ctx.get("udp_enabled", False))
    if enable_debugger_open and (not ready) and open_data.get("http"):
        if not ctx.get("debug_tab_opened", False):
            prev_handles = driver.window_handles
            open_tab_via_debugger(open_data.get("http") or "", platform_url, logger)
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: len(d.window_handles) > len(prev_handles)
                )
                driver.switch_to.window(driver.window_handles[-1])
            except Exception:
                pass
            ctx["debug_tab_opened"] = True

    # Language Menu
    lang_clicked = False
    if (
        safe_click_any(
            driver,
            [xpaths.get("language_menu"), xpaths.get("language_menu_alt")],
            timeout_ms,
            poll_ms,
            logger,
            retries=2,
        )
        or js_click_xpath(driver, xpaths.get("language_menu", ""))
        or js_click_xpath(driver, xpaths.get("language_menu_alt", ""))
    ):
        if logger:
            logger("步骤: 打开语言菜单")
        lang_clicked = True

        if not safe_click_any(
            driver,
            [xpaths.get("english_option"), xpaths.get("english_option_alt")],
            timeout_ms,
            poll_ms,
            logger,
            retries=2,
        ):
            if not js_click_xpath(
                driver, xpaths.get("english_option", "")
            ) and not js_click_xpath(driver, xpaths.get("english_option_alt", "")):
                raise RuntimeError(ERROR_ENGLISH_OPTION_CLICK_FAILED)

    if logger:
        logger(
            "步骤: 选择英文"
            + ("(已点击语言菜单)" if lang_clicked else "(跳过语言菜单点击)")
        )

    if stop_event and stop_event.is_set():
        raise RuntimeError(ERROR_STOPPED)

    if logger:
        logger("步骤: 点击 Creative Studio")
    if not safe_click_any(
        driver,
        [
            xpaths.get("Creative Studio"),
            "//*[contains(text(),'Creative') or contains(text(),'创意') or contains(text(),'工作室')]",
        ],
        timeout_ms,
        poll_ms,
        logger,
        retries=2,
    ):
        raise RuntimeError("Creative Studio 点击失败")

    if logger:
        logger("步骤: 点击 More Tools")
    prev_handles = driver.window_handles
    ctx["gate_state"] = "pre_open_more_tools"
    if not safe_click_any(
        driver,
        [
            xpaths.get("More Tools"),
            "//*[contains(text(),'More') or contains(text(),'更多') or contains(text(),'工具')]",
        ],
        timeout_ms,
        poll_ms,
        logger,
        retries=2,
    ):
        raise RuntimeError(ERROR_MORE_TOOLS_CLICK_FAILED)

    try:
        WebDriverWait(
            driver, min(8, timeout_ms / 1000.0), poll_frequency=poll_ms / 1000.0
        ).until(lambda d: len(d.window_handles) > len(prev_handles))
    except Exception:
        pass

    time.sleep(1.5)
    handle_lock = threading.Lock()
    with handle_lock:
        try:
            new_handles = driver.window_handles
            diff = [h for h in new_handles if h not in prev_handles]
            if diff:
                driver.switch_to.window(diff[-1])
            else:
                driver.switch_to.window(driver.window_handles[-1])
            ctx["gate_state"] = "tab_switched"
        except Exception:
            pass

    if logger:
        logger("步骤: 已切换到新标签")

    wait_page_ready(driver, 20)
    time.sleep(1)

    # ============================================================
    # Sign In 前页面诊断
    #
    # 用来判断无头模式失败到底属于：
    # 1. 切错标签
    # 2. 页面还没加载
    # 3. 响应式布局导致 Sign In 被隐藏
    # 4. Sign In 文案/DOM发生变化
    # 5. Sign In 位于 iframe
    # ============================================================
    try:
        current_url = driver.current_url
    except Exception as e:
        current_url = f"<读取失败: {e}>"

    try:
        current_title = driver.title
    except Exception as e:
        current_title = f"<读取失败: {e}>"

    try:
        current_handle = driver.current_window_handle
    except Exception:
        current_handle = "unknown"

    try:
        all_handles = driver.window_handles
    except Exception:
        all_handles = []

    try:
        current_size = driver.get_window_size()
    except Exception:
        current_size = {}

    try:
        viewport_info = driver.execute_script(
            """
            return {
                innerWidth: window.innerWidth,
                innerHeight: window.innerHeight,
                outerWidth: window.outerWidth,
                outerHeight: window.outerHeight,
                devicePixelRatio: window.devicePixelRatio
            };
            """
        )
    except Exception:
        viewport_info = {}

    try:
        iframe_count = len(driver.find_elements(By.TAG_NAME, "iframe"))
    except Exception:
        iframe_count = -1

    if logger:
        logger(
            f"🔍 SignIn诊断: "
            f"URL={current_url} | "
            f"Title={current_title} | "
            f"Handle={current_handle} | "
            f"Handles={len(all_handles)} | "
            f"Window={current_size} | "
            f"Viewport={viewport_info} | "
            f"Iframes={iframe_count}"
        )

    # ------------------------------------------------------------
    # 检查页面里到底有没有包含 Sign In / 登录 的元素，
    # 包括隐藏元素。
    # ------------------------------------------------------------
    try:
        signin_candidates = driver.find_elements(
            By.XPATH,
            "//*[contains(normalize-space(.),'Sign In') or "
            "contains(normalize-space(.),'登录')]"
        )

        visible_count = 0

        if logger:
            logger(
                f"🔍 SignIn元素扫描: 总匹配={len(signin_candidates)}"
            )

        # 最多输出前10个，避免日志爆炸
        for index, el in enumerate(signin_candidates[:10], start=1):
            try:
                displayed = el.is_displayed()

                if displayed:
                    visible_count += 1

                tag = el.tag_name
                text = (el.text or "").strip().replace("\n", " ")
                cls = el.get_attribute("class") or ""
                rect = el.rect

                if logger:
                    logger(
                        f"🔍 SignIn候选#{index}: "
                        f"tag={tag}, "
                        f"displayed={displayed}, "
                        f"text={text[:120]!r}, "
                        f"class={cls[:160]!r}, "
                        f"rect={rect}"
                    )

            except Exception as e:
                if logger:
                    logger(
                        f"🔍 SignIn候选#{index}: 读取元素信息失败: {e}"
                    )

        if logger:
            logger(
                f"🔍 SignIn元素扫描结果: "
                f"总数={len(signin_candidates)}, "
                f"可见数={visible_count}"
            )

    except Exception as e:
        if logger:
            logger(f"⚠️ SignIn元素扫描失败: {e}")


    if logger:
        logger("步骤: 点击 Sign In / One-Click Sign In")

    # ========================================================
    # 兼容普通模式 + BitBrowser Headless 模式
    #
    # 普通页面可能显示：
    #   Sign In
    #
    # 无头页面当前实际显示：
    #   One-Click Sign In
    #
    # 使用 normalize-space(.) 而不是 text()
    # 可以兼容文字被 span 等子元素包裹的情况。
    # ========================================================
    signin_xpaths = [
        # Headless 当前页面实际按钮
        "//button[contains(normalize-space(.), 'One-Click Sign In')]",

        # role=button 的情况
        "//*[@role='button' and contains(normalize-space(.), 'One-Click Sign In')]",

        # 普通模式 Sign In
        "//button[contains(normalize-space(.), 'Sign In')]",

        # 其他可点击元素
        "//*[@role='button' and contains(normalize-space(.), 'Sign In')]",

        # 原配置保留兼容
        xpaths.get("signin_btn"),

        # 中文页面兼容
        "//button[contains(normalize-space(.), '登录')]",
        "//*[@role='button' and contains(normalize-space(.), '登录')]",
    ]

    signin_clicked = safe_click_any(
        driver,
        signin_xpaths,
        timeout_ms,
        poll_ms,
        logger,
        retries=2,
    )

    # ========================================================
    # XPath没有点击成功时，使用JS进行第二层兜底
    # ========================================================
    if not signin_clicked:
        if logger:
            logger("⚠️ XPath未点击到 Sign In，开始JS兜底查找...")

        try:
            signin_clicked = bool(
                driver.execute_script(
                    """
                    const selectors = [
                        'button',
                        '[role="button"]',
                        'a'
                    ];

                    const elements = Array.from(
                        document.querySelectorAll(selectors.join(','))
                    );

                    for (const el of elements) {
                        const text = (
                            el.innerText ||
                            el.textContent ||
                            ''
                        )
                        .replace(/\\s+/g, ' ')
                        .trim();

                        const isSignin =
                            text.includes('One-Click Sign In') ||
                            text === 'Sign In' ||
                            text.includes('Sign In') ||
                            text.includes('登录');

                        if (!isSignin) {
                            continue;
                        }

                        const rect = el.getBoundingClientRect();
                        const style = window.getComputedStyle(el);

                        const visible =
                            rect.width > 0 &&
                            rect.height > 0 &&
                            style.display !== 'none' &&
                            style.visibility !== 'hidden' &&
                            style.opacity !== '0';

                        if (!visible) {
                            continue;
                        }

                        try {
                            el.scrollIntoView({
                                block: 'center',
                                inline: 'center'
                            });
                        } catch (e) {}

                        el.click();
                        return true;
                    }

                    return false;
                    """
                )
            )

            if signin_clicked:
                if logger:
                    logger("✅ JS成功点击 Sign In / One-Click Sign In")
            else:
                if logger:
                    logger("⚠️ JS也没有找到可点击的 Sign In")

        except Exception as e:
            if logger:
                logger(f"⚠️ JS点击 Sign In 异常: {e}")

    # ========================================================
    # 还是失败 → 保存完整现场
    # ========================================================
    if not signin_clicked:

        try:
            current_url = driver.current_url
        except Exception:
            current_url = "unknown"

        try:
            current_title = driver.title
        except Exception:
            current_title = "unknown"

        try:
            current_size = driver.get_window_size()
        except Exception:
            current_size = {}

        try:
            viewport_info = driver.execute_script(
                """
                return {
                    innerWidth: window.innerWidth,
                    innerHeight: window.innerHeight,
                    scrollWidth: document.documentElement.scrollWidth,
                    scrollHeight: document.documentElement.scrollHeight,
                    readyState: document.readyState
                };
                """
            )
        except Exception:
            viewport_info = {}

        # ====================================================
        # 再输出当前页面所有可能的登录按钮
        # 以后即使Kling再次改文案，也能直接从日志看出来
        # ====================================================
        try:
            signin_debug = driver.execute_script(
                """
                const els = Array.from(
                    document.querySelectorAll(
                        'button, [role="button"], a'
                    )
                );

                return els.slice(0, 100).map((el, index) => {
                    const rect = el.getBoundingClientRect();

                    return {
                        index: index,
                        tag: el.tagName,
                        text: (
                            el.innerText ||
                            el.textContent ||
                            ''
                        ).replace(/\\s+/g, ' ').trim().slice(0, 150),
                        width: Math.round(rect.width),
                        height: Math.round(rect.height),
                        displayed:
                            rect.width > 0 &&
                            rect.height > 0
                    };
                }).filter(
                    item =>
                        item.text.includes('Sign') ||
                        item.text.includes('登录')
                );
                """
            )

        except Exception as e:
            signin_debug = f"读取失败: {e}"

        if logger:
            logger(
                f"❌ Sign In 点击失败最终现场: "
                f"URL={current_url} | "
                f"Title={current_title} | "
                f"Window={current_size} | "
                f"Viewport={viewport_info} | "
                f"LoginElements={signin_debug}"
            )

        # 每个邮箱单独保存截图，8并发不会互相覆盖
        try:
            safe_email_name = (
                email.replace("@", "_at_")
                .replace(".", "_")
                .replace("/", "_")
            )

            take_screenshot(
                driver,
                f"signin_failed_{safe_email_name}.png",
                logger,
            )

        except Exception as e:
            if logger:
                logger(f"SignIn失败截图异常: {e}")

        raise RuntimeError(ERROR_SIGNIN_CLICK_FAILED)

    if logger:
        logger("步骤: 选择邮箱登录")
    if not safe_click_any(
        driver,
        [
            xpaths.get("signin_with_email"),
            # 1. 兼容大小写（防止被前端 CSS 大写欺骗）
            "//*[contains(text(),'邮箱') or contains(text(),'email') or contains(text(),'Email') or contains(text(),'邮件')]",
            # 2. 精准打击你提供的 Vue 专属 class 特征
            "//span[contains(@class, 'caption') and contains(text(), 'Sign')]",
            # 3. 向上越级打击：点击该 span 的外层父级盒子（防止 span 自身不接收点击事件）
            "//span[contains(@class, 'caption') and (contains(text(), 'email') or contains(text(), 'Email'))]/..",
        ],
        timeout_ms,
        poll_ms,
        logger,
        retries=2,
    ):
        raise RuntimeError(ERROR_SIGNIN_EMAIL_CLICK_FAILED)

    if logger:
        logger("步骤: 点击免费注册")
    if not safe_click_any(
        driver,
        [
            xpaths.get("Sign up for free"),
            "//*[contains(text(),'免费') or contains(text(),'注册') or contains(text(),'Sign up')]",
        ],
        timeout_ms,
        poll_ms,
        logger,
        retries=2,
    ):
        raise RuntimeError(ERROR_SIGNUP_CLICK_FAILED)

    # Form Filling
    if logger:
        logger("步骤: 输入邮箱")
    if not safe_send_keys_any(
        driver,
        [
            xpaths.get("Enter Email Address"),
            "//*[@placeholder][contains(@placeholder,'邮箱') or contains(@placeholder,'Email')]",
        ],
        email,
        timeout_ms,
        poll_ms,
        logger,
        retries=1,
    ):
        raise RuntimeError(ERROR_EMAIL_INPUT_FAILED)

    if logger:
        logger("步骤: 输入密码")
    if not safe_send_keys_any(
        driver,
        [
            xpaths.get("password_input"),
            "//*[@placeholder][contains(@placeholder,'密码') or contains(@placeholder,'Password')]",
        ],
        password,
        timeout_ms,
        poll_ms,
        logger,
        retries=1,
    ):
        raise RuntimeError(ERROR_PASSWORD_INPUT_FAILED)

    if logger:
        logger("步骤: 确认密码")
    if not safe_send_keys_any(
        driver,
        [
            xpaths.get("Confirm Password"),
            "//*[@placeholder][contains(@placeholder,'确认') or contains(@placeholder,'Confirm')]",
        ],
        password,
        timeout_ms,
        poll_ms,
        logger,
        retries=1,
    ):
        raise RuntimeError(ERROR_CONFIRM_INPUT_FAILED)

    if logger:
        logger("步骤: 点击下一步")
    if not safe_click_any(
        driver,
        [
            xpaths.get("next_btn"),
            "//*[contains(text(),'下一步') or contains(text(),'Next')]",
        ],
        timeout_ms,
        poll_ms,
        logger,
        retries=2,
    ):
        raise RuntimeError(ERROR_NEXT_CLICK_FAILED)

    ctx["gate_state"] = "registration_started"

    # Check Used Email (极速判断，弃用耗时的 body.text)
    try:
        # 直接定位包含具体错误信息的元素，最大等待 2.5 秒
        error_xpaths: list[str | None] = [
            "//*[contains(text(), 'already registered') or contains(text(), 'already used') or contains(text(), 'account exists')]",
            "//*[contains(text(), '已注册') or contains(text(), '已被使用')]",
        ]
        if first_present_xpath(driver, error_xpaths, 2500, poll_ms):
            if logger:
                logger("🚫 检测到邮箱已被使用提示")
            if email_pool := ctx.get("email_pool"):
                email_pool.update_email_status(email, "fail_used")
            raise RuntimeError(ERROR_EMAIL_USED_PROMPT)
    except RuntimeError as e:
        if str(e) == ERROR_EMAIL_USED_PROMPT:
            raise
    except Exception:
        pass

    # Slider
    if logger:
        logger("步骤: 等待并通过滑块")
    if stop_event and stop_event.is_set():
        raise RuntimeError(ERROR_STOPPED)

    code_input_el_xpath = xpaths.get("code_url_element")
    if not code_input_el_xpath:
        raise RuntimeError("code_input_xpath_missing")

    t_slider = time.time()
    slider_ok = False
    slider_iframe_xpath = xpaths.get("slider_iframe")
    slider_container_xpath = xpaths.get("slider_container")

    for attempt in range(MAX_SLIDER_RETRIES):
        if stop_event and stop_event.is_set():
            raise RuntimeError(ERROR_STOPPED)
        if attempt > 0 and logger:
            logger(f"Slider: 重试 {attempt + 1}/{MAX_SLIDER_RETRIES}")

        if solve_slider(
            driver, xpaths, timeout_ms, poll_ms, logger=logger, stop_event=stop_event
        ):
            slider_ok = True
            break

        if attempt < MAX_SLIDER_RETRIES - 1:
            try:
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                t_wait = time.time()
                while time.time() - t_wait < 6:
                    if stop_event and stop_event.is_set():
                        raise RuntimeError(ERROR_STOPPED)
                    if element_visible(driver, code_input_el_xpath, 400, poll_ms):
                        break
                    if (
                        slider_iframe_xpath
                        and element_visible(driver, slider_iframe_xpath, 400, poll_ms)
                    ) or (
                        slider_container_xpath
                        and element_visible(
                            driver, slider_container_xpath, 400, poll_ms
                        )
                    ):
                        break
                    time.sleep(0.3)
            except Exception:
                pass

    if not slider_ok:
        raise RuntimeError(ERROR_SLIDER_FAILED)
    if logger:
        logger(f"步骤: 滑块通过 (耗时 {time.time() - t_slider:.2f}s)")

    if logger:
        logger("步骤: 滑块通过，强制等待 5 秒...")
    time.sleep(1)
    if logger:
        logger("步骤: 滑块通过，立即跳转接码页等待验证码")


def step_confirm(
    ctx: Dict[str, Any],
    row: Dict[str, Any],
    xpaths: Dict[str, str],
    timeout_ms: int,
    poll_ms: int,
    logger: Optional[Callable[[str], None]],
    stop_event: Optional[threading.Event],
    target_check: Optional[Callable[[], bool]],
    email_pool: Optional[Any],
) -> None:
    """
    Step 3: Confirm and submit registration.

    Extracts verification code and submits the form.
    """
    driver = ctx["driver"]

    if stop_event and stop_event.is_set():
        raise RuntimeError(ERROR_STOPPED)
    if target_check and target_check():
        if logger:
            logger("终止接码：已达到目标注册数量")
        raise RuntimeError(ERROR_TARGET_REACHED)

    default_resend_xp = RESEND_CODE_XPATH_PATTERNS[0]
    resend_xp = xpaths.get("resend_code") or default_resend_xp

    # Prepare Credentials
    auth_code_val = (
        row.get("auth_code")
        or row.get("授权码")
        or row.get("authCode")
        or row.get("code_url")
    )
    pwd_val = row.get("password") or row.get("密码")
    password_for_imap = str(auth_code_val or pwd_val or "").strip()

    if not password_for_imap:
        if logger:
            logger("❌ 错误: 缺少密码/授权码，无法获取验证码。")
        raise RuntimeError(ERROR_MISSING_CREDENTIALS)

    # Proxy for IMAP
    proxy_config = None
    host = str(row.get("host") or row.get("代理IP") or "").strip()
    port = str(row.get("port") or row.get("端口") or "").strip()
    protocol = str(row.get("protocol") or row.get("proxyType") or "socks5").strip()
    proxy_username = str(row.get("proxyUserName") or row.get("用户名") or "").strip()
    proxy_password = str(row.get("proxyPassword") or row.get("密码2") or "").strip()

    if host and port:
        try:
            ptype = socks.SOCKS5
            if protocol.lower() == "http":
                ptype = socks.HTTP
            elif protocol.lower() == "socks4":
                ptype = socks.SOCKS4
            proxy_config = {
                "proxy_type": ptype,
                "addr": host,
                "port": int(port),
                "username": proxy_username,
                "password": proxy_password,
                "rdns": True,
            }
        except (ValueError, TypeError):
            pass

    if logger:
        logger("使用 Unified Captcha 模式获取验证码...")

    # User Request: 3 Retries for Code Extraction (Handled inside extract_verification_code_unified)
    # The extract_verification_code_unified function already implements a retry loop (max_retries=3)
    # with Resend Code clicking. We should not loop here to avoid 3x3=9 attempts.

    if stop_event and stop_event.is_set():
        raise RuntimeError(ERROR_STOPPED)

    code = extract_verification_code_unified(
        driver,
        str(row.get("email") or row.get("账号") or ""),
        password_for_imap,
        resend_xp,
        logger,
        timeout=CODE_EXTRACTION_TIMEOUT_SEC,
        proxy_config=proxy_config,
        stop_event=stop_event,
        email_pool=email_pool,
    )

    if not code:
        # User Request: Mark as problem email and release resources
        if email_pool:
            try:
                email = str(row.get("email") or row.get("账号") or "")
                # Mark as 'problem' with reason
                email_pool.update_email_status(
                    email, "problem", reason="验证码获取超时 (3次重试)"
                )
                ctx["failure_status_set"] = True
                if logger:
                    logger(f"已标记为问题邮箱 (fail_code_timeout): {email}")
            except Exception as e:
                if logger:
                    logger(f"标记问题邮箱失败: {e}")

        if logger:
            logger("步骤: 获取验证码失败(超时)")
        # Diagnosis
        slider_iframe_xpath = xpaths.get("slider_iframe")
        slider_container_xpath = xpaths.get("slider_container")
        try:
            if (
                slider_iframe_xpath
                and element_visible(driver, slider_iframe_xpath, 1000, poll_ms)
            ) or (
                slider_container_xpath
                and element_visible(driver, slider_container_xpath, 1000, poll_ms)
            ):
                if logger:
                    logger("诊断: 滑块验证框重新出现，判定为滑块实际上未通过")
                raise RuntimeError(ERROR_SLIDER_REAPPEARED)
        except RuntimeError as e:
            if str(e) == ERROR_SLIDER_REAPPEARED:
                raise
        except Exception:
            pass
        raise RuntimeError(ERROR_CODE_NOT_FOUND)

    ctx["gate_state"] = "code_received"
    if logger:
        logger("步骤: 获取验证码成功，已切回主窗口")

    try:
        driver.switch_to.default_content()
    except Exception:
        pass

    # Fill Code
    try:
        t0 = time.time()
        if logger:
            logger("正在等待验证码输入框...")
        code_input_el_xpath = xpaths.get("code_url_element")
        input_locators = [
            code_input_el_xpath,
            "//input[@autocomplete='one-time-code']",
            "//input[contains(@placeholder, 'verification') or contains(@placeholder, 'code') or contains(@placeholder, '验证码')]",
            "//input[@type='text' and string-length(@maxlength)='6']",
        ]
        input_locators = list(dict.fromkeys([x for x in input_locators if x]))

        # 直接调用底层的极速并行探测函数，设定 10秒(10000ms) 超时
        found_input_xpath = first_present_xpath(driver, input_locators, 10000, poll_ms)

        if not found_input_xpath:
            if logger:
                logger("验证码输入框未出现(超时)")
            raise RuntimeError(ERROR_CODE_INPUT_NOT_VISIBLE)

        if logger:
            logger(
                f"验证码输入框就绪: {found_input_xpath} (耗时 {time.time() - t0:.2f}s)"
            )
        el = driver.find_element(By.XPATH, found_input_xpath)
        el.clear()
        el.send_keys(code)
    except Exception:
        if logger:
            logger("验证码填写异常")
        raise RuntimeError(ERROR_CODE_INPUT_ERROR)

    if logger:
        logger("步骤: 填写验证码")
    safe_click(driver, xpaths["final_submit_btn"], timeout_ms, poll_ms, logger)

    if logger:
        logger("步骤: 正在多维度校验注册/登录状态 (极速高精模式)...")

    jump_success = False
    # 【效率优化】既然有了精准元素，总宽限期回调到合理的 25 秒即可，无需死等 60 秒
    check_timeout = 25
    end_time = time.time() + check_timeout

    if logger:
        logger("步骤: 正在多维度校验注册/登录状态 (死守积分框模式)...")

    jump_success = False
    check_timeout = 30  # 给足30秒时间等待网页跳转和元素渲染
    end_time = time.time() + check_timeout

    while time.time() < end_time:
        if stop_event and stop_event.is_set():
            raise RuntimeError(ERROR_STOPPED)

        try:
            # === 1. 反向探测: 验证码错误极速熔断 ===
            error_msg_xpaths = [
                "//*[contains(text(), 'incorrect') or contains(text(), '验证码错误') or contains(text(), 'invalid') or contains(text(), 'expired')]",
            ]
            if first_present_xpath(driver, error_msg_xpaths, 200, poll_ms):
                if logger:
                    logger("❌ 验证失败: 捕捉到验证码错误或失效提示")
                break

            # === 2. 核心绝对判定：只认网址跳转 + 积分元素(point-box) ===
            point_box_xpaths = [
                "//div[contains(@class, 'point-box')]",
                "//div[@class='point-box']",
            ]
            curr_url = driver.current_url.lower()

            # 🚨 必须同时满足：网址包含了 /app 且 页面真真切切刷出了积分框！
            if "/app" in curr_url and first_present_xpath(
                driver, point_box_xpaths, 200, poll_ms
            ):
                jump_success = True
                if logger:
                    logger(
                        "✅ 验证通过: 亲眼看到目标URL跳转和专属积分框(point-box)，100%确认成功！"
                    )
                break

            # ⚠️ 删除了所有 Cookie 和 LocalStorage 的判断，绝不让它再“早泄”误判！

        except Exception:
            pass

        time.sleep(1)  # 每秒看一眼，耐心等待网页加载完成

    if logger:
        logger("步骤: 跳转成功，强制等待 3 秒...")
    time.sleep(POST_SUBMIT_WAIT_SEC)

    if email_pool:
        try:
            email_pool.update_email_status(
                str(row.get("email") or row.get("账号") or ""), "submitted"
            )
            if logger:
                logger(
                    f"邮箱 {row.get('email') or row.get('账号')} 已标记为 submitted (注册成功)"
                )
        except Exception as e:
            if logger:
                logger(f"标记邮箱状态失败: {e}")


# =============================================================================
# MAIN REGISTRATION FUNCTION
# =============================================================================


def perform_registration(
    row: Dict[str, Any],
    xpaths: Dict[str, str],
    platform_url: str,
    timeout_ms: int,
    poll_ms: int,
    client: BitBrowserClient,
    logger: Optional[Callable[[str], None]] = None,
    stop_event: Optional[threading.Event] = None,
    target_check: Optional[Callable[[], bool]] = None,
    headless_mode: bool = False,
    udp_enabled: bool = False,
    email_pool: Optional[Any] = None,
    keep_open_on_failure_ms: int = 0,
    allow_hold_on_early_failure: bool = False,
    events: Optional[RegistrationEvents] = None,
    save_traffic: bool = False,  # 新增
) -> Tuple[bool, str]:
    """
    Perform a single registration task using BitBrowser.

    Args:
        row: Dictionary containing email, password, proxy settings, etc.
        xpaths: Dictionary of XPath locators for UI elements
        platform_url: URL of the registration platform
        timeout_ms: Default timeout in milliseconds for element operations
        poll_ms: Polling interval in milliseconds
        client: BitBrowserClient instance
        logger: Optional logging function
        stop_event: Event to signal task cancellation
        target_check: Callable to check if target count is reached
        headless_mode: Whether to run browser in headless mode
        udp_enabled: Whether to enable UDP
        email_pool: Optional EmailPool for email status management
        keep_open_on_failure_ms: Time to keep browser open on failure
        allow_hold_on_early_failure: Whether to hold browser on early failure
        events: RegistrationEvents for callbacks

    Returns:
        Tuple of (success: bool, message: str)
    """
    _logger_orig = logger

    def _log(msg: str, **kwargs):
        # Support level kwarg but ignore it for now or format it
        level = kwargs.get("level", "").upper()
        full_msg = f"[{level}] {msg}" if level else msg

        if _logger_orig:
            # Check if original logger supports kwargs, otherwise just send msg
            try:
                _logger_orig(full_msg)
            except TypeError:
                _logger_orig(full_msg)  # Fallback

        if events and events.on_log:
            events.on_log(full_msg)

    logger = _log

    log_memory_usage(logger)

    email = str(row.get("email") or row.get("账号") or "").strip()

    # Windows Compatibility: Dump Manager & Step Runner
    with WindowsDumpManager(logger, email) as dump_mgr:
        runner = StepRunner(
            logger,
            stop_event,
            max_cpu=DEFAULT_MAX_CPU_PERCENT,
            max_mem=DEFAULT_MAX_MEM_MB,
        )

        # Shared Context for Steps
        ctx: Dict[str, Any] = {
            "driver": None,
            "browser_id": None,
            "gate_state": None,
            "t_attached": 0,
            "open_data": None,
            "debug_tab_opened": False,
            "email_pool": email_pool,
            "udp_enabled": udp_enabled,
            "save_traffic": save_traffic,  # 新增
        }

        # Resolve dynamic settings if they are callables
        current_timeout = timeout_ms() if callable(timeout_ms) else timeout_ms
        current_poll = poll_ms() if callable(poll_ms) else poll_ms

        if logger:
            logger(f"当前任务参数: Timeout={current_timeout}ms, Poll={current_poll}ms")

        # -----------------------------------------------------------------
        # Execution
        # -----------------------------------------------------------------
        result_ok = False
        try:
            runner.run(
                "1. 验证与初始化",
                lambda: step_verify(
                    ctx,
                    row,
                    xpaths,
                    client,
                    platform_url,
                    current_timeout,
                    current_poll,
                    logger,
                    stop_event,
                    target_check,
                    email_pool,
                    headless_mode,
                    udp_enabled,
                ),
                timeout=400,  # 第一阶段：给足约 6.6 分钟让慢速网络加载
            )
            runner.run(
                "2. 填写表单与滑块",
                lambda: step_write(
                    ctx,
                    row,
                    xpaths,
                    platform_url,
                    current_timeout,
                    current_poll,
                    logger,
                    stop_event,
                ),
                timeout=800,  # 第二阶段：给足约 13.3 分钟，彻底解决滑块滑一半被强杀的问题！
            )
            runner.run(
                "3. 确认提交与验证",
                lambda: step_confirm(
                    ctx,
                    row,
                    xpaths,
                    current_timeout,
                    current_poll,
                    logger,
                    stop_event,
                    target_check,
                    email_pool,
                ),
                timeout=800,  # 第三阶段：接码可能要等很久，同样给足约 13.3 分钟
            )
            result_ok = True
            return True, "success"
        except Exception as e:
            result_ok = False
            fail_reason = str(e)
            if logger:
                logger(f"任务执行失败: {fail_reason}")
            return False, fail_reason

        finally:
            driver = ctx.get("driver")
            browser_id = ctx.get("browser_id")
            gate_state = ctx.get("gate_state")
            t_attached = ctx.get("t_attached")

            # Rollback
            # If events.on_failure is provided, let it handle the status update
            if not result_ok and email_pool and not (events and events.on_failure):
                # Check if specific failure status was already set (e.g. fail_code_timeout)
                if not ctx.get("failure_status_set"):
                    try:
                        email_pool.update_email_status(email, "failed")
                        if logger:
                            logger(f"邮箱 {email} 已标记为 failed (注册未完成)")
                    except Exception:
                        pass

            if logger:
                logger(f"正在清理资源: {browser_id}")

            # ============================================================
            # 强制资源清理
            # 规则：
            # 1. 无论成功失败，先退出 Selenium / 关闭浏览器进程
            # 2. 注册失败时，必须删除 BitBrowser Profile
            # 3. 删除失败必须真正重试，绝不假装删除成功
            # ============================================================

            # ---------- 1. 退出 Selenium ----------
            if driver:
                try:
                    try:
                        driver.execute_script(
                            "try{document.activeElement && document.activeElement.blur();}catch(e){}"
                        )
                    except Exception:
                        pass

                    try:
                        driver.get("about:blank")
                    except Exception:
                        pass

                    def _quit_driver():
                        try:
                            driver.quit()
                        except Exception:
                            pass

                    t_quit = threading.Thread(
                        target=_quit_driver,
                        daemon=True,
                    )
                    t_quit.start()
                    t_quit.join(timeout=5.0)

                    if t_quit.is_alive() and logger:
                        logger("⚠️ driver.quit 超时，将继续强制清理浏览器进程")

                except Exception as e:
                    if logger:
                        logger(f"driver.quit error: {e}")

            # ---------- 2. 关闭 BitBrowser ----------
            if browser_id:
                try:
                    client.close_browser(browser_id)
                    if logger:
                        logger(f"已关闭浏览器窗口: {browser_id}")
                except Exception as e:
                    if logger:
                        logger(f"⚠️ 关闭浏览器窗口失败: {browser_id} | {e}")

                # 给 BitBrowser 一点时间释放进程
                time.sleep(1.0)

                # 如果 PID 还活着，强制杀掉
                pid = ctx.get("browser_pid")
                if pid:
                    try:
                        if psutil.pid_exists(pid):
                            if logger:
                                logger(
                                    f"检测到浏览器进程 PID={pid} 仍存在，执行强制清理"
                                )
                            kill_process_tree(pid, logger)
                    except Exception as e:
                        if logger:
                            logger(f"强制清理 PID={pid} 失败: {e}")

            # ---------- 3. 注册失败必须删除 Profile ----------
            if browser_id and not result_ok:
                if logger:
                    logger(f"任务失败，正在删除窗口: {browser_id}")

                delete_success = False
                last_delete_error = None

                for attempt in range(1, MAX_BROWSER_DELETE_RETRIES + 1):
                    try:
                        client.delete_browser(browser_id)

                        delete_success = True

                        if logger:
                            logger(
                                f"✅ 已删除失败窗口: {browser_id} "
                                f"(第 {attempt}/{MAX_BROWSER_DELETE_RETRIES} 次)"
                            )

                        break

                    except Exception as e:
                        last_delete_error = e

                        if logger:
                            logger(
                                f"⚠️ 删除失败窗口失败 "
                                f"{attempt}/{MAX_BROWSER_DELETE_RETRIES}: "
                                f"{browser_id} | {e}"
                            )

                        if attempt < MAX_BROWSER_DELETE_RETRIES:
                            # 429 时不要马上继续轰 BitBrowser API
                            # 依次等待 3、6、9、12 秒
                            wait_sec = attempt * 3

                            if logger:
                                logger(
                                    f"等待 {wait_sec} 秒后重新删除窗口..."
                                )

                            time.sleep(wait_sec)

                if not delete_success:
                    if logger:
                        logger(
                            f"❌ 失败窗口最终删除失败: {browser_id} | "
                            f"{last_delete_error}"
                        )


# =============================================================================
# BATCH PROCESSING
# =============================================================================


def run_batch(
    input_path: str,
    xpaths_path: str,
    platform_url: str,
    base_url: str,
    secret: Optional[str],
    concurrency: int,
    timeout_ms: Any,  # int or Callable
    poll_ms: Any,  # int or Callable
    logger: Optional[Callable[[str], None]] = None,
    stop_event: Optional[threading.Event] = None,
    progress_cb: Optional[Any] = None,
    ip_manager: Optional[Any] = None,
    exhaustion_cb: Optional[Any] = None,
    target_success_count: int = 99999999,
    email_pool: Optional[Any] = None,
    headless_mode: bool = False,
    udp_enabled: bool = False,
    events: Optional[RegistrationEvents] = None,
    save_traffic: bool = False,  # 新增
) -> None:
    """
    Run batch registration tasks.

    Args:
        input_path: Path to input CSV/JSON/XLSX file
        xpaths_path: Path to XPath configuration JSON file
        platform_url: URL of the registration platform
        base_url: BitBrowser API base URL
        secret: BitBrowser API secret
        concurrency: Number of concurrent tasks
        timeout_ms: Timeout in milliseconds (int or callable)
        poll_ms: Polling interval in milliseconds (int or callable)
        logger: Optional logging function
        stop_event: Event to signal cancellation
        progress_cb: Progress callback function
        ip_manager: IP pool manager
        exhaustion_cb: Callback for IP exhaustion
        target_success_count: Target number of successful registrations
        email_pool: Email pool manager
        headless_mode: Run browsers in headless mode
        udp_enabled: Enable UDP
        events: RegistrationEvents for callbacks
    """
    rows = read_rows(input_path)
    if not rows:
        if logger:
            logger("未发现可处理的任务行，直接退出")
        return
    if stop_event is None:
        stop_event = threading.Event()
    with open(xpaths_path, "r", encoding="utf-8") as f:
        xpaths = json.load(f)
    client = BitBrowserClient(base_url, secret)

    def _ping(url: str, timeout: int = 10) -> tuple[bool, str]:
        """增强版ping检查，返回状态和详细信息"""
        try:
            # 优化: 使用 list 接口代替 update/create 来进行健康检查，避免产生残留窗口
            payload = {"page": 0, "pageSize": 1}
            h = client._headers()
            r = requests.post(
                f"{url.rstrip('/')}/browser/list",
                headers=h,
                data=json.dumps(payload),
                timeout=timeout,
            )
            if r.status_code == 200:
                try:
                    data = r.json()
                    if data.get("success") or "data" in data:
                        return True, f"API正常响应"
                    return True, "API响应(可能无数据)"
                except (json.JSONDecodeError, ValueError) as e:
                    return False, f"JSON解析失败: {e}"
            elif r.status_code == 401:
                return False, f"认证失败(401)，请检查secret配置"
            elif r.status_code == 404:
                return False, f"API接口不存在(404)"
            else:
                return False, f"HTTP错误: {r.status_code}"
        except requests.exceptions.ConnectTimeout:
            return False, "连接超时"
        except requests.exceptions.ConnectionError as e:
            return False, f"连接被拒绝: {e}"
        except requests.exceptions.RequestException as e:
            return False, f"请求异常: {e}"

    def _probe_create(url: str, timeout: int = 10) -> tuple[bool, str]:
        """探测「创建窗口」接口是否可用（不会残留窗口）。

        仅 health-check 阶段调用。历史上多次出现 list/open 正常、
        唯有 create 404 的"假正常"：实为比特浏览器客户端/账号侧问题。
        这里提前探一次，把故障挡在正式注册之前，并给出可读的排查指引。
        """
        try:
            h = client._headers()
            r = requests.post(
                f"{url.rstrip('/')}/browser/update",
                headers=h,
                # 故意缺代理方式，触发业务报错而非真正创建窗口，仅用于探测路由是否存在
                data=json.dumps({"name": "__health_probe__", "browserFingerPrint": {}}),
                timeout=timeout,
            )
            if r.status_code == 404:
                return False, (
                    "「创建窗口」接口返回404（路由不存在）—— 通常是比特浏览器客户端"
                    "未登录账号 / 本地API未完整开启 / 客户端版本不支持API创建"
                )
            if r.status_code == 200:
                # 路由存在，可能真创建了测试窗口，立刻清理避免残留
                try:
                    d = r.json().get("data")
                    bid = d.get("id") if isinstance(d, dict) else None
                    if bid:
                        requests.post(
                            f"{url.rstrip('/')}/browser/delete",
                            headers=h,
                            data=json.dumps({"id": bid}),
                            timeout=timeout,
                        )
                except Exception:
                    pass
                return True, "「创建窗口」接口可用"
            # 400/405 等其它状态码都说明路由存在，只是参数问题
            return True, f"「创建窗口」接口存在(HTTP {r.status_code})"
        except requests.exceptions.RequestException as e:
            return False, f"「创建窗口」接口探测异常: {e}"

    if logger:
        logger("开始进行比特浏览器接口连通性检查 (health-check)")

    # 尝试主URL
    ping_ok, ping_msg = _ping(client.base_url, timeout=10)
    if not ping_ok:
        if logger:
            logger(f"主接口 {client.base_url} 检查失败: {ping_msg}")

        # 尝试其他端口
        candidates = [client.base_url] + [
            f"http://127.0.0.1:{p}" for p in (54345, 54346, 54321, 54322, 50325, 55555)
        ]
        found = False
        for c in candidates:
            if c == client.base_url:
                continue
            ping_ok, ping_msg = _ping(c, timeout=5)
            if ping_ok:
                client.base_url = c.rstrip("/")
                if logger:
                    logger(f"已自动切换比特浏览器接口到: {client.base_url}")
                found = True
                break
            else:
                if logger:
                    logger(f"尝试接口 {c}: {ping_msg}")

        if not found:
            if logger:
                logger("=" * 50)
                logger("❌ 比特浏览器接口不可用")
                logger("请检查以下几点：")
                logger("1. 比特浏览器客户端是否已启动")
                logger("2. 是否已开启'启用本地API接口'选项")
                logger("3. 防火墙是否拦截了 54345 端口")
                logger("4. 配置中的API地址是否正确")
                logger("=" * 50)
            return
    else:
        if logger:
            logger(f"✅ 比特浏览器接口连接正常: {client.base_url}")

    # ============================================================
    # 额外探测「创建窗口」接口 —— 根治"假正常"
    # list/open 正常但 create 404 时，注册阶段才会暴露，且会白白消耗
    # 邮箱与IP。这里提前探一次，create 不可用时直接给出排查指引并退出。
    # ============================================================
    create_ok, create_msg = _probe_create(client.base_url, timeout=10)
    if not create_ok:
        if logger:
            logger("=" * 50)
            logger("❌ 比特浏览器「创建窗口」接口不可用")
            logger(create_msg)
            logger("排查步骤：")
            logger("1. 比特浏览器客户端是否已登录账号")
            logger("2. 设置→高级→「本地API接口」是否开启")
            logger("3. 重启比特浏览器客户端后重新运行")
            logger("=" * 50)
        return

    lock = threading.Lock()
    global_success_count = 0

    # 资源预检查: 估算当前 IP 池和邮箱池最大可支持的注册数量
    # 说明: 这里只做告警与日志输出，不改变原有业务决策逻辑
    try:
        ip_stats = ip_manager.get_stats() if ip_manager else None
    except Exception:
        ip_stats = None
    try:
        email_total = len(rows)
    except Exception:
        email_total = 0
    if logger:
        if ip_stats:
            logger(
                f"资源预检查: IP池总数={ip_stats['total_ips']}, 剩余可用名额={ip_stats['remaining_usage_count']}, 邮箱任务数={email_total}, 目标成功数={target_success_count}"
            )
        else:
            logger(
                f"资源预检查: 无IP池(或读取失败)，邮箱任务数={email_total}, 目标成功数={target_success_count}"
            )

    def task(idx: int, r: Dict[str, Any]) -> Tuple[int, bool, str]:
        nonlocal global_success_count
        email = str(r.get("email") or r.get("账号") or "").strip()
        if logger:
            logger(f"任务启动 #{idx + 1}: {email}")

        if stop_event and stop_event.is_set():
            return idx, False, ERROR_STOPPED

        with lock:
            if global_success_count >= target_success_count:
                if logger:
                    logger(
                        f"任务 #{idx + 1} 终止: 已达到目标 ({global_success_count}/{target_success_count})"
                    )
                return idx, False, ERROR_TARGET_REACHED

        email = str(r.get("email") or r.get("账号") or "").strip()

        # Pre-Registration Check (Duplicate/Validation)
        if email_pool:
            is_avail, reason = email_pool.check_email_availability(email)
            if not is_avail:
                if logger:
                    logger(f"跳过 {email}: 邮箱不可用 ({reason})")
                r["status"] = "surplus" if "local_status" in reason else "bad"
                return idx, False, f"email_unavailable_{reason}"

        # Mark as processing
        if email_pool:
            email_pool.update_email_status(email, "processing")

        # IP Allocation Logic
        if ip_manager:
            while True:
                if stop_event and stop_event.is_set():
                    if email_pool:
                        email_pool.update_email_status(email, "stopped")
                    return idx, False, ERROR_STOPPED
                with lock:
                    if global_success_count >= target_success_count:
                        if email_pool:
                            email_pool.update_email_status(email, "skipped")
                        return idx, False, ERROR_TARGET_REACHED

                ip_entry, status = ip_manager.allocate_ip(email)
                if status == "success" and ip_entry:
                    # Inject IP into row
                    r["host"] = ip_entry["host"]
                    r["port"] = ip_entry["port"]
                    r["proxyUserName"] = ip_entry.get("proxyUserName", "")
                    r["proxyPassword"] = ip_entry.get("proxyPassword", "")
                    r["protocol"] = ip_entry.get("protocol", "socks5")
                    if logger:
                        logger(
                            f"已分配IP: {r['host']}:{r['port']} ({r['protocol']}) 给 {email}"
                        )
                    break
                elif status == "email_used":
                    if logger:
                        logger(f"跳过 {email}: 此邮箱已在IP池中使用过")
                    if email_pool:
                        email_pool.update_email_status(email, "skipped_ip_used")
                    return idx, False, "email_already_used_in_pool"
                elif status == "ip_busy":
                    if logger:
                        logger(f"IP资源繁忙 (并发限制)，等待重试: {email}")
                    time.sleep(2)
                    continue
                elif status == "ip_exhausted":
                    if exhaustion_cb:
                        if logger:
                            logger("IP池耗尽，等待用户处理...")
                        action = exhaustion_cb("ip_exhausted", ip_manager.get_stats())
                        if action == "retry":
                            if logger:
                                logger("用户选择重试分配IP")
                            continue
                        else:
                            if logger:
                                logger("用户取消任务")
                            if email_pool:
                                email_pool.update_email_status(email, "cancelled")
                            return idx, False, "ip_exhausted_cancelled"
                    else:
                        if email_pool:
                            email_pool.update_email_status(email, "failed_ip_exhausted")
                        return idx, False, "ip_exhausted_no_handler"
                else:
                    if email_pool:
                        email_pool.update_email_status(
                            email, f"failed_ip_error_{status}"
                        )
                    return idx, False, f"ip_allocation_error: {status}"

        try:
            attempts = 0
            ok = False
            msg = ""
            while attempts < MAX_REGISTRATION_ATTEMPTS and not ok:
                attempts += 1
                ok, msg = perform_registration(
                    r,
                    xpaths,
                    platform_url,
                    timeout_ms,
                    poll_ms,
                    client,
                    logger,
                    stop_event,
                    target_check=lambda: global_success_count >= target_success_count,
                    headless_mode=headless_mode,
                    udp_enabled=udp_enabled,
                    email_pool=email_pool,
                    keep_open_on_failure_ms=0,
                    allow_hold_on_early_failure=False,
                    events=events,
                    save_traffic=save_traffic,  # 新增
                )
                if not ok:
                    if events and events.on_failure:
                        events.on_failure(email, msg)
                    if logger:
                        logger(
                            f"重试 {attempts}/{MAX_REGISTRATION_ATTEMPTS}: {email} 失败原因: {msg}"
                        )

                    # Check if marked as 'problem' to abort retry immediately
                    if email_pool:
                        try:
                            cfg = email_pool.get_email_config(email)
                            if cfg and cfg.get("status") == "problem":
                                if logger:
                                    logger(
                                        f"检测到邮箱 {email} 已标记为问题邮箱，停止重试"
                                    )
                                break
                        except Exception:
                            pass

                    # 自动重试条件：proxy/network 错误 或 超时错误
                    if msg and (
                        "proxy" in msg or "network" in msg or "timed out" in msg.lower()
                    ):
                        if "timed out" in msg.lower():
                            if logger:
                                logger(f"检测到超时，自动释放资源并准备重试...")
                        time.sleep(3)
                        continue
                    else:
                        # 非重试类错误，结束重试
                        break

            if ip_manager and not ok:
                # Release IP on failure
                ip_manager.release_ip(r.get("host"), r.get("port"), email)
                if logger:
                    logger(f"注册失败，已释放IP资源: {email}")

            with lock:
                if ok:
                    if events and events.on_success:
                        events.on_success(email)
                    global_success_count += 1
                    if global_success_count > target_success_count:
                        r["status"] = "surplus"
                        msg = "Excess registration (target exceeded)"
                        if logger:
                            logger(
                                f"⚠️ 超量注册检测: {email} (当前成功数: {global_success_count}, 目标: {target_success_count})"
                            )
                        if email_pool:
                            email_pool.update_email_status(email, "surplus")
                    else:
                        r["status"] = "good"
                        if email_pool:
                            email_pool.update_email_status(email, "success")

                    if global_success_count >= target_success_count:
                        if not stop_event.is_set():
                            stop_event.set()
                            if logger:
                                logger(
                                    f"🛑 终止条件触发: 任务完成目标 ({target_success_count})，已触发停止信号"
                                )
                else:
                    r["status"] = "fail"
                    if email_pool:
                        if msg == "email_already_registered_local":
                            pass
                        elif msg == "email_used_prompt":
                            email_pool.update_email_status(email, "fail_used")
                        else:
                            # Check if email is already marked as 'problem' (e.g., by on_failure callback)
                            # If so, don't reset it to 'new'
                            current_status = None
                            if email_pool:
                                try:
                                    cfg = email_pool.get_email_config(email)
                                    if cfg:
                                        current_status = cfg.get("status")
                                except Exception:
                                    pass

                            if current_status == "problem":
                                if logger:
                                    logger(
                                        f"邮箱 {email} 保持问题邮箱状态，不重置为未使用"
                                    )
                            else:
                                # User requested to reset to unused on failure
                                email_pool.update_email_status(email, "new")
        finally:
            # Check if we need to release IP mapping if task failed
            if ip_manager and "ip_entry" in locals() and ip_entry:
                # If not successful, we should ensure the specific IP binding is released
                # so it doesn't stay 'bound' to this email if the email is going back to 'new' state.
                # release_ip handles the logic of "unbinding" this email from this IP.
                if not ok:
                    try:
                        ip_manager.release_ip(ip_entry["host"], ip_entry["port"], email)
                    except Exception:
                        pass

                # Always release the active count
                ip_manager.release_active_ip(ip_entry["host"], ip_entry["port"])

        r["msg"] = msg
        if events and events.on_finish:
            events.on_finish(email, ok, msg)
        return idx, ok, msg

    rounds = 0

    # Outer Loop: Rounds (Passes through the list)
    while True:
        # --- 优先级 1: 目标数量限制 ---
        with lock:
            if global_success_count >= target_success_count:
                if logger:
                    logger(
                        f"🛑 终止条件触发: 已达到目标注册数量 ({target_success_count})"
                    )
                break

        # --- 优先级 2: 待处理任务 ---
        # 🚨 核心修复：定义“终态”（彻底没救的状态）。包含：成功、坏号(被封)、已被注册、超量、问题邮箱
        terminal_statuses = ["good", "bad", "surplus", "fail_used", "problem"]

        pending_idx = [
            i
            for i, r in enumerate(rows)
            # 只有不在“终态”列表里的任务（比如空白状态、普通的 fail 等），才允许进入下一轮重试
            if str(r.get("status", "")).strip().lower() not in terminal_statuses
        ]

        if not pending_idx:
            if logger:
                logger("🛑 终止条件触发: 所有任务均已处于终态 (无待处理项)，任务结束")
            break

        if stop_event and stop_event.is_set():
            if logger:
                logger("🛑 终止条件触发: 收到外部停止信号")
            break

        if logger:
            logger(f"🚀 开始第 {rounds + 1} 轮 (持续运行直至达到目标)")
            logger(
                f"📊 当前状态: 成功 {global_success_count}/{target_success_count}, 待处理 {len(pending_idx)}"
            )

        # Progress callback start of round
        if progress_cb and callable(progress_cb):
            total = len(rows)
            succ = sum(1 for rr in rows if str(rr.get("status", "")).strip() == "good")
            fail = sum(1 for rr in rows if str(rr.get("status", "")).strip() == "fail")
            progress_cb(total, succ, fail, rounds + 1, 0)

        # Check target count before round processing (Double Check)
        with lock:
            remaining_needed = target_success_count - global_success_count
            if remaining_needed <= 0:
                break

        # --- 流水线持续并发模式 (解决“一波一波跑”的问题) ---
        with lock:
            remaining_needed = target_success_count - global_success_count

        actual_tasks_indices = pending_idx[:remaining_needed]

        if not actual_tasks_indices:
            break

        if logger:
            logger(
                f"🚀 任务队列已载入: 共 {len(actual_tasks_indices)} 个任务, 保持并发数: {concurrency}"
            )

        futures = []
        # 核心：整个轮次只创建一个线程池，设定最大并发数
        with ThreadPoolExecutor(max_workers=concurrency) as ex:
            # 瞬间把所有任务塞进队列，线程池会自动控制同时运行的数量
            for idx in actual_tasks_indices:
                futures.append(ex.submit(task, idx, rows[idx]))

            # as_completed 会在【任何一个】窗口完成时立刻返回结果
            # 并且线程池会自动从队列里拉取下一个任务去填补空缺窗口！
            for fut in as_completed(futures):
                if stop_event and stop_event.is_set():
                    # 取消队列中还没开始的任务
                    ex.shutdown(wait=False, cancel_futures=True)
                    break

                try:
                    idx_res, ok, msg = fut.result()

                    # 实时刷新 UI 进度
                    if progress_cb and callable(progress_cb):
                        total = len(rows)
                        succ = sum(
                            1
                            for rr in rows
                            if str(rr.get("status", "")).strip() == "good"
                        )
                        fail = sum(
                            1
                            for rr in rows
                            if str(rr.get("status", "")).strip() == "fail"
                        )
                        progress_cb(total, succ, fail, rounds + 1, 0)

                    # 实时检测是否达到目标
                    with lock:
                        if global_success_count >= target_success_count:
                            if not stop_event.is_set():
                                stop_event.set()
                                if logger:
                                    logger(
                                        f"🛑 终止条件触发: 已达到目标 ({target_success_count})"
                                    )
                            ex.shutdown(wait=False, cancel_futures=True)
                            break
                except Exception:
                    pass
        # --------------------------------------------------------

        write_rows_csv(input_path, rows)
        rounds += 1

        if logger:
            logger(f"🏁 第 {rounds} 轮结束")

        # Check if we should continue to next round
        # The loop condition at the top will handle max_rounds and target checks.
        # We just need to break if stop_event is set.
        if stop_event.is_set():
            break

        time.sleep(1)  # Brief pause between rounds


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================


def main() -> None:
    """Main entry point for command-line execution."""
    import argparse

    p = argparse.ArgumentParser()
    base_dir = os.path.dirname(__file__)
    p.add_argument("--input", default=os.path.join(base_dir, "kl-mail.csv"))
    p.add_argument("--xpaths", default=os.path.join(base_dir, "kling_xpaths.json"))
    p.add_argument("--platform-url", default="https://klingai.com/global")
    p.add_argument("--bitbrowser-url", default="http://127.0.0.1:54345")
    p.add_argument("--bitbrowser-secret", default=os.environ.get("BITBROWSER_SECRET"))
    p.add_argument("--concurrency", type=int, default=1)
    p.add_argument("--timeout-ms", type=int, default=DEFAULT_TIMEOUT_MS)
    p.add_argument("--poll-ms", type=int, default=DEFAULT_POLL_MS)
    args = p.parse_args()

    def stdout_logger(msg: str) -> None:
        print(msg)

    run_batch(
        args.input,
        args.xpaths,
        args.platform_url,
        args.bitbrowser_url,
        args.bitbrowser_secret,
        args.concurrency,
        args.timeout_ms,
        args.poll_ms,
        logger=stdout_logger,
    )


if __name__ == "__main__":
    main()
