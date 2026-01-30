import json
import os
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, List, Optional, Tuple, Callable

import requests
import urllib.parse
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class BitBrowserClient:
    def __init__(self, base_url: str, secret: Optional[str] = None):
        self.base_url = base_url.rstrip('/')
        self.secret = secret

    def _headers(self) -> Dict[str, str]:
        h = {'Content-Type': 'application/json', 'Accept': 'application/json'}
        if self.secret:
            h['Authorization'] = self.secret
        return h

    def _request(self, method: str, endpoint: str, data: Optional[Dict[str, Any]] = None, timeout: int = 30) -> requests.Response:
        url = f"{self.base_url}{endpoint}"
        try:
            if method.upper() == 'POST':
                r = requests.post(url, headers=self._headers(), data=json.dumps(data) if data else None, timeout=timeout)
            else:
                r = requests.get(url, headers=self._headers(), timeout=timeout)
            
            if r.status_code == 405:
                print(f"HTTP 405 Method Not Allowed: {method} {url}")
            r.raise_for_status()
            return r
        except Exception as e:
            if isinstance(e, requests.exceptions.HTTPError) and e.response.status_code == 405:
                print(f"HTTP 405 Error Details: {method} {url} - {e.response.text}")
            raise e

    def update_browser(self, name: str, proxy: Optional[Dict[str, Any]] = None, enable_udp: bool = False) -> str:
        payload: Dict[str, Any] = {
            'name': name,
            'remark': '',
            'proxyMethod': 2,
            'proxyType': 'noproxy',
            'isUDP': 1 if enable_udp else 0,
            'syncTabs': True,
            'syncCookies': True,
            'syncLocalStorage': True,
            'syncIndexedDb': True,
            'browserFingerPrint': {
                'coreVersion': '124'
            }
        }
        if proxy:
            payload.update(proxy)
        r = self._request('POST', '/browser/update', payload, timeout=30)
        data = r.json()
        d = data.get('data')
        bid = None
        if isinstance(d, dict):
            bid = d.get('id')
        elif isinstance(d, str):
            bid = d
        if not bid:
            bid = data.get('id')
        if not bid:
            raise RuntimeError(f"update_browser failed: {data}")
        return bid

    def create_browser(self, name: str, proxy: Optional[Dict[str, Any]] = None, enable_udp: bool = False) -> str:
        payload: Dict[str, Any] = {
            'name': name,
            'remark': '',
            'proxyMethod': 2,
            'proxyType': 'noproxy',
            'isUDP': 1 if enable_udp else 0,
            'syncTabs': True,
            'syncCookies': True,
            'syncLocalStorage': True,
            'syncIndexedDb': True,
            'browserFingerPrint': {
                'coreVersion': '124'
            }
        }
        if proxy:
            payload.update(proxy)
        r = self._request('POST', '/browser/create', payload, timeout=30)
        data = r.json()
        d = data.get('data')
        bid = None
        if isinstance(d, dict):
            bid = d.get('id')
        elif isinstance(d, str):
            bid = d
        if not bid:
            bid = data.get('id')
        if not bid:
            raise RuntimeError(f"create_browser failed: {data}")
        return bid

    def open_browser(self, browser_id: str) -> Dict[str, Any]:
        payload = {'id': browser_id}
        r = self._request('POST', '/browser/open', payload, timeout=60)
        data = r.json()
        d = data.get('data')
        if isinstance(d, dict):
            return d
        return {}

    def close_browser(self, browser_id: str) -> None:
        try:
            self._request('POST', '/browser/close', {'id': browser_id}, timeout=30)
        except Exception:
            pass

    def delete_browser(self, browser_id: str) -> None:
        try:
            self._request('POST', '/browser/delete', {'id': browser_id}, timeout=30)
        except Exception:
            pass

    def detail_browser(self, browser_id: str) -> Dict[str, Any]:
        try:
            r = self._request('POST', '/browser/detail', {'id': browser_id}, timeout=30)
            data = r.json()
            d = data.get('data')
            if isinstance(d, dict):
                return d
            return {}
        except Exception:
            return {}


def read_rows(input_path: str) -> List[Dict[str, Any]]:
    if input_path.lower().endswith('.csv'):
        import csv
        rows: List[Dict[str, Any]] = []
        with open(input_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        return rows
    if input_path.lower().endswith('.json'):
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            if isinstance(data, dict) and 'rows' in data:
                return data['rows']
        return []
    if input_path.lower().endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(input_path)
            sheet = wb.active
            headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
            rows = []
            for r in sheet.iter_rows(min_row=2):
                row = {headers[i]: (r[i].value if i < len(r) else None) for i in range(len(headers))}
                rows.append(row)
            return rows
        except Exception:
            return []
    return []


def write_rows_csv(input_path: str, rows: List[Dict[str, Any]]) -> None:
    import csv
    if not rows:
        return
    # Fix: Collect all unique keys from all rows to ensure fieldnames cover everything
    headers = set()
    for r in rows:
        headers.update(r.keys())
    
    # Sort headers for consistency, prioritize common fields
    header_list = sorted(list(headers))
    priority_fields = ['email', 'password', 'status', 'host', 'port', 'msg']
    for f in reversed(priority_fields):
        if f in header_list:
            header_list.remove(f)
            header_list.insert(0, f)

    with open(input_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=header_list)
        writer.writeheader()
        for r in rows:
            # Ensure row has all keys (fill missing with None or empty)
            # DictWriter handles missing keys by ignoring them if not in fieldnames (but we ensure they are)
            # or putting '' if restval is set. Default raises error if key in dict but not in fieldnames.
            # We already covered all keys.
            writer.writerow(r)


def element_exists(driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int) -> bool:
    try:
        eff_timeout = min(timeout_ms, 60000)
        eff_poll = max(0.2, poll_ms / 1000.0)
        WebDriverWait(driver, eff_timeout / 1000.0, poll_frequency=eff_poll).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return True
    except Exception:
        return False


def element_visible(driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int) -> bool:
    try:
        eff_timeout = min(timeout_ms, 60000)
        eff_poll = max(0.2, poll_ms / 1000.0)
        WebDriverWait(driver, eff_timeout / 1000.0, poll_frequency=eff_poll).until(EC.visibility_of_element_located((By.XPATH, xpath)))
        return True
    except Exception:
        return False

def element_exists_visible(driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int) -> bool:
    if not element_exists(driver, xpath, timeout_ms, poll_ms):
        return False
    return element_visible(driver, xpath, timeout_ms, poll_ms)

def wait_page_ready(driver: webdriver.Remote, timeout_sec: int = 12) -> bool:
    try:
        WebDriverWait(driver, timeout_sec).until(lambda d: d.execute_script("return document.readyState") in ("interactive", "complete"))
        return True
    except Exception:
        return False

def log_page_timing(driver: webdriver.Remote, logger: Optional[Any]) -> None:
    if not logger:
        return
    try:
        data = driver.execute_script("""
            try {
              var t = performance.timing || {};
              var navs = performance.getEntriesByType && performance.getEntriesByType('navigation');
              var nav = (navs && navs.length) ? navs[0] : null;
              var ns = nav ? nav.startTime : (t.navigationStart || 0);
              var dcl = nav ? nav.domContentLoadedEventEnd : (t.domContentLoadedEventEnd || 0);
              var le = nav ? nav.loadEventEnd : (t.loadEventEnd || 0);
              return {
                readyState: document.readyState,
                navigationStart: ns,
                domContentLoadedEventEnd: dcl,
                loadEventEnd: le
              };
            } catch (e) {
              return { readyState: document.readyState };
            }
        """)
        ns = float(data.get('navigationStart') or 0.0)
        dcl = float(data.get('domContentLoadedEventEnd') or 0.0)
        le = float(data.get('loadEventEnd') or 0.0)
        rs = data.get('readyState')
        def fmt(ms: float) -> str:
            return f"{int(ms)}ms" if ms > 0 else "n/a"
        logger(f"PageTiming: readyState={rs}, domContentLoaded={fmt(dcl - ns)}, loadComplete={fmt(le - ns)}")
    except Exception:
        pass

def log_resource_status(driver: webdriver.Remote, logger: Optional[Any], limit: int = 200) -> None:
    if not logger:
        return
    try:
        data = driver.execute_script("""
            try {
              var es = performance.getEntries ? performance.getEntries() : [];
              var out = {};
              for (var i=0;i<es.length;i++){
                var e = es[i];
                var t = e.entryType || e.initiatorType || 'unknown';
                out[t] = (out[t]||0) + 1;
              }
              return {counts: out, total: es.length};
            } catch (e) { return {error: ''+e}; }
        """)
        counts = data.get('counts') or {}
        total = data.get('total') or 0
        try:
            logger(f"Resources: total={total}, counts={json.dumps(counts, ensure_ascii=False)}")
        except Exception:
            logger(f"Resources: total={total}")
    except Exception:
        pass

def log_resource_phase_timings(driver: webdriver.Remote, logger: Optional[Any], contains: Optional[str] = None, limit: int = 30) -> None:
    if not logger:
        return
    try:
        data = driver.execute_script("""
            try {
              var es = performance.getEntriesByType ? performance.getEntriesByType('resource') : [];
              var out = [];
              var kw = arguments[0] || '';
              var lim = arguments[1] || 30;
              for (var i=0;i<es.length;i++){
                var e = es[i];
                if (kw && e.name && e.name.indexOf(kw) === -1) continue;
                out.push({
                  name: e.name,
                  initiatorType: e.initiatorType,
                  duration: e.duration,
                  dns: (e.domainLookupEnd && e.domainLookupStart) ? (e.domainLookupEnd - e.domainLookupStart) : 0,
                  tcp: (e.connectEnd && e.connectStart) ? (e.connectEnd - e.connectStart) : 0,
                  ssl: (e.connectEnd && e.secureConnectionStart) ? (e.connectEnd - e.secureConnectionStart) : 0,
                  ttfb: (e.responseStart && e.requestStart) ? (e.responseStart - e.requestStart) : 0,
                  download: (e.responseEnd && e.responseStart) ? (e.responseEnd - e.responseStart) : 0
                });
              }
              if (out.length > lim) out = out.slice(out.length - lim);
              return out;
            } catch (e) { return []; }
        """, contains or "", limit)
        if not isinstance(data, list) or not data:
            return
        key = contains or "all"
        logger(f"ResourceTiming[{key}] count={len(data)}")
        for e in data:
            try:
                name = str(e.get('name') or '')
                it = str(e.get('initiatorType') or '')
                dur = float(e.get('duration') or 0.0)
                dns = float(e.get('dns') or 0.0)
                tcp = float(e.get('tcp') or 0.0)
                ssl = float(e.get('ssl') or 0.0)
                ttfb = float(e.get('ttfb') or 0.0)
                dl = float(e.get('download') or 0.0)
                logger(f"Timing {it} {int(dur)}ms dns={int(dns)} tcp={int(tcp)} ssl={int(ssl)} ttfb={int(ttfb)} dl={int(dl)} url={name}")
            except Exception:
                pass
    except Exception:
        pass

def ensure_artifact_dir() -> str:
    base = os.path.dirname(__file__)
    p = os.path.join(base, 'test_artifacts')
    try:
        os.makedirs(p, exist_ok=True)
    except Exception:
        pass
    return p

def take_screenshot(driver: webdriver.Remote, name: str, logger: Optional[Any] = None) -> None:
    d = ensure_artifact_dir()
    path = os.path.join(d, name)
    try:
        driver.save_screenshot(path)
        if logger:
            logger(f"Screenshot: {path}")
    except Exception as e:
        if logger:
            logger(f"Screenshot error: {e}")

def write_artifact_json(name: str, data: Any, logger: Optional[Any] = None) -> None:
    d = ensure_artifact_dir()
    path = os.path.join(d, name)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        if logger:
            logger(f"Artifact: {path}")
    except Exception as e:
        if logger:
            logger(f"Artifact error: {e}")

def check_connectivity(driver: webdriver.Remote, url: str, logger: Optional[Any], max_wait: int = 20) -> bool:
    ok = True
    try:
        try:
            driver.get(url)
        except Exception:
            driver.execute_script("window.location.href=arguments[0];", url)
        wait_page_ready(driver, max_wait)
        log_page_timing(driver, logger)
        log_resource_status(driver, logger)
        try:
            init_handles = driver.window_handles
            if logger:
                logger(f"初始窗口数量: {len(init_handles)}")
            if len(init_handles) > 1 and (not udp_enabled):
                try:
                    driver.switch_to.window(init_handles[0])
                    if logger:
                        logger("检测到初始阶段存在额外标签，已切回主标签")
                except Exception:
                    pass
        except Exception:
            pass
        take_screenshot(driver, "bitbrowser_initial.png", logger)
    except Exception as e:
        ok = False
        if logger:
            logger(f"Connectivity nav error: {e}")
    return ok

def xpath_count(driver: webdriver.Remote, xpath: str) -> int:
    try:
        return int(driver.execute_script("""
            try {
              var r = document.evaluate(arguments[0], document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
              return r ? r.snapshotLength : 0;
            } catch (e) { return -1; }
        """, xpath))
    except Exception:
        return -1

def safe_click(driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int, logger: Optional[Any] = None, retries: int = 2) -> bool:
    for i in range(max(1, retries)):
        try:
            WebDriverWait(driver, min(timeout_ms, 12000) / 1000.0, poll_frequency=max(0.2, poll_ms/1000.0)).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
            return True
        except Exception:
            try:
                el = WebDriverWait(driver, min(timeout_ms, 12000) / 1000.0, poll_frequency=max(0.2, poll_ms/1000.0)).until(EC.presence_of_element_located((By.XPATH, xpath)))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                try:
                    el.click()
                    return True
                except Exception:
                    driver.execute_script("arguments[0].click();", el)
                    return True
            except Exception as e:
                if logger:
                    logger(f"点击失败重试 {i+1}/{retries}: {e}")
        time.sleep(0.5)
    return False

def first_present_xpath(driver: webdriver.Remote, xpaths: List[Optional[str]], timeout_ms: int, poll_ms: int) -> Optional[str]:
    for xp in xpaths:
        if not xp:
            continue
        try:
            if element_exists_visible(driver, xp, min(timeout_ms, 12000), poll_ms):
                return xp
        except Exception:
            pass
    return None

def safe_click_any(driver: webdriver.Remote, xpaths: List[Optional[str]], timeout_ms: int, poll_ms: int, logger: Optional[Any] = None, retries: int = 2) -> bool:
    xp = first_present_xpath(driver, xpaths, timeout_ms, poll_ms)
    if not xp:
        return False
    return safe_click(driver, xp, timeout_ms, poll_ms, logger, retries)

def js_click_xpath(driver: webdriver.Remote, xpath: str) -> bool:
    try:
        r = driver.execute_script("""
          var xp = arguments[0];
          try {
            var res = document.evaluate(xp, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            var el = res && res.singleNodeValue;
            if (el) {
              try { el.scrollIntoView({block:'center'}); } catch(e) {}
              try { el.click(); return true; } catch(e) {}
              try { el.dispatchEvent(new MouseEvent('click', {bubbles:true})); return true; } catch(e) {}
            }
          } catch (e) {}
          return false;
        """, xpath)
        return bool(r)
    except Exception:
        return False

def safe_send_keys(driver: webdriver.Remote, xpath: str, text: str, timeout_ms: int, poll_ms: int, logger: Optional[Any] = None, retries: int = 1) -> bool:
    for i in range(max(1, retries)):
        try:
            el = WebDriverWait(driver, min(timeout_ms, 12000) / 1000.0, poll_frequency=max(0.2, poll_ms/1000.0)).until(EC.presence_of_element_located((By.XPATH, xpath)))
            el.send_keys(text)
            return True
        except Exception as e:
            if logger:
                logger(f"输入失败重试 {i+1}/{retries}: {e}")
        time.sleep(0.5)
    return False

def safe_send_keys_any(driver: webdriver.Remote, xpaths: List[Optional[str]], text: str, timeout_ms: int, poll_ms: int, logger: Optional[Any] = None, retries: int = 1) -> bool:
    for xp in xpaths:
        if not xp:
            continue
        if safe_send_keys(driver, xp, text, timeout_ms, poll_ms, logger, retries):
            return True
        try:
            el = WebDriverWait(driver, min(timeout_ms, 12000) / 1000.0, poll_frequency=max(0.2, poll_ms/1000.0)).until(EC.presence_of_element_located((By.XPATH, xp)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            try:
                el.send_keys(text)
                return True
            except Exception:
                pass
        except Exception:
            pass
    return False

def validate_xpaths(driver: webdriver.Remote, xpaths: Dict[str, str], logger: Optional[Any], keys: List[str], timeout_ms: int, poll_ms: int) -> Dict[str, bool]:
    result: Dict[str, bool] = {}
    for k in keys:
        xp = xpaths.get(k) or ""
        ok = bool(xp) and element_exists_visible(driver, xp, min(timeout_ms, 8000), poll_ms)
        result[k] = ok
        if logger:
            logger(f"XPath验证 {k}: {'OK' if ok else 'FAIL'}")
    return result

def find_click(driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int) -> None:
    eff_timeout = min(timeout_ms, 60000)
    WebDriverWait(driver, eff_timeout / 1000.0, poll_frequency=poll_ms / 1000.0).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()


def find_click_any(driver: webdriver.Remote, xpath: str, timeout_ms: int, poll_ms: int) -> None:
    try:
        eff_timeout = min(timeout_ms, 60000)
        WebDriverWait(driver, eff_timeout / 1000.0, poll_frequency=poll_ms / 1000.0).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        return
    except Exception:
        eff_timeout = min(timeout_ms, 60000)
        el = WebDriverWait(driver, eff_timeout / 1000.0, poll_frequency=poll_ms / 1000.0).until(EC.presence_of_element_located((By.XPATH, xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
        try:
            el.click()
        except Exception:
            driver.execute_script("arguments[0].click();", el)


def _human_drag_track(distance: int, duration_ms: int = 900, jitter_px: int = 1, overshoot_px: int = 5) -> List[int]:
    steps = 15
    arr: List[int] = []
    moved = 0
    for i in range(steps):
        remain = distance - moved
        if remain <= 0:
            break
        prog = i / steps
        base = int((1 - (prog - 0.5) ** 2) * 8) + 1
        jitter = ((i % 3) - 1) * jitter_px
        dx = max(1, base + jitter)
        moved += dx
        arr.append(dx)
    while sum(arr) < distance:
        arr.append(1)
    arr.append(overshoot_px)
    arr.append(-overshoot_px + 1)
    return arr


_SLIDER_PASS_CACHE: Dict[str, Dict[str, Any]] = {}


def solve_slider(driver: webdriver.Remote, xpaths: Dict[str, str], timeout_ms: int, poll_ms: int, logger: Optional[Any] = None) -> bool:
    t0 = time.time()
    ok_xpath = xpaths.get('code_url_element') or xpaths.get('next_btn') or xpaths.get('password_input')

    def _slider_container_visible(short_timeout_ms: int = 600) -> bool:
        try:
            try:
                driver.switch_to.default_content()
            except Exception:
                pass
            if element_visible(driver, xpaths['slider_container'], short_timeout_ms, poll_ms):
                return True
        except Exception:
            pass
        try:
            try:
                driver.switch_to.default_content()
            except Exception:
                pass
            if element_exists(driver, xpaths['slider_iframe'], short_timeout_ms, poll_ms):
                try:
                    iframe = driver.find_element(By.XPATH, xpaths['slider_iframe'])
                    driver.switch_to.frame(iframe)
                except Exception:
                    return False
                try:
                    return element_visible(driver, xpaths['slider_container'], short_timeout_ms, poll_ms)
                finally:
                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass
        except Exception:
            pass
        return False

    def _ok_visible(short_timeout_ms: int) -> bool:
        if not ok_xpath:
            return False
        try:
            try:
                driver.switch_to.default_content()
            except Exception:
                pass
            return element_visible(driver, ok_xpath, short_timeout_ms, poll_ms)
        except Exception:
            return False

    try:
        if _ok_visible(min(1500, timeout_ms)):
            if logger:
                logger("Slider: 已处于后续步骤，跳过滑块等待")
            return True
    except Exception:
        pass

    cache_key = ""
    domain_filter = None
    try:
        sid = str(getattr(driver, 'session_id', '') or '')
        cur = ''
        try:
            cur = driver.current_url or ''
        except Exception:
            cur = ''
        host = ''
        try:
            host = (urllib.parse.urlparse(cur).netloc or '').strip()
        except Exception:
            host = ''
        domain_filter = host or None
        cache_key = f"{sid}|{host}"
    except Exception:
        cache_key = ""

    try:
        cached = _SLIDER_PASS_CACHE.get(cache_key) if cache_key else None
        if cached and (time.time() - float(cached.get('t') or 0.0) < 300.0):
            if not _slider_container_visible(600):
                if logger:
                    logger("Slider: 命中通过缓存，且当前未检测到滑块")
                if cached.get('ok') is True:
                    return True
    except Exception:
        pass

    appear_wait_ms = min(max(timeout_ms, 15000), 60000)
    start = time.time()
    while (time.time() - start) * 1000 < appear_wait_ms:
        try:
            if _ok_visible(800):
                if logger:
                    logger("Slider: 等待期间检测到后续步骤，视为通过")
                if cache_key:
                    _SLIDER_PASS_CACHE[cache_key] = {'ok': True, 't': time.time()}
                return True
        except Exception:
            pass
        if element_visible(driver, xpaths['slider_iframe'], 800, poll_ms) or element_visible(driver, xpaths['slider_container'], 800, poll_ms):
            break
        time.sleep(max(0.2, poll_ms / 1000.0))
    else:
        if logger:
            logger(f"Slider: {int(appear_wait_ms)}ms 内未出现滑块")
            log_page_timing(driver, logger)
            log_resource_phase_timings(driver, logger, domain_filter, 30)
        return False
    try:
        iframe_probe_ms = min(timeout_ms, 6000)
        if element_exists(driver, xpaths['slider_iframe'], iframe_probe_ms, poll_ms):
            iframe = driver.find_element(By.XPATH, xpaths['slider_iframe'])
            driver.switch_to.frame(iframe)
        for i in range(10):
            # 每次尝试重新定位容器与句柄，避免刷新后引用失效
            try:
                # 确保位于正确的文档或iframe中
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                if element_exists(driver, xpaths['slider_iframe'], 1500, poll_ms):
                    try:
                        iframe = driver.find_element(By.XPATH, xpaths['slider_iframe'])
                        driver.switch_to.frame(iframe)
                    except Exception:
                        pass
                container = None
                for xp in [xpaths['slider_container'], "//*[contains(@class,'slider-shadow')]", "//*[contains(@class,'kwai-captcha-slider-wrapper')]"]:
                    try:
                        container = driver.find_element(By.XPATH, xp)
                        if container:
                            break
                    except Exception:
                        pass
                if not container:
                    time.sleep(1.5)  # 找不到容器时增加等待
                    continue
                handle_wait_ms = min(timeout_ms, 4000)
                handle = None
                handle_candidates = [xpaths['slider_handle'], "//*[contains(@class,'slider-btn')]", "//*[contains(@class,'btn-icon')]"]
                # 显式等待滑块句柄出现
                try:
                    WebDriverWait(driver, 5).until(lambda d: any(element_exists(d, hx, 500, poll_ms) for hx in handle_candidates))
                except Exception:
                    pass

                for hx in handle_candidates:
                    try:
                        handle = WebDriverWait(driver, handle_wait_ms / 1000.0, poll_frequency=poll_ms / 1000.0).until(EC.presence_of_element_located((By.XPATH, hx)))
                        if handle:
                            break
                    except Exception:
                        handle = None
                try:
                    width = driver.execute_script("return Math.floor(arguments[0].getBoundingClientRect().width)||arguments[0].offsetWidth||200;", container)
                except Exception:
                    width = container.size.get('width') or 200
                handle_w = 24
                try:
                    if handle is not None:
                        hw = handle.size.get('width')
                        if hw:
                            handle_w = hw
                except Exception:
                    pass
                dist = 238
            except Exception:
                dist = 238
            try:
                actions = ActionChains(driver)
                if handle is not None:
                    actions.move_to_element(handle).pause(0.01).move_by_offset(2, 0).click_and_hold(handle).pause(0.01)
                else:
                    h = container.size.get('height') or 20
                    actions.move_to_element_with_offset(container, 5, int(h/2)).pause(0.01).click_and_hold().pause(0.01)
                # Faster slider: reduce steps significantly
                steps = max(2, int(dist / 100))
                step_len = max(40, int(dist / steps))
                moved = 0
                # Use a single action chain for smoother and faster movement
                for _ in range(steps):
                    left = dist - moved
                    dx = min(step_len, left)
                    actions.move_by_offset(dx, 0)
                    moved += dx
                actions.release().perform()
                
                if _ok_visible(1500):
                    if cache_key:
                        _SLIDER_PASS_CACHE[cache_key] = {'ok': True, 't': time.time()}
                    if logger:
                        logger(f"Slider: 通过(检测到后续元素)，耗时 {time.time() - t0:.2f}s")
                    return True

                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass

                if not _slider_container_visible(600):
                    if _ok_visible(1200):
                        if cache_key:
                            _SLIDER_PASS_CACHE[cache_key] = {'ok': True, 't': time.time()}
                        if logger:
                            logger(f"Slider: 通过(检测到后续元素)，耗时 {time.time() - t0:.2f}s")
                        return True
                    
                    # 优化: 滑块消失且一段时间未重现，视为通过，避免死等后续元素导致重试
                    time.sleep(1.0)
                    if not _slider_container_visible(500):
                         # 二次确认: 等待 3s 确保不是因刷新导致的短暂消失
                         time.sleep(3.0)
                         if not _slider_container_visible(500):
                             if logger:
                                 logger(f"Slider: 通过(滑块消失并确认)，耗时 {time.time() - t0:.2f}s")
                             if cache_key:
                                _SLIDER_PASS_CACHE[cache_key] = {'ok': True, 't': time.time()}
                             return True

                    time.sleep(0.4)
                    continue

                time.sleep(0.1)
            except Exception:
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                time.sleep(0.2)
    except Exception:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    if _ok_visible(min(5000, timeout_ms)):
        if cache_key:
            _SLIDER_PASS_CACHE[cache_key] = {'ok': True, 't': time.time()}
        if logger:
            logger(f"Slider: 通过(最终兜底检测)，耗时 {time.time() - t0:.2f}s")
        return True
    if logger:
        logger(f"Slider: 失败，耗时 {time.time() - t0:.2f}s")
        log_page_timing(driver, logger)
        log_resource_phase_timings(driver, logger, domain_filter, 30)
        log_performance_network(driver, logger, limit=120, domain_filter=domain_filter)
    return False


def extract_code_from_page_text(driver: webdriver.Remote) -> Optional[str]:
    try:
        body_text = driver.find_element(By.TAG_NAME, 'body').text
        m = re.search(r"\b(\d{6})\b", body_text)
        if m:
            return m.group(1)
    except Exception:
        pass
    return None


def extract_code_using_xpath(driver: webdriver.Remote, xpath: str) -> Optional[str]:
    try:
        el = driver.find_element(By.XPATH, xpath)
        txt = el.text or el.get_attribute('value') or ''
        m = re.search(r"\b(\d{6})\b", txt)
        if m:
            return m.group(1)
    except Exception:
        pass
    return None


def wait_extract_code(driver: webdriver.Remote, xpath: Optional[str], max_wait_sec: int = 20, logger: Optional[Any] = None) -> Optional[str]:
    end = time.time() + max_wait_sec
    code: Optional[str] = None
    while time.time() < end:
        if logger:
            logger("等待验证码...")
        if xpath:
            try:
                WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, xpath)))
                code = extract_code_using_xpath(driver, xpath)
                if code:
                    if logger:
                        logger(f"验证码通过XPath提取: {code}")
                    return code
            except Exception:
                pass
        code = extract_code_from_page_text(driver)
        if code:
            if logger:
                logger(f"验证码通过文本提取: {code}")
            return code
        time.sleep(1)
    return None


def extract_code_attempts(driver: webdriver.Remote, xpath: Optional[str], logger: Optional[Any], attempts: int = 3) -> Optional[str]:
    code: Optional[str] = None
    for i in range(max(1, attempts)):
        if logger:
            logger(f"尝试提取验证码 {i+1}/{attempts}")
        try:
            if xpath:
                try:
                    el = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    txt = (el.text or el.get_attribute('value') or '').strip()
                    if logger:
                        logger(f"元素文本: '{txt}'")
                    m = re.search(r"\b(\d{6})\b", txt)
                    if m:
                        code = m.group(1)
                        return code
                    clean = txt.replace(' ', '').replace('\n', '')
                    if clean.isdigit() and len(clean) == 6:
                        code = clean
                        return code
                except Exception as e:
                    if logger:
                        logger(f"元素提取失败: {e}")
            page_txt = ''
            try:
                body = driver.find_element(By.TAG_NAME, 'body')
                page_txt = body.text
            except Exception:
                pass
            if page_txt:
                m2 = re.search(r"\b(\d{6})\b", page_txt)
                if m2:
                    code = m2.group(1)
                    if logger:
                        logger(f"页面文本提取验证码: {code}")
                    return code
        except Exception as e:
            if logger:
                logger(f"提取尝试异常: {e}")
        time.sleep(1 if i < attempts-1 else 0)
    return code


def extract_verification_code_flow(driver: webdriver.Remote, code_url: str, code_xpath: Optional[str], logger: Optional[Any], debugger_http: Optional[str] = None, timeout: int = 60) -> Optional[str]:
    main_handle = driver.current_window_handle
    try:
        # Set shorter page load timeout for verification code page
        driver.set_page_load_timeout(20)
        
        # 立即开始尝试，无需额外等待
        for open_try in range(3):
            t_start = time.time()
            created = False
            prev_len = len(driver.window_handles)
            if debugger_http:
                try:
                    created = create_cdp_tab(debugger_http, code_url, logger)
                except Exception as e:
                    if logger:
                        logger(f"调试接口创建失败: {e}")
            if created:
                try:
                    WebDriverWait(driver, 3).until(lambda d: len(d.window_handles) > prev_len)
                    driver.switch_to.window(driver.window_handles[-1])
                    if logger:
                        try:
                            logger(f"已切换到新标签: {driver.current_url}")
                        except Exception:
                            logger("已切换到新标签")
                except Exception:
                    created = False
            if not created:
                driver.switch_to.new_window('tab')
                if logger:
                    logger(f"打开接码链接: {code_url}")
                try:
                    driver.get(code_url)
                except Exception as nav_err:
                    if logger:
                        logger(f"导航失败: {nav_err}")
                    try:
                        driver.execute_script("window.location.href=arguments[0];", code_url)
                    except Exception:
                        pass
            
            if logger: logger(f"页面打开耗时: {time.time()-t_start:.2f}s")
            
            loaded = False
            try:
                WebDriverWait(driver, 5).until(lambda d: d.execute_script("return document.readyState") in ("interactive", "complete"))
                loaded = True
                if logger:
                    logger("接码页已加载")
            except Exception:
                if logger:
                    logger("接码页加载超时")
            if loaded:
                break
            try:
                driver.refresh()
            except Exception:
                pass
            time.sleep(0.5)
            
        code: Optional[str] = None
        # 使用总体超时控制
        end_time = time.time() + timeout
        
        while time.time() < end_time and not code:
            try:
                if code_xpath:
                    try:
                        el = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, code_xpath)))
                        txt = (el.text or el.get_attribute('value') or '').strip()
                        if logger:
                            logger(f"元素文本: '{txt}'")
                        m = re.search(r"\b(\d{6})\b", txt)
                        if m:
                            code = m.group(1)
                            break
                        clean = txt.replace(' ', '').replace('\n', '')
                        if clean.isdigit() and len(clean) == 6:
                            code = clean
                            break
                    except Exception:
                        pass
                if not code:
                    try:
                        body_txt = driver.find_element(By.TAG_NAME, 'body').text
                        m2 = re.search(r"\b(\d{6})\b", body_txt)
                        if m2:
                            code = m2.group(1)
                            if logger:
                                logger(f"页面文本提取验证码: {code}")
                            break
                    except Exception:
                        pass
            except Exception:
                pass
            
            if not code:
                # 每隔3秒刷新一次 (用户要求)
                if int(time.time()) % 3 == 0:
                    try:
                        driver.refresh()
                        if logger: logger("刷新接码页面(3s间隔)...")
                        WebDriverWait(driver, 5).until(lambda d: d.execute_script("return document.readyState") in ("interactive", "complete"))
                    except Exception:
                        pass
                time.sleep(1)
        
        try:
            if len(driver.window_handles) > 1:
                driver.close()
        except Exception:
            pass
        try:
            driver.switch_to.window(main_handle)
        except Exception:
            pass
        return code
    except Exception as e:
        if logger:
            logger(f"接码流程异常: {e}")
        try:
            if len(driver.window_handles) > 1:
                driver.close()
        except Exception:
            pass
        try:
            driver.switch_to.window(main_handle)
        except Exception:
            pass
        return None


def get_verification_code(driver: webdriver.Remote, code_url: str, code_xpath: Optional[str] = None, retries: int = 3, wait_seconds: int = 5, logger: Optional[Any] = None) -> Optional[str]:
    main_handle = driver.current_window_handle
    for i in range(retries):
        try:
            if logger:
                logger(f"打开接码标签尝试 {i+1}/{retries}: {code_url}")
            driver.switch_to.new_window('tab')
            try:
                driver.get(code_url)
            except Exception as nav_err:
                if logger:
                    logger(f"直接导航失败，尝试JS打开: {nav_err}")
                try:
                    driver.execute_script("window.open(arguments[0], '_blank')", code_url)
                except Exception as js_err:
                    if logger:
                        logger(f"JS window.open 失败: {js_err}")
            time.sleep(1) # Reduced from wait_seconds
            code = wait_extract_code(driver, code_xpath, max_wait_sec=15, logger=logger)
            driver.close()
            driver.switch_to.window(main_handle)
            if code:
                if logger:
                    logger("接码成功，返回主标签")
                return code
        except Exception as e:
            if logger:
                logger(f"接码标签异常: {e}")
            try:
                driver.close()
            except Exception:
                pass
            try:
                driver.switch_to.window(main_handle)
            except Exception:
                pass
        time.sleep(2)
    return None


def open_attached_driver(open_data: Dict[str, Any]) -> webdriver.Chrome:
    driver_path = open_data.get('driver')
    debugger_address = open_data.get('http')
    if not driver_path or not debugger_address:
        raise RuntimeError(f"No driver/http returned: {open_data}")
    from selenium.webdriver.chrome.service import Service
    options = webdriver.ChromeOptions()
    options.debugger_address = debugger_address
    try:
        options.page_load_strategy = 'none'
    except Exception:
        pass
    try:
        options.set_capability('goog:loggingPrefs', {'browser': 'ALL', 'performance': 'ALL'})
    except Exception:
        pass
    service = Service(executable_path=driver_path)
    driver = webdriver.Chrome(service=service, options=options)
    try:
        driver.set_page_load_timeout(25)
    except Exception:
        pass
    try:
        driver.implicitly_wait(0)
    except Exception:
        pass
    return driver


def create_cdp_tab(debugger_address: str, url: str, logger: Optional[Any] = None) -> bool:
    try:
        if debugger_address.startswith('http://'):
            base = debugger_address
        else:
            base = f"http://{debugger_address}"
        encoded = urllib.parse.quote(url, safe='')
        target_url = f"{base}/json/new?{encoded}"
        
        if logger:
            logger(f"调试接口(PUT): {target_url}")
        
        # Try PUT first (Standard CDP)
        try:
            r = requests.put(target_url, timeout=2)
            if r.status_code == 200:
                return True
            if r.status_code != 405:
                if logger: logger(f"PUT失败: {r.status_code}")
        except Exception as e:
            if logger: logger(f"PUT请求异常: {e}")

        # Fallback to GET if PUT failed (Compatibility)
        if logger:
            logger(f"调试接口(GET重试): {target_url}")
        r = requests.get(target_url, timeout=2)
        if logger:
            logger(f"GET响应: {r.status_code}")
        return r.status_code == 200
    except Exception as e:
        if logger:
            logger(f"调试接口最终异常: {e}")
        return False


def open_tab_via_debugger(debugger_address: str, url: str, logger: Optional[Any] = None) -> bool:
    return create_cdp_tab(debugger_address, url, logger)


def proxy_payload(host: str, port: str, username: Optional[str], password: Optional[str], protocol: str = 'socks5') -> Dict[str, Any]:
    if not host or not port:
        return {'proxyType': 'noproxy'}
    
    # Map protocol to bitbrowser proxyType if needed
    # Common types: socks5, http, https
    ptype = protocol.lower() if protocol else 'socks5'
    if ptype not in ('socks5', 'http', 'https', 'ssh'):
        ptype = 'socks5'

    p: Dict[str, Any] = {
        'proxyType': ptype,
        'host': host,
        'port': str(port),
        'proxyHost': host,
        'proxyPort': str(port),
    }
    if username:
        p['proxyUserName'] = username
    if password:
        p['proxyPassword'] = password
    
    # Construct proxy string URL
    if username and password:
        p['proxy'] = f"{ptype}://{username}:{password}@{host}:{port}"
    else:
        p['proxy'] = f"{ptype}://{host}:{port}"
    return p


def log_window_urls(driver: webdriver.Remote, logger: Optional[Any]) -> None:
    if not logger:
        return
    try:
        handles = driver.window_handles
        logger(f"当前窗口句柄数量: {len(handles)}")
        for i, h in enumerate(handles):
            try:
                driver.switch_to.window(h)
                url = driver.current_url
            except Exception:
                url = '(不可获取URL)'
            logger(f"句柄[{i}]: {url}")
    except Exception as e:
        logger(f"枚举窗口失败: {e}")

def log_browser_console(driver: webdriver.Remote, logger: Optional[Any], limit: int = 50) -> None:
    if not logger:
        return
    try:
        logs = driver.get_log('browser')
        if not logs:
            return
        size = len(logs)
        start = max(0, size - limit)
        for entry in logs[start:]:
            try:
                ts = entry.get('timestamp')
                lvl = entry.get('level')
                msg = entry.get('message')
                logger(f"Console[{lvl}]: {msg}")
            except Exception:
                pass
    except Exception:
        pass

def log_performance_network(driver: webdriver.Remote, logger: Optional[Any], limit: int = 200, domain_filter: Optional[str] = None) -> None:
    if not logger:
        return
    def redact(h: Any) -> Any:
        if not isinstance(h, dict):
            return h
        out = {}
        for k, v in h.items():
            try:
                kl = str(k).lower()
            except Exception:
                kl = ''
            if kl in ('authorization', 'cookie', 'set-cookie', 'x-api-key', 'proxy-authorization'):
                out[k] = '[REDACTED]'
            else:
                out[k] = v
        return out
    try:
        logs = driver.get_log('performance')
        if not logs:
            return
        size = len(logs)
        start = max(0, size - limit)
        for entry in logs[start:]:
            try:
                msg_str = entry.get('message') or ''
                data = json.loads(msg_str)
                m = data.get('message', {})
                method = m.get('method')
                params = m.get('params', {})
                if method in ('Network.requestWillBeSent', 'Network.responseReceived'):
                    req = params.get('request', {})
                    res = params.get('response', {})
                    url = req.get('url') or res.get('url') or ''
                    if domain_filter and (domain_filter not in url):
                        continue
                    if method == 'Network.requestWillBeSent':
                        logger(f"HTTP[REQ] {req.get('method')} {url}")
                        hdrs = req.get('headers') or {}
                        try:
                            logger(f"HTTP[REQ-HEADERS] {json.dumps(redact(hdrs), ensure_ascii=False)}")
                        except Exception:
                            pass
                    if method == 'Network.responseReceived':
                        status = res.get('status')
                        logger(f"HTTP[RES] {status} {url}")
                        hdrs = res.get('headers') or {}
                        try:
                            logger(f"HTTP[RES-HEADERS] {json.dumps(redact(hdrs), ensure_ascii=False)}")
                        except Exception:
                            pass
            except Exception:
                pass
    except Exception:
        pass

def log_response_bodies(driver: webdriver.Remote, logger: Optional[Any], limit: int = 10, domain_filter: Optional[str] = None) -> None:
    if not logger:
        return
    try:
        logs = driver.get_log('performance')
        if not logs:
            return
        ids = []
        for entry in reversed(logs):
            try:
                data = json.loads(entry.get('message') or '{}')
                m = data.get('message', {})
                method = m.get('method')
                params = m.get('params', {})
                if method == 'Network.responseReceived':
                    res = params.get('response', {})
                    url = res.get('url') or ''
                    if domain_filter and (domain_filter not in url):
                        continue
                    rid = params.get('requestId')
                    if rid:
                        ids.append((rid, url))
                        if len(ids) >= limit:
                            break
            except Exception:
                pass
        for rid, url in ids:
            try:
                body = driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': rid})
                txt = body.get('body', '')
                logger(f"HTTP[BODY] {url} len={len(txt)}")
            except Exception:
                pass
    except Exception:
        pass


def perform_registration(
    row: Dict[str, Any],
    xpaths: Dict[str, str],
    platform_url: str,
    timeout_ms: int,
    poll_ms: int,
    client: BitBrowserClient,
    logger: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
    target_check: Optional[Callable[[], bool]] = None,
    silent_mode: bool = False,
    udp_enabled: bool = False,
    email_manager: Optional[Any] = None,
    keep_open_on_failure_ms: int = 0,
    allow_hold_on_early_failure: bool = False,
) -> Tuple[bool, str]:
    email = str(row.get('email') or row.get('账号') or '').strip()
    password = str(row.get('password') or row.get('密码') or '').strip()
    code_url = str(row.get('code_url') or row.get('接收验证码链接') or '').strip()
    host = str(row.get('host') or row.get('代理IP') or '').strip()
    port = str(row.get('port') or row.get('端口') or '').strip()
    proxy_username = str(row.get('proxyUserName') or row.get('用户名') or '').strip()
    proxy_password = str(row.get('proxyPassword') or row.get('密码2') or '').strip()
    protocol = str(row.get('protocol') or row.get('proxyType') or 'socks5').strip()
    window_name = str(row.get('windowName') or row.get('窗口名称') or email or 'win').strip()
    
    # Resolve dynamic settings if they are callables
    current_timeout = timeout_ms() if callable(timeout_ms) else timeout_ms
    current_poll = poll_ms() if callable(poll_ms) else poll_ms
    
    if logger:
        logger(f"当前任务参数: Timeout={current_timeout}ms, Poll={current_poll}ms")

    # Early target check
    if target_check and target_check():
        return False, 'target_reached'
        
    if stop_event and stop_event.is_set():
        return False, 'stopped'
    if logger:
        logger(f"create_profile {window_name}")
    browser_id = None
    driver = None
    result_ok = False
    try:
        browser_id = client.update_browser(window_name, proxy_payload(host, port, proxy_username, proxy_password, protocol))
    except Exception:
        # 如果更新失败（如不存在），尝试创建
        try:
            browser_id = client.create_browser(window_name, proxy_payload(host, port, proxy_username, proxy_password, protocol))
        except Exception as create_err:
            if logger:
                logger(f"创建窗口失败: {create_err}")
            return False, f"create_browser failed: {create_err}"
    if logger:
        logger(f"profile_id {browser_id}")
    open_data = client.open_browser(browser_id)
    if logger:
        logger(f"open_browser {browser_id} -> {open_data}")
    if not (open_data.get('driver') and open_data.get('http')):
        raise RuntimeError(f"open_browser 未返回 driver/http: {open_data}")
    
    # Wait for browser to fully start
    time.sleep(5)
    
    try:
        driver = open_attached_driver(open_data)
    except Exception as attach_err:
        if logger:
            logger(f"连接浏览器失败，尝试重试: {attach_err}")
        time.sleep(3)
        driver = open_attached_driver(open_data)
    t_attached = time.time()
    gate_state = 'attached'
    debug_tab_opened = False
    
    # Silent Mode: Move window off-screen instead of minimize for stability
    if silent_mode:
        try:
            driver.set_window_position(-3000, 0)
        except Exception:
            try:
                driver.minimize_window()
            except Exception:
                pass

    result_ok = False
    try:
        if logger:
            logger("步骤: 打开平台网址")
        if not check_connectivity(driver, platform_url, logger, max_wait=20):
            take_screenshot(driver, f"{window_name}_connectivity_failed.png", logger)
            return False, 'proxy_connectivity_failed'
        # 页面就绪与加载时间
        wait_page_ready(driver, 20)
        log_page_timing(driver, logger)
        log_resource_status(driver, logger)
        # DOM 可见性校验，确保关键入口元素已渲染
        try:
            if not element_visible(driver, xpaths.get('language_menu', ''), min(current_timeout, 10000), current_poll):
                if logger:
                    logger("DOM状态校验失败: language_menu 未可见，尝试刷新后重试")
                try:
                    driver.refresh()
                except Exception:
                    pass
                wait_page_ready(driver, 20)
                log_page_timing(driver, logger)
                log_resource_status(driver, logger)
                if not element_visible(driver, xpaths.get('language_menu', ''), min(current_timeout, 10000), current_poll):
                    if logger:
                        logger("DOM状态校验仍失败: language_menu 未可见，继续后续入口以免阻塞")
        except Exception:
            pass
        ready = False
        for xp_key in ('language_menu', 'signin_btn', 'Creative Studio'):
            try:
                xp_val = xpaths.get(xp_key)
                if xp_val and element_exists(driver, xp_val, 8000, current_poll):
                    ready = True
                    break
            except Exception:
                pass
        if not ready:
            try:
                driver.refresh()
            except Exception:
                pass
            wait_page_ready(driver, 20)
            log_page_timing(driver, logger)
            log_resource_status(driver, logger)
            for xp_key in ('language_menu', 'signin_btn', 'Creative Studio'):
                try:
                    xp_val = xpaths.get(xp_key)
                    if xp_val and element_exists(driver, xp_val, 8000, current_poll):
                        ready = True
                        break
                except Exception:
                    pass
        # 为降低早期关闭概率，首页阶段默认不尝试通过调试接口创建新标签
        # 若启用 UDP 调试模式，则允许使用调试接口打开新标签以作为回退
        enable_debugger_open = bool(udp_enabled)
        if enable_debugger_open and (not ready) and open_data.get('http'):
            if not debug_tab_opened:
                prev_handles = driver.window_handles
                open_tab_via_debugger(open_data.get('http') or '', platform_url, logger)
                try:
                    WebDriverWait(driver, 8).until(lambda d: len(d.window_handles) > len(prev_handles))
                    driver.switch_to.window(driver.window_handles[-1])
                except Exception:
                    pass
                debug_tab_opened = True
        # 语言菜单存在性验证与点击
        if logger:
            try:
                cnt = xpath_count(driver, xpaths['language_menu'])
                logger(f"XPath检测 language_menu count={cnt}")
            except Exception:
                pass
        lang_clicked = False
        if safe_click_any(driver, [xpaths.get('language_menu'), xpaths.get('language_menu_alt')], current_timeout, current_poll, logger, retries=2) or js_click_xpath(driver, xpaths.get('language_menu', '')) or js_click_xpath(driver, xpaths.get('language_menu_alt', '')):
            if logger:
                logger("步骤: 打开语言菜单")
            lang_clicked = True
            try:
                WebDriverWait(driver, min(current_timeout, 8000) / 1000.0).until(EC.presence_of_element_located((By.XPATH, xpaths.get('english_option') or xpaths.get('english_option_alt') or "//*[@role='menu']")))
            except Exception:
                pass
            if not safe_click_any(driver, [xpaths.get('english_option'), xpaths.get('english_option_alt')], current_timeout, current_poll, logger, retries=2):
                if not js_click_xpath(driver, xpaths.get('english_option', '')) and not js_click_xpath(driver, xpaths.get('english_option_alt', '')):
                    return False, 'english_option_click_failed'
        else:
            # 未能点击语言菜单，但继续尝试后续入口
            pass
        if logger:
            logger("步骤: 选择英文" + ("(已点击语言菜单)" if lang_clicked else "(跳过语言菜单点击)"))
        if stop_event and stop_event.is_set():
            return False, 'stopped'
        if logger:
            logger("步骤: 点击 Creative Studio")
        if not safe_click_any(driver, [
            xpaths.get('Creative Studio'),
            "//*[contains(text(),'Creative') or contains(text(),'创意') or contains(text(),'工作室')]"
        ], current_timeout, current_poll, logger, retries=2):
            try:
                find_click(driver, xpaths.get('Creative Studio') or "//*[contains(text(),'Creative') or contains(text(),'创意') or contains(text(),'工作室')]", current_timeout, current_poll)
            except Exception as e:
                return False, str(e)
        if logger:
            logger("步骤: 点击 More Tools")
        prev_handles = driver.window_handles
        gate_state = 'pre_open_more_tools'
        if not safe_click_any(driver, [
            xpaths.get('More Tools'),
            "//*[contains(text(),'More') or contains(text(),'更多') or contains(text(),'工具')]"
        ], current_timeout, current_poll, logger, retries=2):
            return False, 'more_tools_click_failed'
        try:
            WebDriverWait(driver, min(8, current_timeout / 1000.0), poll_frequency=current_poll / 1000.0).until(lambda d: len(d.window_handles) > len(prev_handles))
        except Exception:
            pass
        time.sleep(1.5)
        handle_lock = threading.Lock()
        with handle_lock:
            try:
                new_handles = driver.window_handles
                diff = [h for h in new_handles if h not in prev_handles]
                if diff:
                    driver.switch_to.window(diff[-1])
                else:
                    driver.switch_to.window(driver.window_handles[-1])
                gate_state = 'tab_switched'
            except Exception:
                pass
        if logger:
            logger("步骤: 已切换到新标签")
        wait_page_ready(driver, 20)
        try:
            log_browser_console(driver, logger, 20)
        except Exception:
            pass
        try:
            log_performance_network(driver, logger, 100, domain_filter=None)
        except Exception:
            pass
        try:
            log_resource_status(driver, logger)
        except Exception:
            pass
        time.sleep(0.5)
        if logger:
            logger("步骤: 点击 Sign In")
        if not safe_click_any(driver, [
            xpaths.get('signin_btn'),
            "//*[contains(text(),'Sign In') or contains(text(),'登录')]"
        ], current_timeout, current_poll, logger, retries=2):
            return False, 'signin_click_failed'
        try:
            WebDriverWait(driver, min(current_timeout, 10000) / 1000.0).until(EC.presence_of_element_located((By.XPATH, xpaths.get('signin_with_email') or xpaths.get('signin_with_email_alt') or "//*")))
        except Exception:
            pass
        if logger:
            logger("步骤: 选择邮箱登录")
        if not safe_click_any(driver, [
            xpaths.get('signin_with_email'),
            "//*[contains(text(),'邮箱') or contains(text(),'email') or contains(text(),'邮件')]"
        ], current_timeout, current_poll, logger, retries=2):
            return False, 'signin_email_click_failed'
        try:
            WebDriverWait(driver, min(current_timeout, 10000) / 1000.0).until(EC.presence_of_element_located((By.XPATH, xpaths.get('Sign up for free') or "//*[contains(text(),'免费') or contains(text(),'注册') or contains(text(),'Sign up')]")))
        except Exception:
            pass
        if logger:
            logger("步骤: 点击免费注册")
        if not safe_click_any(driver, [
            xpaths.get('Sign up for free'),
            "//*[contains(text(),'免费') or contains(text(),'注册') or contains(text(),'Sign up')]"
        ], current_timeout, current_poll, logger, retries=2):
            return False, 'signup_click_failed'
        try:
            WebDriverWait(driver, min(current_timeout, 12000) / 1000.0).until(EC.presence_of_element_located((By.XPATH, xpaths.get('Enter Email Address') or "//*[@placeholder][contains(@placeholder,'邮箱') or contains(@placeholder,'Email')]")))
        except Exception:
            pass
        if logger:
            logger("步骤: 输入邮箱")
        if not safe_send_keys_any(driver, [
            xpaths.get('Enter Email Address'),
            "//*[@placeholder][contains(@placeholder,'邮箱') or contains(@placeholder,'Email')]"
        ], email, current_timeout, current_poll, logger, retries=1):
            return False, 'email_input_failed'
        if logger:
            logger("步骤: 输入密码")
        if not safe_send_keys_any(driver, [
            xpaths.get('password_input'),
            "//*[@placeholder][contains(@placeholder,'密码') or contains(@placeholder,'Password')]"
        ], password, current_timeout, current_poll, logger, retries=1):
            return False, 'password_input_failed'
        if logger:
            logger("步骤: 确认密码")
        if not safe_send_keys_any(driver, [
            xpaths.get('Confirm Password'),
            "//*[@placeholder][contains(@placeholder,'确认') or contains(@placeholder,'Confirm')]"
        ], password, current_timeout, current_poll, logger, retries=1):
            return False, 'confirm_input_failed'
        if logger:
            logger("步骤: 点击下一步")
        if not safe_click_any(driver, [
            xpaths.get('next_btn'),
            "//*[contains(text(),'下一步') or contains(text(),'Next')]"
        ], current_timeout, current_poll, logger, retries=2):
            return False, 'next_click_failed'
        gate_state = 'registration_started'

        try:
            # Check for generic error message
            if element_visible(driver, "//*[contains(text(), 'registered') or contains(text(), 'used')]", 2000, current_poll):
                body_text = driver.find_element(By.TAG_NAME, 'body').text.lower()
                if "already registered" in body_text or "already used" in body_text or "account exists" in body_text:
                    if logger: logger("检测到邮箱已被使用提示")
                    if email_manager:
                        email_manager.update_email_status(email, 'fail_used')
                    return False, 'email_used_prompt'
        except Exception:
            pass

        if logger:
            logger("步骤: 等待并通过滑块")
        if stop_event and stop_event.is_set():
            return False, 'stopped'
        code_input_el_xpath = xpaths.get('code_url_element')
        if not code_input_el_xpath:
            return False, 'code_input_xpath_missing'
        t_slider = time.time()
        slider_ok = False
        max_slider_retries = 8
        slider_iframe_xpath = xpaths.get('slider_iframe')
        slider_container_xpath = xpaths.get('slider_container')
        for attempt in range(max_slider_retries):
            if stop_event and stop_event.is_set():
                return False, 'stopped'
            if attempt > 0 and logger:
                logger(f"Slider: 重试 {attempt+1}/{max_slider_retries}")
            if solve_slider(driver, xpaths, current_timeout, current_poll, logger=logger):
                slider_ok = True
                break
            if attempt < max_slider_retries - 1:
                try:
                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass
                    t_wait = time.time()
                    while time.time() - t_wait < 6:
                        if stop_event and stop_event.is_set():
                            return False, 'stopped'
                        if element_visible(driver, code_input_el_xpath, 400, current_poll):
                            break
                        if (slider_iframe_xpath and element_visible(driver, slider_iframe_xpath, 400, current_poll)) or (slider_container_xpath and element_visible(driver, slider_container_xpath, 400, current_poll)):
                            break
                        time.sleep(0.3)
                except Exception:
                    pass
        if not slider_ok:
            return False, 'slider_failed'
        if logger:
            logger(f"步骤: 滑块通过 (耗时 {time.time()-t_slider:.2f}s)")
            try:
                log_resource_phase_timings(driver, logger, contains="captcha", limit=30)
            except Exception:
                pass
        # 优化: 滑块通过后，直接跳转接码页等待验证码
        if logger:
            logger("步骤: 滑块通过，立即跳转接码页等待验证码")
        
        # 在等待过程中动态检查停止与目标达成
        if stop_event and stop_event.is_set():
            return False, 'stopped'
        if target_check and target_check():
            if logger: logger("终止接码：已达到目标注册数量")
            return False, 'target_reached'

        # 直接调用接码流程 (内部负责打开、刷新、提取)
        code = extract_verification_code_flow(driver, code_url, xpaths.get('code_input'), logger, open_data.get('http'), timeout=60)
        
        if not code:
            if logger:
                logger("步骤: 获取验证码失败(超时)")
            # 诊断: 检查是否滑块重现(导致根本没发码)
            try:
                if (slider_iframe_xpath and element_visible(driver, slider_iframe_xpath, 1000, current_poll)) or \
                   (slider_container_xpath and element_visible(driver, slider_container_xpath, 1000, current_poll)):
                    if logger: logger("诊断: 滑块验证框重新出现，判定为滑块实际上未通过")
                    return False, 'slider_reappeared'
            except Exception:
                pass
            return False, 'code_not_found'

        gate_state = 'code_received'
        if logger:
            logger("步骤: 获取验证码成功，已切回主窗口")

        # 3. 拿到验证码后，等待输入框出现并填入
        try:
            t0 = time.time()
            if logger: logger("正在等待验证码输入框...")
            # 因为之前已经等待了验证码提取时间，这里通常应该已经就绪，但为了稳健仍给予等待
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, code_input_el_xpath))
            )
            if logger: logger(f"验证码输入框就绪 (耗时 {time.time()-t0:.2f}s)")
        except Exception:
            if logger: logger("验证码输入框未出现(超时)")
            # 再次诊断
            try:
                body_text = driver.find_element(By.TAG_NAME, 'body').text
                if "frequent" in body_text.lower() or "try again" in body_text.lower():
                     if logger: logger(f"诊断: 页面包含错误提示")
            except Exception:
                pass
            return False, 'code_input_not_visible'
        
        # 3. Input code
        if target_check and target_check():
            if logger: logger("终止提交：已达到目标注册数量")
            return False, 'target_reached'
        WebDriverWait(driver, min(current_timeout, 10000) / 1000.0).until(
            EC.visibility_of_element_located((By.XPATH, code_input_el_xpath))
        ).send_keys(code)
        if logger:
            logger("步骤: 填写验证码")
        
        # Secondary email check before final submit
        if email_manager and email_manager.is_email_registered(email):
             if logger: logger(f"提交前检测: 邮箱 {email} 已被注册")
             return False, 'email_already_registered_secondary'

        if target_check and target_check():
            if logger: logger("终止提交：已达到目标注册数量")
            return False, 'target_reached'
        find_click_any(driver, xpaths['final_submit_btn'], current_timeout, current_poll)
        try:
            log_performance_network(driver, logger, 100, domain_filter=None)
        except Exception:
            pass
        try:
            log_response_bodies(driver, logger, 5, domain_filter=None)
        except Exception:
            pass
        
        if logger:
            logger("步骤: 提交注册")
        # Ensure we find the close popup
        # Optimize: Retry loop for closing popup to ensure it is closed
        popup_closed = False
        
        # Define locators: Config > User Preference > Fallback
        popup_xpaths = [
            xpaths.get('close_popup_svg'), # From config
            "//svg[@xmlns='http://www.w3.org/2000/svg' and contains(@class, 'el-icon')]",
            "//*[contains(@class, 'el-dialog__close')]"
        ]
        # Filter None
        popup_xpaths = [x for x in popup_xpaths if x]
        
        for attempt in range(3):
            try:
                if logger and attempt == 0: logger("检查关闭弹窗状态...")
                
                # Check if popup exists AND is visible
                found_xpath = None
                # Use a shorter timeout for checking existence (e.g., 2s)
                # If it's not there, we assume it's closed or never appeared.
                for xp in popup_xpaths:
                    # Fix: Use element_visible instead of element_exists
                    # This ensures we don't try to close a popup that is in DOM but hidden (display: none)
                    if element_visible(driver, xp, 2000, current_poll):
                         found_xpath = xp
                         break
                
                if not found_xpath:
                    if logger: logger("未检测到可见弹窗，视为已关闭")
                    popup_closed = True
                    break
                
                if logger: logger(f"检测到关闭弹窗 ({found_xpath})，尝试关闭 (第 {attempt+1} 次)")
                
                # Try explicit wait for clickable
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, found_xpath)))
                except Exception:
                    pass

                # Try normal click first
                clicked = find_click_any(driver, found_xpath, 2000, current_poll)
                if not clicked:
                    # Try JS click
                    js_click_xpath(driver, found_xpath)
                
                time.sleep(1) # Wait for animation
                
                # Verify if closed
                # Fix: Check for INVISIBILITY instead of non-existence
                # Element UI often hides dialogs (display: none) rather than removing them from DOM
                try:
                    WebDriverWait(driver, 2, poll_frequency=0.2).until(
                        EC.invisibility_of_element_located((By.XPATH, found_xpath))
                    )
                    if logger: logger("弹窗已成功关闭 (验证通过)")
                    popup_closed = True
                    break
                except Exception:
                    if logger: logger("弹窗关闭验证失败(仍可见)，准备重试...")
                    
            except Exception as e:
                if logger: logger(f"关闭弹窗过程异常: {e}")
            
            time.sleep(1)

        result_ok = True # Default to True as per "submit clicked" logic
        
        if popup_closed:
             # Strict Email Status Update: Only mark as submitted AFTER successful popup close
            if email_manager:
                try:
                    email_manager.update_email_status(email, 'submitted')
                    if logger: logger(f"邮箱 {email} 已标记为 submitted (注册流程完成)")
                except Exception as e:
                    if logger: logger(f"标记邮箱状态失败: {e}")
            
            return True, 'success'
        else:
            if logger:
                logger("步骤: 未找到关闭弹窗按钮或关闭超时，但提交按钮已点击，视为成功")
            
            # 即使未找到弹窗，也标记为成功，因为"final_submit_btn"已经点击
            # User Request: "Correct to: Fill Email -> Click Submit -> Mark on Success"
            # Since we treat "Submit Clicked" as Success (even if popup fails), we mark it here.
            if email_manager:
                try:
                    email_manager.update_email_status(email, 'submitted')
                    if logger: logger(f"邮箱 {email} 已标记为 submitted (强制成功)")
                except Exception as e:
                    if logger: logger(f"标记邮箱状态失败: {e}")
            
            return True, '成功(弹窗未关闭)'

    except Exception as e:
        result_ok = False # Mark as failed for exception
        return False, str(e)
    finally:
        # Rollback Mechanism: If not successful, reset email status if needed
        if not result_ok and email_manager:
            try:
                # If failure, and we haven't determined it's a permanent failure (like email used), reset.
                # If msg is 'email_used_prompt' or 'email_already_registered', we should NOT reset.
                # result_ok is False here.
                # But 'msg' variable is local to try block? No, it's not in scope of finally if exception raised before definition.
                # We can check the actual status in email_manager.
                
                # However, perform_registration doesn't update 'fail_used' itself usually, it returns msg.
                # The caller (run_batch) handles status updates based on return value.
                # EXCEPT for the 'submitted' status we just added inside the function.
                
                # So if we marked 'submitted' but then crashed/failed (unlikely given the flow above), we should rollback.
                # Or if we marked nothing, we ensure it's clean.
                
                # Check if email is currently marked as 'submitted' or 'success' but we are failing?
                # If we are failing, we shouldn't have marked it 'submitted' yet, unless exception occurred AFTER marking.
                
                # But the user request says: "If registration flow fails at ANY stage, auto reset email status to unused".
                # This implies if it was 'new' and we failed, keep it 'new'.
                # If we marked it 'used' prematurely, reset it.
                # Since we strictly moved marking to the end, the only risk is if we marked it and then failed in finally block?
                # Or if previous steps marked it? (Previous steps don't mark it).
                
                # So this is mostly a safety net.
                pass
            except Exception:
                pass

        if logger:
             logger(f"正在清理资源: {browser_id}")
        
        # 诊断: 检查窗口是否在清理前仍然存活
        try:
            if browser_id and logger:
                detail = client.detail_browser(browser_id)
                logger(f"清理前浏览器状态: {detail}")
        except Exception:
            pass

        elapsed = time.time() - t_attached if 't_attached' in locals() else 0
        early_failure = (not result_ok) and (gate_state in ('attached', 'pre_open_more_tools', 'tab_switched'))
        hold_close = (not result_ok) and ((elapsed < (keep_open_on_failure_ms / 1000.0)) or (allow_hold_on_early_failure and early_failure))
        if hold_close and logger:
            if allow_hold_on_early_failure and early_failure:
                logger("因早期失败，保持窗口并等待后续重试")
            else:
                logger(f"因失败且未达到最小保持时间({keep_open_on_failure_ms}ms)，暂不关闭窗口")
        
        # 优先使用 driver.quit() 关闭窗口，这通常比 API 更干净
        try:
            if driver and not hold_close:
                # 检查连接是否还在，如果还在则退出
                try:
                    driver.quit()
                    if logger: logger("driver.quit() 执行完成")
                except Exception:
                    pass
        except Exception as e:
            if logger: logger(f"driver.quit error: {e}")

        try:
            if browser_id and not hold_close:
                if result_ok:
                    if logger: logger("注册成功，等待数据同步...")
                    time.sleep(3)
                client.close_browser(browser_id)
                if logger: logger(f"已请求关闭浏览器API: {browser_id}")
        except Exception as e:
            if logger: logger(f"close_browser error: {e}")

        # 必须删除未成功的窗口
        if browser_id and not result_ok:
            if logger:
                logger(f"任务失败，正在删除窗口: {browser_id}")
            # time.sleep(2)
            for _ in range(3):
                try:
                    client.delete_browser(browser_id)
                    if logger:
                        logger(f"已删除失败窗口: {browser_id}")
                    break
                except Exception as del_err:
                    if logger:
                        logger(f"删除窗口失败(重试): {del_err}")
                    time.sleep(2.0)


def run_batch(
    input_path: str,
    xpaths_path: str,
    platform_url: str,
    base_url: str,
    secret: Optional[str],
    concurrency: int,
    timeout_ms: Any, # int or Callable
    poll_ms: Any, # int or Callable
    logger: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
    progress_cb: Optional[Any] = None,
    ip_manager: Optional[Any] = None,
    exhaustion_cb: Optional[Any] = None,
    target_success_count: int = 99999999,
    email_manager: Optional[Any] = None,
    silent_mode: bool = False,
    udp_enabled: bool = False,
) -> None:
    rows = read_rows(input_path)
    if not rows:
        if logger:
            logger("未发现可处理的任务行，直接退出")
        return
    if stop_event is None:
        stop_event = threading.Event()
    with open(xpaths_path, 'r', encoding='utf-8') as f:
        xpaths = json.load(f)
    client = BitBrowserClient(base_url, secret)
    def _ping(url: str) -> bool:
        try:
            # 优化: 使用 list 接口代替 update/create 来进行健康检查，避免产生残留窗口
            payload = {'page': 0, 'pageSize': 1}
            h = client._headers()
            r = requests.post(f"{url.rstrip('/')}/browser/list", headers=h, data=json.dumps(payload), timeout=5)
            if r.status_code == 200:
                try:
                    r.json()
                    return True
                except:
                    pass
            return False
        except Exception as e:
            return False
    if logger:
        logger("开始进行比特浏览器接口连通性检查 (health-check)")
    if not _ping(client.base_url):
        candidates = [client.base_url] + [f"http://127.0.0.1:{p}" for p in (54345, 54346, 54321, 54322, 50325, 55555)]
        for c in candidates:
            if _ping(c):
                client.base_url = c.rstrip('/')
                if logger:
                    logger(f"已自动切换比特浏览器接口到: {client.base_url}")
                break
        else:
            if logger:
                logger("比特浏览器接口不可用，请确认应用已启动并开启本地API")
            return
    
    lock = threading.Lock()
    global_success_count = 0

    # 资源预检查: 估算当前 IP 池和邮箱池最大可支持的注册数量
    # 说明: 这里只做告警与日志输出，不改变原有业务决策逻辑
    try:
        ip_stats = ip_manager.get_stats() if ip_manager else None
    except Exception:
        ip_stats = None
    try:
        email_total = len(rows)
    except Exception:
        email_total = 0
    if logger:
        if ip_stats:
            logger(f"资源预检查: IP池总数={ip_stats['total_ips']}, 剩余可用名额={ip_stats['remaining_usage_count']}, 邮箱任务数={email_total}, 目标成功数={target_success_count}")
        else:
            logger(f"资源预检查: 无IP池(或读取失败)，邮箱任务数={email_total}, 目标成功数={target_success_count}")
    
    def task(idx: int, r: Dict[str, Any]) -> Tuple[int, bool, str]:
        nonlocal global_success_count
        email = str(r.get('email') or r.get('账号') or '').strip()
        if logger: logger(f"任务启动 #{idx+1}: {email}")
        
        if stop_event and stop_event.is_set():
            return idx, False, 'stopped'
            
        with lock:
            if global_success_count >= target_success_count:
                if logger: logger(f"任务 #{idx+1} 终止: 已达到目标 ({global_success_count}/{target_success_count})")
                return idx, False, 'target_reached'
        
        email = str(r.get('email') or r.get('账号') or '').strip()
        
        # Check if email is already used in EmailManager (source of truth)
        if email_manager and email_manager.is_email_registered(email):
             if logger: logger(f"跳过 {email}: 邮箱状态已标记为已使用/成功")
             r['status'] = 'surplus' 
             return idx, False, 'email_already_used_persistent'
        
        # Mark as processing
        if email_manager:
            email_manager.update_email_status(email, 'processing')
        
        # IP Allocation Logic
        if ip_manager:
            while True:
                if stop_event and stop_event.is_set():
                    if email_manager:
                        email_manager.update_email_status(email, 'stopped')
                    return idx, False, 'stopped'
                with lock:
                    if global_success_count >= target_success_count:
                        if email_manager:
                            email_manager.update_email_status(email, 'skipped')
                        return idx, False, 'target_reached'
                
                ip_entry, status = ip_manager.allocate_ip(email)
                if status == 'success' and ip_entry:
                    # Inject IP into row
                    r['host'] = ip_entry['host']
                    r['port'] = ip_entry['port']
                    r['proxyUserName'] = ip_entry.get('proxyUserName', '')
                    r['proxyPassword'] = ip_entry.get('proxyPassword', '')
                    r['protocol'] = ip_entry.get('protocol', 'socks5')
                    if logger:
                        logger(f"已分配IP: {r['host']}:{r['port']} ({r['protocol']}) 给 {email}")
                    break
                elif status == 'email_used':
                    if logger:
                        logger(f"跳过 {email}: 此邮箱已在IP池中使用过")
                    if email_manager:
                        email_manager.update_email_status(email, 'skipped_ip_used')
                    return idx, False, 'email_already_used_in_pool'
                elif status == 'ip_busy':
                    if logger:
                        logger(f"IP资源繁忙 (并发限制)，等待重试: {email}")
                    time.sleep(2)
                    continue
                elif status == 'ip_exhausted':
                    if exhaustion_cb:
                        if logger:
                            logger("IP池耗尽，等待用户处理...")
                        action = exhaustion_cb('ip_exhausted', ip_manager.get_stats())
                        if action == 'retry':
                            if logger:
                                logger("用户选择重试分配IP")
                            continue
                        else:
                            if logger:
                                logger("用户取消任务")
                            if email_manager:
                                email_manager.update_email_status(email, 'cancelled')
                            return idx, False, 'ip_exhausted_cancelled'
                    else:
                        if email_manager:
                            email_manager.update_email_status(email, 'failed_ip_exhausted')
                        return idx, False, 'ip_exhausted_no_handler'
                else:
                    if email_manager:
                        email_manager.update_email_status(email, f'failed_ip_error_{status}')
                    return idx, False, f'ip_allocation_error: {status}'

        try:
            attempts = 0
            ok = False
            msg = ''
            while attempts < 3 and not ok:
                attempts += 1
                ok, msg = perform_registration(
                    r,
                    xpaths,
                    platform_url,
                    timeout_ms,
                    poll_ms,
                    client,
                    logger,
                    stop_event,
                    target_check=lambda: global_success_count >= target_success_count,
                    silent_mode=silent_mode,
                    udp_enabled=udp_enabled,
                    email_manager=email_manager,
                    keep_open_on_failure_ms=0,
                    allow_hold_on_early_failure=False,
                )
                if not ok:
                    if logger:
                        logger(f"重试 {attempts}/3: {email} 失败原因: {msg}")
                    if msg and ('proxy' in msg or 'network' in msg):
                        time.sleep(3)
                        continue
                    else:
                        break
            
            if ip_manager and not ok:
                # Release IP on failure
                 ip_manager.release_ip(r.get('host'), r.get('port'), email)
                 if logger:
                     logger(f"注册失败，已释放IP资源: {email}")
                     
            with lock:
                if ok:
                    global_success_count += 1
                    if global_success_count > target_success_count:
                        r['status'] = 'surplus'
                        msg = 'Excess registration (target exceeded)'
                        if logger: logger(f"⚠️ 超量注册检测: {email} (当前成功数: {global_success_count}, 目标: {target_success_count})")
                        if email_manager:
                            email_manager.update_email_status(email, 'surplus')
                    else:
                        r['status'] = 'good'
                        if email_manager:
                            email_manager.update_email_status(email, 'success')

                    if global_success_count >= target_success_count:
                        if not stop_event.is_set():
                            stop_event.set()
                            if logger: logger(f"🛑 终止条件触发: 任务完成目标 ({target_success_count})，已触发停止信号")
                else:
                    r['status'] = 'fail'
                    if email_manager:
                        if msg == 'email_already_registered_local':
                            pass
                        elif msg == 'email_used_prompt':
                            email_manager.update_email_status(email, 'fail_used')
                        else:
                            # User requested to reset to unused on failure
                            email_manager.update_email_status(email, 'new')
        finally:
             if ip_manager and 'ip_entry' in locals() and ip_entry:
                 ip_manager.release_active_ip(ip_entry['host'], ip_entry['port'])
                
        row['msg'] = msg
        return idx, ok, msg

    pending_idx = [i for i, r in enumerate(rows) if str(r.get('status', '')).strip() != 'good']
    rounds = 0
    
    # Resolve initial max_rounds to avoid reference error if loop doesn't run (though unlikely here)
    # But mainly for the loop logic.
    
    # Outer Loop: Rounds (Passes through the list)
    while True:
        # --- 优先级 1: 目标数量限制 ---
        with lock:
            if global_success_count >= target_success_count:
                if logger: logger(f"🛑 终止条件触发: 已达到目标注册数量 ({target_success_count})")
                break

        # --- 优先级 2: 待处理任务 ---
        pending_idx = [i for i, r in enumerate(rows) if str(r.get('status', '')).strip() != 'good']
        if not pending_idx:
            if logger: logger("🛑 终止条件触发: 所有任务已完成 (无待处理项)")
            break
            
        if stop_event and stop_event.is_set():
            if logger: logger("🛑 终止条件触发: 收到外部停止信号")
            break
            
        if logger:
            logger(f"🚀 开始第 {rounds + 1} 轮 (持续运行直至达到目标)")
            logger(f"📊 当前状态: 成功 {global_success_count}/{target_success_count}, 待处理 {len(pending_idx)}")
            
        # Progress callback start of round
        if progress_cb and callable(progress_cb):
             total = len(rows)
             succ = sum(1 for rr in rows if str(rr.get('status', '')).strip() == 'good')
             fail = sum(1 for rr in rows if str(rr.get('status', '')).strip() == 'fail')
             progress_cb(total, succ, fail, rounds + 1, 0)

        # Check target count before round processing (Double Check)
        with lock:
            remaining_needed = target_success_count - global_success_count
            if remaining_needed <= 0:
                break

        # Inner Loop: Process Pending Tasks in Batches
        # We process 'pending_idx' in chunks of 'concurrency'
        round_tasks_indices = pending_idx[:]
        chunk_size = concurrency
        
        # Iterate chunks
        for i in range(0, len(round_tasks_indices), chunk_size):
            if stop_event and stop_event.is_set():
                break
            
            # Check Global Target (Pre-batch check)
            with lock:
                if global_success_count >= target_success_count:
                    if logger: logger(f"🛑 批次前检查: 已达到目标 ({global_success_count}/{target_success_count})")
                    break
                remaining_needed = target_success_count - global_success_count
            
            # Determine batch
            batch_indices = round_tasks_indices[i : i + chunk_size]
            # Cap batch size by remaining needed to avoid over-execution
            current_batch_limit = min(len(batch_indices), remaining_needed)
            actual_batch_indices = batch_indices[:current_batch_limit]
            
            if not actual_batch_indices:
                break
            
            # Submit Batch
            futures = []
            if logger: logger(f"Process Batch: {len(actual_batch_indices)} tasks")
            with ThreadPoolExecutor(max_workers=len(actual_batch_indices)) as ex:
                for idx in actual_batch_indices:
                    futures.append(ex.submit(task, idx, rows[idx]))
                
                for fut in as_completed(futures):
                    try:
                        idx_res, ok, msg = fut.result()
                        # Log result
                        
                        if progress_cb and callable(progress_cb):
                             total = len(rows)
                             succ = sum(1 for rr in rows if str(rr.get('status', '')).strip() == 'good')
                             fail = sum(1 for rr in rows if str(rr.get('status', '')).strip() == 'fail')
                             progress_cb(total, succ, fail, rounds + 1, current_max_rounds)

                        with lock:
                             if global_success_count >= target_success_count:
                                 if not stop_event.is_set():
                                     stop_event.set()
                                     if logger: logger(f"🛑 终止条件触发: 已达到目标注册数量 ({target_success_count})")
                    except Exception:
                        pass
            
            if stop_event.is_set():
                break

        write_rows_csv(input_path, rows)
        rounds += 1
        
        if logger: logger(f"🏁 第 {rounds} 轮结束")
        
        # Check if we should continue to next round
        # The loop condition at the top will handle max_rounds and target checks.
        # We just need to break if stop_event is set.
        if stop_event.is_set():
            break
            
        time.sleep(1) # Brief pause between rounds


def main() -> None:
    import argparse
    p = argparse.ArgumentParser()
    import os
    base_dir = os.path.dirname(__file__)
    p.add_argument('--input', default=os.path.join(base_dir, 'kl-mail.csv'))
    p.add_argument('--xpaths', default=os.path.join(base_dir, 'kling_xpaths.json'))
    p.add_argument('--platform-url', default='https://klingai.com/global')
    p.add_argument('--bitbrowser-url', default='http://127.0.0.1:54345')
    p.add_argument('--bitbrowser-secret', default=os.environ.get('BITBROWSER_SECRET'))
    p.add_argument('--concurrency', type=int, default=1)
    p.add_argument('--timeout-ms', type=int, default=100000)
    p.add_argument('--poll-ms', type=int, default=500)
    # p.add_argument('--max-rounds', type=int, default=5) # Deprecated
    args = p.parse_args()
    def stdout_logger(msg: str) -> None:
        print(msg)
    run_batch(args.input, args.xpaths, args.platform_url, args.bitbrowser_url, args.bitbrowser_secret, args.concurrency, args.timeout_ms, args.poll_ms, logger=stdout_logger)


if __name__ == '__main__':
    main()
